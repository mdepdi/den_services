import os
import time
import geopandas as gpd
import pandas as pd
import numpy as np
import sys
import shutil
import math
import re
from pathlib import Path
from tqdm import tqdm
from enum import Enum
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side, Font, Alignment

root = Path(__file__).resolve().parents[2]
sys.path.append(root)

from modules.kml import validate_kmz_design, validate_kmz_ipl
from modules.utils import admin_information
from modules.geometry import geodesic_length
from core.config import settings

MAINDATA_DIR = settings.MAINDATA_DIR
DATA_DIR = settings.DATA_DIR

# ------------------------------------------------------
# LOGGER
# ------------------------------------------------------
from core.logger import create_logger
logger = create_logger(__file__)

# ENUMS
# ------------------------------------------------------
class Operator(str, Enum):
    IOH = "ioh"
    XL = "xl"
    SURGE = "surge"
    TSEL = "tsel"

class DeviceType(str, Enum):
    OTB = "OTB"
    ODP = "ODP"

class ConnectorType(str, Enum):
    SC = "SC"
    FC = "FC"

class BoQType(str, Enum):
    INTERSITE = "intersite"
    MMP = "mmp"

def drm_format(
    kmz_path: str,
    export_dir: str,
    sep: str = ";",
):
        
    validated = validate_kmz_design(kmz_path, sep=sep)
    if validated is None:
        logger.info(f"❌ Invalid KMZ Design: {kmz_path}")
        return

    # ---------------------------
    # Inputs (GeoDataFrames)
    # ---------------------------
    points_kmz, lines_kmz = validated

    # Check Admin Level
    admin_col = {'Provinsi','Kabkot', 'Kecamatan', 'Desa'}
    miss_col = admin_col - set(points_kmz.columns)

    if miss_col:
        logger.info(f"🌏 Add Admin Information (missing: {(", ").join(sorted(miss_col))})")
        points_kmz = admin_information(points_kmz, level="desa")
        lines_kmz = admin_information(lines_kmz, level="kabkot")
        points_kmz = points_kmz.drop_duplicates(subset=["site_id", "ring_name"]).reset_index(drop=True)
        lines_kmz = lines_kmz.drop_duplicates(subset=["segment", "ring_name"]).reset_index(drop=True)

    lines_kmz['length'] = lines_kmz['geometry'].to_crs(epsg=4326).apply(geodesic_length)

    colopriming_data = pd.read_excel(f"{DATA_DIR}/Sitelist Dec 2025.xlsx")
    colopriming_data['site_id'] = colopriming_data['site_id'].astype(str)
    colopriming_data = colopriming_data.set_index("site_id")

    target_crs = 3857
    points_kmz = points_kmz.to_crs(epsg=target_crs)
    lines_kmz = lines_kmz.to_crs(epsg=target_crs)

    points_grouped = points_kmz.groupby("ring_name")

    # Container
    segments_records = []
    sites_records = []
    ring_records = []
    recorded_sites = set()
    for ring, ring_points in tqdm(points_grouped, desc="DRM Format Process Ring", total=len(points_grouped)):
        ring_lines = lines_kmz[lines_kmz['ring_name'] == ring].copy()

        if ring_lines.empty:
            logger.error(f"🔴 Ring {ring} lines data not found.")
            continue

        # Ring Metadata
        province = ring_points['Provinsi'].mode()[0]
        city = ring_points['Kabkot'].mode()[0]
        vendor = "TBG"

        ne_ids = set(ring_lines['near_end'].astype(str))
        fe_ids = set(ring_lines['far_end'].astype(str))

        # Route
        routes_length = ring_lines['length'].sum()

        # Hub Sitelist
        hubs = ring_points[ring_points['site_type'].astype(str).str.lower().str.contains("hub")].copy()
        sitelist = ring_points[~(ring_points.index.isin(hubs.index))].copy()
        hubs_ids = set(hubs['site_id'].astype(str))
        sites_ids = set(sitelist['site_id'].astype(str))
        qty_hubs = len(hubs)
        qty_sites = len(sitelist)

        match qty_hubs:
            case 0:
                logger.error(f"🔴 Ring {ring} FO Hub not found.")
                continue
            case 1:
                hub_1 = hubs['site_id'].astype(str).values[0]
                hub_2 = None
            case 2:
                hub_1 = hubs['site_id'].astype(str).values[0]
                hub_2 = hubs['site_id'].astype(str).values[-1]
            case _:
                hub_1 = hubs['site_id'].astype(str).values[0]
                hub_2 = None
                logger.error(f"🟠 Ring {ring} FO Hub exceeds, found {qty_hubs}.")
            
        # Enrich Metadata
        route_type = None
        if qty_hubs == 1 and qty_sites == 1:
            route_type = "Star"
        elif qty_hubs == 1 and qty_sites > 1:
            route_type = "Chain"
        elif qty_hubs > 1 and qty_sites >= 1:
            route_type = "Ring"

        # Segments
        for idx, route in ring_lines.iterrows():
            ne_id = route['near_end']
            fe_id = route['far_end']

            ne_site = ring_points[ring_points['site_id'].astype(str) == str(ne_id)].copy()
            fe_site = ring_points[ring_points['site_id'].astype(str) == str(fe_id)].copy()
            
            if ne_site.empty:
                raise ValueError(f"🔴 Near End site {ne_id} not found in ring {ring}.")

            if fe_site.empty:
                raise ValueError(f"🔴 Far End site {fe_id} not found in ring {ring}.")

            ne_site = ne_site.iloc[0]
            fe_site = fe_site.iloc[0]

            # Enrich Metadata
            match route_type:
                case "Star" | "Chain":
                    ne_name = str(ne_site.get("site_name", ""))
                    fe_name = str(fe_site.get("site_name", ""))

                    ne_station = bool(re.fullmatch(r'^[A-Za-z ]+', ne_name))
                    fe_station = bool(re.fullmatch(r'^[A-Za-z ]+', fe_name))

                    if ne_station or fe_station:
                        ne_status = "Station" if ne_station else "Direct to Station"
                        fe_status = "Station" if fe_station else "Direct to Station"
                    else:
                        ne_status = ne_site.get('site_type')
                        fe_status = fe_site.get('site_type')

                case "Ring":
                    ne_status = ne_site.get('site_type')
                    fe_status = fe_site.get('site_type')
            
            segment = {
                "no": None,
                "segment": route['segment'],
                "ne_site": ne_site['site_id'],
                "ne_long": ne_site['long'],
                "ne_lat": ne_site['lat'],
                "fe_site": fe_site['site_id'],
                "fe_long": fe_site['long'],
                "fe_lat": fe_site['lat'],
                "status_ne": ne_status,
                "status_fe": fe_status,
                "ring_name": ring,
                "route_type": route_type,
                "length": route['length'],
                "hub_1": hub_1,
                "hub_2": hub_2,
                "pop_type": None,
                "rfs_plan":None,
                "province": province,
                "city": city,
                "vendor": vendor
            }
            segments_records.append(segment)
        
            # Site
            colo_ne = colopriming_data.loc[ne_site['site_id']].copy() if ne_site['site_id'] in colopriming_data.index else pd.DataFrame()
            colo_fe = colopriming_data.loc[fe_site['site_id']].copy() if fe_site['site_id'] in colopriming_data.index else pd.DataFrame()
            ne_record = {
                "no": None,
                "site_id": ne_site['site_id'],
                "site_name": ne_site['site_name'],
                "site_id_cust": None,
                "province": ne_site['Provinsi'],
                "city": ne_site['Kabkot'],
                "kecamatan": ne_site['Kecamatan'],
                "kelurahan": ne_site['Desa'],
                "lat": ne_site['lat'],
                "long": ne_site['long'],
                "tower_owner": colo_ne.get("tp", None),
                "field_type": colo_ne.get("field_type", None),
                "pop_interconnection": (",").join(hubs['site_id'].unique().tolist()),
                "location_pop": None,
                "antenna_height": colo_ne.get("total_tower_height", None),
                "total_sector": None,
                "fo_type": route_type,
                "ring_id": ring,
                "distance_per_site": route['length'] if str(ne_site['site_id']) not in hubs_ids else None
            }
            fe_record = {
                "no": None,
                "site_id": fe_site['site_id'],
                "site_name": fe_site['site_name'],
                "site_id_cust": None,
                "province": fe_site['Provinsi'],
                "city": fe_site['Kabkot'],
                "kecamatan": fe_site['Kecamatan'],
                "kelurahan": fe_site['Desa'],
                "lat": fe_site['lat'],
                "long": fe_site['long'],
                "tower_owner": colo_fe.get("tp", None),
                "field_type": colo_fe.get("field_type", None),
                "pop_interconnection": (",").join(hubs['site_id'].unique().tolist()),
                "location_pop": None,
                "antenna_height": colo_fe.get("total_tower_height", None),
                "total_sector": None,
                "fo_type": route_type,
                "ring_id": ring,
                "distance_per_site": route['length'] if str(fe_site['site_id']) not in hubs_ids else None
            }

            if (str(ne_site['site_id']) not in hubs_ids) and (str(ne_site['site_id']) not in recorded_sites):
                sites_records.append(ne_record)
                recorded_sites.add(str(ne_site['site_id']))

            if (str(fe_site['site_id']) not in hubs_ids) and (str(fe_site['site_id']) not in recorded_sites):
                sites_records.append(fe_record)
                recorded_sites.add(str(fe_site['site_id']))
    
        # Ring Record
        ring_record = {
            "no": None,
            "province": province,
            "ring_id": ring,
            "pop_interconnection": (",").join(hubs['site_id'].unique().tolist()),
            "total_site": qty_sites,
            "route_type": route_type,
            "total_distance": routes_length,
            "avg_per_site": round(routes_length / qty_sites, 3),
        }
        ring_records.append(ring_record)
    
    # Compile DRM Summary
    segments_records = pd.DataFrame(segments_records)
    sites_records = pd.DataFrame(sites_records)
    ring_records = pd.DataFrame(ring_records)

    segments_records['no'] = segments_records.index + 1
    sites_records['no'] = sites_records.index + 1
    ring_records['no'] = ring_records.index + 1

    logger.info(f"ℹ️ Excel sheet 'Lampiran Segment' written with {len(segments_records):,} records.")
    logger.info(f"ℹ️ Excel sheet 'Lampiran Site' written with {len(sites_records):,} records.")
    logger.info(f"ℹ️ Excel sheet 'Summary Ring' written with {len(ring_records):,} records.")


    # ---------------------------
    # Write Excel
    # ---------------------------
    template_path = os.path.join(DATA_DIR, "template", "drm", "Template_DRM_Report.xlsx")
    output_path = os.path.join(export_dir, "Summary Report_Design Review Format.xlsx")

    if not os.path.exists(template_path):
        raise ValueError("BOQ template file not found in template directory.")

    shutil.copy2(template_path, output_path)

    # ---------------------------
    # PROCESS BOQ EXCEL
    # ---------------------------
    wb = load_workbook(output_path)

    # Style
    named_style = NamedStyle(name="RowStyle")
    side_style = Side(style="thin", border_style="thin")
    border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
    named_style.font = Font(name="Arial", size=10)
    named_style.border = border

    if "RowStyle" not in [s for s in wb.named_styles]:
        wb.add_named_style(named_style)

    # Segment Sheet
    segment_sheet = wb["Lampiran Segment"]
    start_data_row = 3
    col_index = {col: num for num, col in enumerate(segments_records.columns, start=1)}
    for idx, record in enumerate(segments_records.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = segment_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "RowStyle"

            if isinstance(value, (int, float)) and value is not None:
                if ('long' in str(key).lower()) or ('lat' in str(key).lower()):
                    continue

                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sites Sheet
    segment_sheet = wb["Lampiran Site"]
    start_data_row = 4
    col_index = {col: num for num, col in enumerate(sites_records.columns, start=1)}
    for idx, record in enumerate(sites_records.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = segment_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "RowStyle"

            if isinstance(value, (int, float)) and value is not None:
                if ('long' in str(key).lower()) or ('lat' in str(key).lower()):
                    continue

                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Summary Ring Sheet
    segment_sheet = wb["Summary Ring"]
    start_data_row = 3
    col_index = {col: num for num, col in enumerate(ring_records.columns, start=1)}
    for idx, record in enumerate(ring_records.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = segment_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "RowStyle"

            if isinstance(value, (int, float)) and value is not None:
                if ('long' in str(key).lower()) or ('lat' in str(key).lower()):
                    continue

                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    logger.info("✅ DRM Excel format saved.")

def boq_mmp(
    kmz_path: str,
    export_dir: str,
    operator: Operator | str = Operator.XL,
    sep: str = ";",
    interval_pole_m: int = 60,
    cable_percentage: int = 15,
    cable_multiplier: int = 2,
    sclc_enabled: bool = False,
    device_in_site: DeviceType = DeviceType.ODP,
    device_in_branch: DeviceType = DeviceType.ODP,
    connector_in_site: ConnectorType = ConnectorType.SC,
    connector_in_branch: ConnectorType = ConnectorType.SC,
    program_name: str = "TBG MMP"
):
    def even_excel(x:float|int):
        return math.ceil(x/2) * 2
    
    validated = validate_kmz_ipl(kmz_path, sep=sep)
    if validated is None:
        logger.info(f"❌ BOQ generation failed (invalid KMZ): {kmz_path}")
        return

    # ---------------------------
    # Inputs (GeoDataFrames)
    # ---------------------------
    gdf_points = validated["points_data"]
    gdf_lines = validated["lines_data"]

    gdf_hub = validated["fo_hub"]
    gdf_sitelist = validated["sitelist"]
    gdf_odp = validated["odp"]
    gdf_otb = validated["otb"]
    gdf_closure = validated["closure"]
    gdf_topology = validated["topology"]

    gdf_route = validated["route"]
    gdf_backbone = validated["backbone"]
    gdf_access = validated["access"]
    gdf_fo_exist = validated["fo_exist"]
    gdf_pole_exist = validated["pole_exist"]
    gdf_obstacle = validated["obstacle"]

    # ---------------------------
    # Enrich Metadata and CRS normalize
    # ---------------------------
    colopriming_data = pd.read_excel(f"{DATA_DIR}/Sitelist Dec 2025.xlsx")
    gdf_hub = admin_information(gdf_hub, level="kabkot")

    target_crs = 3857
    gdf_points = gdf_points.to_crs(epsg=target_crs)
    gdf_lines = gdf_lines.to_crs(epsg=target_crs)

    gdf_hub = gdf_hub.to_crs(epsg=target_crs)
    gdf_sitelist = gdf_sitelist.to_crs(epsg=target_crs)
    gdf_odp = gdf_odp.to_crs(epsg=target_crs)
    gdf_otb = gdf_otb.to_crs(epsg=target_crs)
    gdf_closure = gdf_closure.to_crs(epsg=target_crs)
    gdf_topology = gdf_topology.to_crs(epsg=target_crs)

    gdf_route = gdf_route.to_crs(epsg=target_crs)
    gdf_backbone = gdf_backbone.to_crs(epsg=target_crs)
    gdf_access = gdf_access.to_crs(epsg=target_crs)
    gdf_fo_exist = gdf_fo_exist.to_crs(epsg=target_crs)
    gdf_pole_exist = gdf_pole_exist.to_crs(epsg=target_crs)
    gdf_obstacle = gdf_obstacle.to_crs(epsg=target_crs)

    # ---------------------------
    # Compile BOQ records
    # ---------------------------
    boq_records: list[dict] = []
    recorded_segment = set()
    num = 1
    gdf_route = gdf_route.reset_index(drop=True)
    for ring_name, gdf_ring_route in gdf_route.groupby("ring_name"):
        gdf_ring_segments = gdf_ring_route.drop_duplicates().copy()
        gdf_ring_sites = gdf_sitelist[gdf_sitelist["ring_name"] == ring_name].copy()
        gdf_ring_hubs = gdf_hub[gdf_hub["ring_name"] == ring_name].copy()

        # Metadata
        program = (gdf_ring_sites["program"].mode().iloc[0] if "program" in gdf_ring_sites.columns and not gdf_ring_sites.empty else None)
        region = (gdf_ring_sites["region"].mode().iloc[0] if "region" in gdf_ring_sites.columns and not gdf_ring_sites.empty else None)
        city = (gdf_ring_hubs["Kabkot"].mode().iloc[0] if "Kabkot" in gdf_ring_hubs.columns and not gdf_ring_hubs.empty else None)
        
        is_otb = DeviceType.OTB in [device_in_site, device_in_branch]
        is_odp = DeviceType.ODP in [device_in_site, device_in_branch]
        is_sc = ConnectorType.SC in [connector_in_branch, connector_in_site]
        is_fc = ConnectorType.FC in [connector_in_branch, connector_in_site]

        # Ring-level slices
        gdf_ring_backbone = gdf_backbone[gdf_backbone["ring_name"] == ring_name].copy()
        gdf_ring_access = gdf_access[gdf_access["ring_name"] == ring_name].copy()
        gdf_ring_fo_exist = gdf_fo_exist[gdf_fo_exist["ring_name"] == ring_name].copy()
        gdf_ring_pole_exist = gdf_pole_exist[gdf_pole_exist["ring_name"] == ring_name].copy()
        gdf_ring_otb = gdf_otb[gdf_otb["ring_name"] == ring_name].copy()
        gdf_ring_odp = gdf_odp[gdf_odp["ring_name"] == ring_name].copy()
        gdf_ring_closure = gdf_closure[gdf_closure["ring_name"] == ring_name].copy()
        gdf_ring_obstacle = gdf_obstacle[gdf_obstacle["ring_name"] == ring_name].copy()

        is_first = True
        for idx, seg_row in gdf_ring_segments.iterrows():
            
            # Segment Metadata
            seg_name = seg_row["name"]
            seg_ne = seg_row["near_end"]
            seg_fe = seg_row["far_end"]
            seg_ctx = f"ring={ring_name} seg={seg_name} ne={seg_ne} fe={seg_fe}"

            site_row = colopriming_data.loc[colopriming_data["site_id"].astype(str) == str(seg_ne)].copy()
            site_row = site_row.iloc[0] if not site_row.empty else pd.Series()
            site_name = site_row.get("site_name", None)
            lat = site_row.get("lat", None)
            long = site_row.get("long", None)
            site_type = site_row.get("site_type", None)
            tower_type = site_row.get("tower_type", None)
            region = site_row.get("region", None)
            kabupaten = site_row.get("kabupaten", None)
            provinsi = site_row.get("provinsi", None)

            mmp_row = gdf_ring_sites[gdf_ring_sites["site_id"].astype(str) == str(seg_fe)].copy()
            mmp_row = mmp_row.to_crs(epsg=4326)
            mmp_row = mmp_row.iloc[0] if not mmp_row.empty else pd.Series()
            mmp_long = mmp_row.get("geometry", None).x
            mmp_lat = mmp_row.get("geometry", None).y

            len_route_m = (
                round(float(seg_row['length']), 3)
                if seg_row.length is not None
                else 0.0
            )
            seg_core = int(seg_row.get("core", 24) or 24)
            
            # Previous Route
            if idx == 0:
                prev_ring = None
                len_prev_access_m = 0
                len_prev_access_ext_m = 0
            else:
                prev_seg = gdf_route.loc[idx-1, :]
                prev_ring = prev_seg['ring_name']
                prev_ne = prev_seg['near_end']
                prev_fe = prev_seg['far_end']
                prev_df_access = gdf_ring_access[(gdf_ring_access["near_end"] == prev_ne) & (gdf_ring_access["far_end"] == prev_fe)].copy()
                len_prev_access_m = (float(sum(prev_df_access['length'])) if not prev_df_access.empty else 0.0)
                len_prev_access_ext_m = 0

            # ---------------------------
            # Segment slices
            # ---------------------------
            df_bb = gdf_ring_backbone[
                (gdf_ring_backbone["near_end"] == seg_ne)
                & (gdf_ring_backbone["far_end"] == seg_fe)
            ].copy()
            df_access = gdf_ring_access[
                (gdf_ring_access["near_end"] == seg_ne)
                & (gdf_ring_access["far_end"] == seg_fe)
            ].copy()
            df_overlap = gdf_ring_fo_exist[
                (gdf_ring_fo_exist["near_end"] == seg_ne)
                & (gdf_ring_fo_exist["far_end"] == seg_fe)
            ].copy()
            df_pole = gdf_ring_pole_exist[
                (gdf_ring_pole_exist["near_end"] == seg_ne)
                & (gdf_ring_pole_exist["far_end"] == seg_fe)
            ].copy()

            df_otb = gdf_ring_otb[gdf_ring_otb["segment"] == seg_name].copy()
            df_otb_new = df_otb[df_otb["ext_note"] == 0].copy()
            df_otb_ext = df_otb[df_otb["ext_note"] == 1].copy()

            df_odp = gdf_ring_odp[gdf_ring_odp["segment"] == seg_name].copy()
            df_odp_new = df_odp[df_odp["ext_note"] == 0].copy()
            df_odp_ext = df_odp[df_odp["ext_note"] == 1].copy()

            if df_odp.empty:
                raise ValueError(
                    f"[ODP_NOT_FOUND] {seg_ctx}\n"
                    f"Segment snapshot:\n"
                    f"{gdf_ring_segments[['segment', 'near_end', 'far_end']].head(10).to_string(index=False)}"
                )

            logger.info(f"🟢 Processing {seg_ctx}")

            df_closure = gdf_ring_closure[gdf_ring_closure["segment"] == seg_name].copy()
            df_closure_new = df_closure[df_closure["ext_note"] == 0].copy()
            df_closure_ext = df_closure[df_closure["ext_note"] == 1].copy()

            df_obs_seg = gdf_ring_obstacle[(gdf_ring_obstacle["near_end"] == seg_ne) & (gdf_ring_obstacle["far_end"] == seg_fe)].copy()
            df_obs_toll = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("toll", case=False, na=False)].copy()
            df_obs_rail = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("rail", case=False, na=False)].copy()
            df_obs_bridge = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("bridge", case=False, na=False)].copy()

            # ---------------------------
            # Core parsing
            # ---------------------------
            core_bb = 24
            if not df_bb.empty and "name" in df_bb.columns:
                raw_name = str(df_bb["name"].iloc[0])
                tail = raw_name.split("_FO")[-1].replace("C", "")
                core_bb = int(tail) if tail.isdigit() else 24

            # ---------------------------
            # Length metrics
            # ---------------------------
            len_bb_m = float(sum(df_bb['length'])) if not df_bb.empty else 0.0
            len_access_m = (float(sum(df_access['length'])) if not df_access.empty else 0.0)
            len_overlap_m = ( float(sum(df_overlap['length'])) if not df_overlap.empty else 0.0)
            len_pole_m = (float(sum(df_pole['length'])) if not df_pole.empty else 0.0)
            len_access_ext_m = 0.0

            # Cable length by backbone core
            len_cable_by_core_m = {c: (len_route_m if core_bb == c else 0.0) for c in (24, 48, 72, 96, 120, 144)}
            
            # ---------------------------
            # Quantity metrics
            # ---------------------------
            qty_otb = len(df_otb)
            qty_otb_new = len(df_otb_new)
            qty_otb_ext = len(df_otb_ext)

            qty_odp = len(df_odp)
            qty_odp_new = len(df_odp_new)
            qty_odp_ext = len(df_odp_ext)

            df_otb_by_core = {
                c: df_otb_new[df_otb_new["core"] == c] for c in (24, 48, 72, 96, 120, 144)
            }
            df_odp_by_core = {
                c: df_odp_new[df_odp_new["core"] == c] for c in (24, 48, 72, 96, 120, 144)
            }
            qty_otb_by_core = {c: len(df_) for c, df_ in df_otb_by_core.items()}
            qty_odp_by_core = {c: len(df_) for c, df_ in df_odp_by_core.items()}

            qty_closure = len(df_closure)
            qty_closure_new = len(df_closure_new)
            qty_closure_ext = len(df_closure_ext)

            qty_obs_toll = len(df_obs_toll)
            qty_obs_rail = len(df_obs_rail)
            qty_obs_bridge = len(df_obs_bridge)

            # ---------------------------
            # Calculations
            # ---------------------------
            fo_factor = 1 + (cable_percentage/100)
            calc_permission_pu = max(0, math.floor(len_bb_m + len_access_m - len_pole_m + sum(len_cable_by_core_m.get(core, 0) for core in len_cable_by_core_m.keys() if int(core) != 24)))
            calc_fo_cable_m = math.ceil(math.ceil(len_bb_m + len_access_m) * cable_multiplier * fo_factor / 100) * 100
            calc_closure_24_qty = qty_closure_new + (math.floor(calc_fo_cable_m / 4000) if calc_fo_cable_m >= 4000 else 0)
            calc_total_overlap_m = round((len_overlap_m + len_access_ext_m + len_prev_access_m + len_prev_access_ext_m if ring_name == prev_ring else len_overlap_m + len_access_ext_m) * fo_factor, 0)

            # Material
            calc_mat_hdpe_subduct_32_27_qty = 20 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 20 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0) + 70 * qty_obs_rail
            calc_mat_gi_pipe_1p5in_qty = 3 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 3 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0) + 3 * (2 * qty_obs_rail)
            calc_mat_pole_fo_9m_3step_qty = 0 if ((calc_permission_pu / interval_pole_m) < 3) else even_excel((calc_permission_pu / interval_pole_m) * 0.05) #=IF((S10/80)<3;0;EVEN(((S10)/80)*0,05))
            calc_mat_pole_fo_7m_2step_qty = 0 if calc_permission_pu < 0 else even_excel(calc_permission_pu/interval_pole_m) - calc_mat_pole_fo_9m_3step_qty #=IF((S10)<0;0;EVEN(((S10)/70)-DV10))
            calc_mat_slack_support = math.ceil((calc_fo_cable_m + calc_total_overlap_m)/400) + 1
            calc_mat_slack_support_70x70x3_qty = 1 + math.floor((calc_mat_pole_fo_7m_2step_qty + calc_mat_pole_fo_9m_3step_qty)/4) if calc_mat_pole_fo_7m_2step_qty + calc_mat_pole_fo_9m_3step_qty > 0 else 0 # =IF(SUM(DU10;DV10)>0;1+ROUNDDOWN(SUM(DU10;DV10)/4;0);0)
            
            # Services
            otb_factor = (is_otb and (is_sc or is_fc))
            is_sc_odp = (is_sc and is_odp)

            calc_svc_pulling_fo_aerial_incl_pole_m = (calc_fo_cable_m + (len_cable_by_core_m.get(core_bb, 0) if core_bb != 24 else 0) - len_pole_m - calc_mat_hdpe_subduct_32_27_qty + 0 ) if (calc_fo_cable_m + (len_cable_by_core_m.get(core_bb, 0) if core_bb != 24 else 0) >= 20) else 0
            calc_splicing_fusion = ((calc_closure_24_qty + (qty_odp_by_core.get(24, 0) if is_sc_odp else 0)) * 24 + (qty_odp_by_core.get(48, 0) if is_sc_odp else 0) * 48 + (qty_odp_by_core.get(4, 0)  if is_sc_odp else 0) * 4 + (qty_odp_by_core.get(8, 0)  if is_sc_odp else 0) * 8 + (qty_odp_by_core.get(16, 0) if is_sc_odp else 0) * 16)
            calc_termination_fusion = sum((qty_otb_by_core.get(core, 0) if otb_factor else 0) * core for core in (12, 24, 48, 96, 144, 288))

            calc_svc_splicing_fusion_qty = 24 if (calc_splicing_fusion == 0 and calc_fo_cable_m > 0) else calc_splicing_fusion
            calc_svc_termination_fusion_qty = calc_termination_fusion
            
            # Testing
            calc_test_otdr_2lambda_2way_ls = (calc_svc_termination_fusion_qty if calc_svc_termination_fusion_qty > 0 else 96 if len_cable_by_core_m.get(96, 0) > 0 else 48 if len_cable_by_core_m.get(48, 0) > 0 else 24)

            seg_record = {
            "spk_site": program_name,
            "boq_sent_date": None,
            "no": num,
            "tp": "TBG",
            "site_id": None,
            "site_name": site_name,
            "lat_site": lat,
            "lon_site": long,
            "site_type": site_type,
            "tower_type": tower_type,
            "region": region,
            "area_city": kabupaten,
            "area_province": provinsi,
            "operator": operator.upper(),
            "operator_site_id": seg_ne,
            "spk_wo": program,
            "e2e_distance_m": len_route_m,
            "prep_general_cost": 1,
            "prep_transport_material_cost": 1,
            "mat_fo_adss_24_m": calc_fo_cable_m,
            "mat_odp_24_sc_upc_qty": qty_odp_by_core.get(24, 0) if is_odp else 0,
            "mat_otb_24_sc_upc_qty": qty_otb_by_core.get(24, 0) if is_otb else 0,
            "mat_closure_dome_24_qty": calc_closure_24_qty,
            "mat_hdpe_subduct_32_27_qty": calc_mat_hdpe_subduct_32_27_qty,
            "mat_pole_fo_7m_2step_qty": calc_mat_pole_fo_7m_2step_qty,
            "mat_pole_fo_9m_3step_qty": calc_mat_pole_fo_9m_3step_qty,
            "mat_pole_accesories": calc_mat_pole_fo_7m_2step_qty,
            "mat_slack_support_qty": calc_mat_slack_support,
            "mat_gi_pipe_3in_qty": 1,
            "mat_patch_outdoor_sm_sc_upc_lc_upc_20m": 2 if (seg_name not in recorded_segment) else 0,
            "mat_patch_outdoor_sm_sc_upc_sc_upc_5m": 2 if (seg_name not in recorded_segment) else 0,
            "mat_adapter_lc_lc_duplex_qty": 0,
            "svc_pulling_fo_aerial_m": calc_fo_cable_m,
            "svc_pulling_fo_duct_m": 15,
            "svc_pulling_fo_inbuilding_m": 0,
            "svc_trenching_reinstate_makadam_m": 15,
            "svc_trenching_reinstate_makadam_out_m": 0,
            "svc_install_hdpe_placement": 15,
            "svc_install_pole_7m_2step_unit": calc_mat_pole_fo_7m_2step_qty,
            "svc_install_pole_9m_3step_unit": calc_mat_pole_fo_9m_3step_qty,
            "svc_install_pole_accesories": calc_mat_pole_fo_7m_2step_qty,
            "svc_install_slack_support_qty": calc_mat_slack_support,
            "svc_install_riser_galvanized_3m_qty": 1,
            "svc_splicing_termination_qty": 24*2,
            "svc_install_odp_otb_qty": qty_odp_by_core.get(24, 0),
            "svc_install_patch_outdoor_sc_lc_20m_qty": 2 if (seg_name not in recorded_segment) else 0,
            "svc_install_patch_outdoor_sc_sc_5m_qty": 2 if (seg_name not in recorded_segment) else 0,
            "svc_support_integration_qty": 1,
            "prep_special_permit_cost": None,
            "sow": "New MMP",
            "remark_site": None,
            "lat_mmp": mmp_lat,
            "lon_mmp": mmp_long,
            "remark_mmp": None,
            "existing_route_flag": len_overlap_m,
            }

            boq_records.append(seg_record)

            # Update Segment Info
            recorded_segment.add(seg_name)
            is_first = False
            num += 1

    boq_df = pd.DataFrame(boq_records)

    # ---------------------------
    # Write into Excel template
    # ---------------------------
    template_path = os.path.join(DATA_DIR, "template", "boq", "Template_BOQ_MMP.xlsx")
    output_path = os.path.join(export_dir, "BOQ MMP Report.xlsx")

    if not os.path.exists(template_path):
        raise ValueError("BOQ template file not found in template directory.")

    shutil.copy2(template_path, output_path)

    # ---------------------------
    # PROCESS BOQ EXCEL
    # ---------------------------
    wb = load_workbook(output_path)

    # Style
    named_style = NamedStyle(name="BOQ Row")
    side_style = Side(style="thin", border_style="thin")
    border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
    named_style.font = Font(name="Arial", size=10)
    named_style.border = border

    if "BOQ Row" not in [s for s in wb.named_styles]:
        wb.add_named_style(named_style)

    # BOQ Sheet
    boq_sheet = wb["BOQ"]
    start_data_row = 2
    col_index = {col: num for num, col in enumerate(boq_df.columns, start=1)}
    for idx, record in enumerate(boq_df.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = boq_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "BOQ Row"

            if isinstance(value, (int, float)) and value is not None:
                if ('long' in str(key).lower()) or ('lat' in str(key).lower()):
                    continue

                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    logger.info("✅ BOQ Excel saved.")

def boq_generation(
    kmz_path: str,
    export_dir: str,
    operator: Operator | str = Operator.XL,
    sep: str = ";",
    interval_pole_m: int = 80,
    cable_percentage: int = 10,
    sclc_enabled: bool = False,
    device_in_site: DeviceType = DeviceType.OTB,
    device_in_branch: DeviceType = DeviceType.ODP,
    connector_in_site: ConnectorType = ConnectorType.SC,
    connector_in_branch: ConnectorType = ConnectorType.SC,
    program_name: str = "Intersite FO"
):
    def even_excel(x:float|int):
        return math.ceil(x/2) * 2
    
    validated = validate_kmz_ipl(kmz_path, sep=sep)
    if validated is None:
        logger.info(f"❌ BOQ generation failed (invalid KMZ): {kmz_path}")
        return

    # ---------------------------
    # Inputs (GeoDataFrames)
    # ---------------------------
    gdf_points = validated["points_data"]
    gdf_lines = validated["lines_data"]

    gdf_hub = validated["fo_hub"]
    gdf_sitelist = validated["sitelist"]
    gdf_odp = validated["odp"]
    gdf_otb = validated["otb"]
    gdf_closure = validated["closure"]
    gdf_topology = validated["topology"]

    gdf_route = validated["route"]
    gdf_backbone = validated["backbone"]
    gdf_access = validated["access"]
    gdf_fo_exist = validated["fo_exist"]
    gdf_pole_exist = validated["pole_exist"]
    gdf_obstacle = validated["obstacle"]

    # ---------------------------
    # Enrich Metadata and CRS normalize
    # ---------------------------
    gdf_hub = admin_information(gdf_hub, level="kabkot")

    target_crs = 3857
    gdf_points = gdf_points.to_crs(epsg=target_crs)
    gdf_lines = gdf_lines.to_crs(epsg=target_crs)

    gdf_hub = gdf_hub.to_crs(epsg=target_crs)
    gdf_sitelist = gdf_sitelist.to_crs(epsg=target_crs)
    gdf_odp = gdf_odp.to_crs(epsg=target_crs)
    gdf_otb = gdf_otb.to_crs(epsg=target_crs)
    gdf_closure = gdf_closure.to_crs(epsg=target_crs)
    gdf_topology = gdf_topology.to_crs(epsg=target_crs)

    gdf_route = gdf_route.to_crs(epsg=target_crs)
    gdf_backbone = gdf_backbone.to_crs(epsg=target_crs)
    gdf_access = gdf_access.to_crs(epsg=target_crs)
    gdf_fo_exist = gdf_fo_exist.to_crs(epsg=target_crs)
    gdf_pole_exist = gdf_pole_exist.to_crs(epsg=target_crs)
    gdf_obstacle = gdf_obstacle.to_crs(epsg=target_crs)

    # ---------------------------
    # Compile BOQ records
    # ---------------------------
    piv_records: list[dict] = []
    boq_records: list[dict] = []
    boq_pr_records: list[dict] = []
    recorded_segment = set()
    num = 1
    gdf_route = gdf_route.reset_index(drop=True)
    for ring_name, gdf_ring_route in gdf_route.groupby("ring_name"):
        gdf_ring_segments = gdf_ring_route.drop_duplicates().copy()
        gdf_ring_sites = gdf_sitelist[gdf_sitelist["ring_name"] == ring_name].copy()
        gdf_ring_hubs = gdf_hub[gdf_hub["ring_name"] == ring_name].copy()

        # Metadata
        program = (gdf_ring_sites["program"].mode().iloc[0] if "program" in gdf_ring_sites.columns and not gdf_ring_sites.empty else None)
        region = (gdf_ring_sites["region"].mode().iloc[0] if "region" in gdf_ring_sites.columns and not gdf_ring_sites.empty else None)
        city = (gdf_ring_hubs["Kabkot"].mode().iloc[0] if "Kabkot" in gdf_ring_hubs.columns and not gdf_ring_hubs.empty else None)
        
        is_otb = DeviceType.OTB in [device_in_site, device_in_branch]
        is_odp = DeviceType.ODP in [device_in_site, device_in_branch]
        is_sc = ConnectorType.SC in [connector_in_branch, connector_in_site]
        is_fc = ConnectorType.FC in [connector_in_branch, connector_in_site]

        # Ring-level slices
        gdf_ring_backbone = gdf_backbone[gdf_backbone["ring_name"] == ring_name].copy()
        gdf_ring_access = gdf_access[gdf_access["ring_name"] == ring_name].copy()
        gdf_ring_fo_exist = gdf_fo_exist[gdf_fo_exist["ring_name"] == ring_name].copy()
        gdf_ring_pole_exist = gdf_pole_exist[gdf_pole_exist["ring_name"] == ring_name].copy()
        gdf_ring_otb = gdf_otb[gdf_otb["ring_name"] == ring_name].copy()
        gdf_ring_odp = gdf_odp[gdf_odp["ring_name"] == ring_name].copy()
        gdf_ring_closure = gdf_closure[gdf_closure["ring_name"] == ring_name].copy()
        gdf_ring_obstacle = gdf_obstacle[gdf_obstacle["ring_name"] == ring_name].copy()

        is_first = True
        for idx, seg_row in gdf_ring_segments.iterrows():
            seg_name = seg_row["name"]
            seg_ne = seg_row["near_end"]
            seg_fe = seg_row["far_end"]
            seg_ctx = f"ring={ring_name} seg={seg_name} ne={seg_ne} fe={seg_fe}"

            len_route_m = (
                round(float(seg_row['length']), 3)
                if seg_row.length is not None
                else 0.0
            )
            seg_core = int(seg_row.get("core", 24) or 24)
            
            # Previous Route
            if idx == 0:
                prev_ring = None
                len_prev_access_m = 0
                len_prev_access_ext_m = 0
            else:
                prev_seg = gdf_route.loc[idx-1, :]
                prev_ring = prev_seg['ring_name']
                prev_ne = prev_seg['near_end']
                prev_fe = prev_seg['far_end']
                prev_df_access = gdf_ring_access[(gdf_ring_access["near_end"] == prev_ne) & (gdf_ring_access["far_end"] == prev_fe)].copy()
                len_prev_access_m = (float(sum(prev_df_access['length'])) if not prev_df_access.empty else 0.0)
                len_prev_access_ext_m = 0

            # ---------------------------
            # Segment slices
            # ---------------------------
            df_bb = gdf_ring_backbone[
                (gdf_ring_backbone["near_end"] == seg_ne)
                & (gdf_ring_backbone["far_end"] == seg_fe)
            ].copy()
            df_access = gdf_ring_access[
                (gdf_ring_access["near_end"] == seg_ne)
                & (gdf_ring_access["far_end"] == seg_fe)
            ].copy()
            df_overlap = gdf_ring_fo_exist[
                (gdf_ring_fo_exist["near_end"] == seg_ne)
                & (gdf_ring_fo_exist["far_end"] == seg_fe)
            ].copy()
            df_pole = gdf_ring_pole_exist[
                (gdf_ring_pole_exist["near_end"] == seg_ne)
                & (gdf_ring_pole_exist["far_end"] == seg_fe)
            ].copy()

            df_otb = gdf_ring_otb[gdf_ring_otb["segment"] == seg_name].copy()
            df_otb_new = df_otb[df_otb["ext_note"] == 0].copy()
            df_otb_ext = df_otb[df_otb["ext_note"] == 1].copy()

            df_odp = gdf_ring_odp[gdf_ring_odp["segment"] == seg_name].copy()
            df_odp_new = df_odp[df_odp["ext_note"] == 0].copy()
            df_odp_ext = df_odp[df_odp["ext_note"] == 1].copy()

            if df_odp.empty:
                raise ValueError(
                    f"[ODP_NOT_FOUND] {seg_ctx}\n"
                    f"Segment snapshot:\n"
                    f"{gdf_ring_segments[['segment', 'near_end', 'far_end']].head(10).to_string(index=False)}"
                )

            logger.info(f"🟢 Processing {seg_ctx}")

            df_closure = gdf_ring_closure[gdf_ring_closure["segment"] == seg_name].copy()
            df_closure_new = df_closure[df_closure["ext_note"] == 0].copy()
            df_closure_ext = df_closure[df_closure["ext_note"] == 1].copy()

            df_obs_seg = gdf_ring_obstacle[(gdf_ring_obstacle["near_end"] == seg_ne) & (gdf_ring_obstacle["far_end"] == seg_fe)].copy()
            df_obs_toll = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("toll", case=False, na=False)].copy()
            df_obs_rail = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("rail", case=False, na=False)].copy()
            df_obs_bridge = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("bridge", case=False, na=False)].copy()

            # ---------------------------
            # Core parsing
            # ---------------------------
            core_bb = 24
            if not df_bb.empty and "name" in df_bb.columns:
                raw_name = str(df_bb["name"].iloc[0])
                tail = raw_name.split("_FO")[-1].replace("C", "")
                core_bb = int(tail) if tail.isdigit() else 24

            # ---------------------------
            # Length metrics
            # ---------------------------
            len_bb_m = float(sum(df_bb['length'])) if not df_bb.empty else 0.0
            len_access_m = (float(sum(df_access['length'])) if not df_access.empty else 0.0)
            len_overlap_m = ( float(sum(df_overlap['length'])) if not df_overlap.empty else 0.0)
            len_pole_m = (float(sum(df_pole['length'])) if not df_pole.empty else 0.0)
            len_access_ext_m = 0.0

            # Cable length by backbone core
            len_cable_by_core_m = {c: (len_route_m if core_bb == c else 0.0) for c in (24, 48, 72, 96, 120, 144)}
            
            # ---------------------------
            # Quantity metrics
            # ---------------------------
            qty_otb = len(df_otb)
            qty_otb_new = len(df_otb_new)
            qty_otb_ext = len(df_otb_ext)

            qty_odp = len(df_odp)
            qty_odp_new = len(df_odp_new)
            qty_odp_ext = len(df_odp_ext)

            df_otb_by_core = {
                c: df_otb_new[df_otb_new["core"] == c] for c in (24, 48, 72, 96, 120, 144)
            }
            df_odp_by_core = {
                c: df_odp_new[df_odp_new["core"] == c] for c in (24, 48, 72, 96, 120, 144)
            }
            qty_otb_by_core = {c: len(df_) for c, df_ in df_otb_by_core.items()}
            qty_odp_by_core = {c: len(df_) for c, df_ in df_odp_by_core.items()}

            qty_closure = len(df_closure)
            qty_closure_new = len(df_closure_new)
            qty_closure_ext = len(df_closure_ext)

            qty_obs_toll = len(df_obs_toll)
            qty_obs_rail = len(df_obs_rail)
            qty_obs_bridge = len(df_obs_bridge)

            # ---------------------------
            # Calculations
            # ---------------------------
            fo_factor = 1 + (cable_percentage/100)
            calc_permission_pu = max(0, math.floor(len_bb_m + len_access_m - len_pole_m + sum(len_cable_by_core_m.get(core, 0) for core in len_cable_by_core_m.keys() if int(core) != 24)))
            calc_fo_cable_m = math.ceil(math.ceil(len_bb_m + len_access_m) * fo_factor / 100) * 100
            calc_closure_24_qty = qty_closure_new + (math.floor(calc_fo_cable_m / 4000) if calc_fo_cable_m >= 4000 else 0)
            calc_total_overlap_m = round((len_overlap_m + len_access_ext_m + len_prev_access_m + len_prev_access_ext_m if ring_name == prev_ring else len_overlap_m + len_access_ext_m) * fo_factor, 0)

            # Material
            calc_mat_hdpe_subduct_32_27_qty = 20 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 20 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0) + 70 * qty_obs_rail
            calc_mat_gi_pipe_1p5in_qty = 3 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 3 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0) + 3 * (2 * qty_obs_rail)
            calc_mat_pole_fo_9m_3step_qty = 0 if ((calc_permission_pu / interval_pole_m) < 3) else even_excel((calc_permission_pu / interval_pole_m) * 0.05) #=IF((S10/80)<3;0;EVEN(((S10)/80)*0,05))
            calc_mat_pole_fo_7m_2step_qty = 0 if calc_permission_pu < 0 else even_excel(calc_permission_pu/interval_pole_m) - calc_mat_pole_fo_9m_3step_qty #=IF((S10)<0;0;EVEN(((S10)/70)-DV10))
            calc_mat_slack_support_70x70x3_qty = 1 + math.floor((calc_mat_pole_fo_7m_2step_qty + calc_mat_pole_fo_9m_3step_qty)/4) if calc_mat_pole_fo_7m_2step_qty + calc_mat_pole_fo_9m_3step_qty > 0 else 0 # =IF(SUM(DU10;DV10)>0;1+ROUNDDOWN(SUM(DU10;DV10)/4;0);0)
            
            # Services
            otb_factor = (is_otb and (is_sc or is_fc))
            is_sc_odp = (is_sc and is_odp)

            calc_svc_pulling_fo_aerial_incl_pole_m = (calc_fo_cable_m + (len_cable_by_core_m.get(core_bb, 0) if core_bb != 24 else 0) - len_pole_m - calc_mat_hdpe_subduct_32_27_qty + 0 ) if (calc_fo_cable_m + (len_cable_by_core_m.get(core_bb, 0) if core_bb != 24 else 0) >= 20) else 0
            calc_splicing_fusion = ((calc_closure_24_qty + (qty_odp_by_core.get(24, 0) if is_sc_odp else 0)) * 24 + (qty_odp_by_core.get(48, 0) if is_sc_odp else 0) * 48 + (qty_odp_by_core.get(4, 0)  if is_sc_odp else 0) * 4 + (qty_odp_by_core.get(8, 0)  if is_sc_odp else 0) * 8 + (qty_odp_by_core.get(16, 0) if is_sc_odp else 0) * 16)
            calc_termination_fusion = sum((qty_otb_by_core.get(core, 0) if otb_factor else 0) * core for core in (12, 24, 48, 96, 144, 288))

            calc_svc_splicing_fusion_qty = 24 if (calc_splicing_fusion == 0 and calc_fo_cable_m > 0) else calc_splicing_fusion
            calc_svc_termination_fusion_qty = calc_termination_fusion
            
            # Testing
            calc_test_otdr_2lambda_2way_ls = (calc_svc_termination_fusion_qty if calc_svc_termination_fusion_qty > 0 else 96 if len_cable_by_core_m.get(96, 0) > 0 else 48 if len_cable_by_core_m.get(48, 0) > 0 else 24)

            # ---------------------------
            # Pivot record
            # ---------------------------
            piv_record = {
                "program": program,
                "operator": operator.upper(),
                "region": region,
                "city": city,
                "ring_name": ring_name,
                "segment_name": seg_name,
                "near_end": seg_ne,
                "far_end": seg_fe,
                "route_qty": None,
                "route_length_m": len_route_m,
                "bb_qty": None,
                "bb_length_m": len_bb_m,
                "access_qty": None,
                "access_length_m": len_access_m,
                "access_ext_qty": None,
                "access_ext_length_m": len_access_ext_m,
                "overlap_qty": None,
                "overlap_length_m": len_overlap_m,
                "pole_ext_length_m": len_pole_m,
                "otb_qty": qty_otb,
                "otb_new_qty": qty_otb_new,
                "otb_ext_qty": qty_otb_ext,
                "odp_qty": qty_odp,
                "odp_new_qty": qty_odp_new,
                "odp_ext_qty": qty_odp_ext,
                "odp_24_qty": qty_odp_by_core.get(24, 0),
                "odp_48_qty": qty_odp_by_core.get(48, 0),
                "odp_72_qty": qty_odp_by_core.get(72, 0),
                "odp_96_qty": qty_odp_by_core.get(96, 0),
                "odp_120_qty": qty_odp_by_core.get(120, 0),
                "odp_144_qty": qty_odp_by_core.get(144, 0),
                "closure_qty": qty_closure,
                "closure_new_qty": qty_closure_new,
                "closure_ext_qty": qty_closure_ext,
                "obstacle_toll_qty": qty_obs_toll,
                "obstacle_railway_qty": qty_obs_rail,
                "obstacle_bridge_qty": qty_obs_bridge,
                "permission_pu": calc_permission_pu,
                "fo_cable_m": calc_fo_cable_m,
                "total_overlap_m": calc_total_overlap_m,
                "cable_24_m": len_cable_by_core_m[24],
                "cable_48_m": len_cable_by_core_m[48],
                "cable_72_m": len_cable_by_core_m[72],
                "cable_96_m": len_cable_by_core_m[96],
                "cable_120_m": len_cable_by_core_m[120],
                "cable_144_m": len_cable_by_core_m[144],
            }
            piv_records.append(piv_record)

            # BOQ Record
            seg_record = {
                # Segment Info
                "no": num,
                "site_type": program if program != "NA" else program_name,
                "stip_category": None,
                "operator": operator.upper(),
                "spk_wo": None,
                "segment_type": "Segment",
                "sonumb": None,
                "ring_id": ring_name,
                "segment_id": seg_name,
                "area_city": city,
                "program": region,
                "work_code": None,
                "region_procurement": None,
                "e2e_distance_m": calc_fo_cable_m + calc_total_overlap_m, # len_cable_by_core_m[48] + len_cable_by_core_m[72] + len_cable_by_core_m[96] + len_cable_by_core_m[120] + len_cable_by_core_m[144]
                "boq_id": None,
                "subset_separator": None,

                # PREPARATION
                "prep_a1_rambu_papan": int(is_first),
                "prep_a2_direksi_keet": int(is_first),
                "prep_a3_koordinasi_pu_m": calc_permission_pu,
                "prep_a4_koordinasi_pjka": qty_obs_rail,
                "prep_a5_koordinasi_toll": qty_obs_toll,
                "prep_a6_koordinasi_private_supervise": None,
                "prep_a7_survey_abd": 1,
                "prep_a8_mobilisasi_transpor": int(is_first),

                # MATERIAL SUPPLY
                # Fo Cable
                "mat_fo_aerial_12_m": None,
                "mat_fo_aerial_24_m": None,
                "mat_fo_aerial_48_m": None,
                "mat_fo_aerial_72_m": None,
                "mat_fo_aerial_96_m": None,
                "mat_fo_aerial_144_m": None,
                "mat_fo_adss_12_m": len_cable_by_core_m.get(12, 0),
                "mat_fo_adss_24_m": calc_fo_cable_m, #  len_cable_by_core_m.get(24, 100) if calc_fo_cable_m is None else 
                "mat_fo_adss_48_m": len_cable_by_core_m.get(48, 0),
                "mat_fo_adss_72_m": len_cable_by_core_m.get(72, 0),
                "mat_fo_adss_96_m": len_cable_by_core_m.get(96, 0),
                "mat_fo_adss_144_m": len_cable_by_core_m.get(144, 0),
                "mat_fo_db_sj_12_m": None,
                "mat_fo_db_sj_24_m": None,
                "mat_fo_db_sj_48_m": None,
                "mat_fo_db_sj_72_m": None,
                "mat_fo_db_sj_96_m": None,
                "mat_fo_db_sj_144_m": None,
                "mat_fo_db_sj_264_m": None,
                "mat_fo_db_sj_288_m": None,
                "mat_fo_duct_12_m": None,
                "mat_fo_duct_24_m": None,
                "mat_fo_duct_48_m": None,
                "mat_fo_duct_72_m": None,
                "mat_fo_duct_96_m": None,
                "mat_fo_duct_144_m": None,
                "mat_fo_duct_264_m": None,
                "mat_fo_duct_288_m": None,
                "mat_fo_micro_24_m": None,
                "mat_fo_micro_48_m": None,
                "mat_fo_micro_96_m": None,
                "mat_fo_micro_144_m": None,
                "mat_fo_micro_288_m": None,
                "mat_fo_dropwire_1_m": None,
                "mat_fo_dropwire_2_m": None,
                "mat_closure_dome_24_qty": calc_closure_24_qty,
                "mat_closure_dome_48_qty": None,
                "mat_closure_dome_144_qty": None,
                "mat_closure_dome_96_qty": None,
                "mat_closure_dome_864_hd_qty": None,
                "mat_closure_inline_24_qty": None,
                "mat_closure_inline_48_qty": None,
                "mat_closure_inline_96_qty": None,
                "mat_closure_inline_144_qty": None,
                "mat_closure_inline_288_qty": None,
                "mat_otb_12_sc_upc_qty": qty_otb_by_core.get(12, 0) if (is_sc and is_otb) else None,
                "mat_otb_12_fc_upc_qty": qty_otb_by_core.get(12, 0) if (is_fc and is_otb) else None,
                "mat_otb_24_sc_upc_qty": qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else None,
                "mat_otb_24_fc_upc_qty": qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else None,
                "mat_otb_48_sc_upc_qty": qty_otb_by_core.get(48, 0) if (is_sc and is_otb) else None,
                "mat_otb_96_sc_upc_qty": qty_otb_by_core.get(96, 0) if (is_sc and is_otb) else None,
                "mat_otb_144_sc_upc_qty": qty_otb_by_core.get(144, 0) if (is_sc and is_otb) else None,
                "mat_otb_288_sc_upc_qty": qty_otb_by_core.get(288, 0) if (is_sc and is_otb) else None,
                "mat_otb_288_lc_upc_hd_qty": None,
                "mat_odp_4_sc_upc_qty": qty_odp_by_core.get(4, 0) if (is_sc and is_odp) else None,
                "mat_odp_8_sc_upc_qty": qty_odp_by_core.get(8, 0) if (is_sc and is_odp) else None,
                "mat_odp_16_sc_upc_qty": qty_odp_by_core.get(16, 0) if (is_sc and is_odp) else None,
                "mat_odp_24_sc_upc_qty": qty_odp_by_core.get(24, 0) if (is_sc and is_odp) else None,
                "mat_odp_96_sc_upc_qty": qty_odp_by_core.get(96, 0) if (is_sc and is_odp) else None,
                "mat_odp_pedestal_16_ug_qty": None,
                "mat_splitter_1to2_sc_upc_qty": None,
                "mat_splitter_1to4_sc_upc_qty": None,
                "mat_splitter_1to8_sc_upc_qty": None,
                "mat_splitter_1to16_sc_upc_qty": None,
                "mat_splitter_1to32_sc_upc_qty": None,
                "mat_rosset_2port_sc_upc_qty": None,
                "mat_hdpe_subduct_32_27_qty": calc_mat_hdpe_subduct_32_27_qty,
                "mat_hdpe_subduct_40_33_qty": None,
                "mat_hdpe_subduct_50_43_qty": None,
                "mat_hdpe_microduct_4way_12_10_qty": None,
                "mat_hdpe_microduct_7way_12_10_qty": None,
                "mat_hdpe_microduct_14way_12_10_qty": None,
                "mat_hdpe_microduct_4way_18_14_qty": None,
                "mat_hdpe_microduct_7way_18_14_qty": None,
                "mat_hdpe_microduct_14way_18_14_qty": None,
                "mat_gi_pipe_1in_qty": None,
                "mat_gi_pipe_1p5in_qty": calc_mat_gi_pipe_1p5in_qty,
                "mat_gi_pipe_2in_qty": None,
                "mat_gi_pipe_3in_qty": None,
                "mat_gi_pipe_4in_qty": None,
                "mat_gi_pipe_5in_qty": None,
                "mat_pvc_pipe_20mm_qty": None,
                "mat_pvc_pipe_1p5in_qty": None,
                "mat_pvc_pipe_2in_qty": None,
                "mat_pvc_pipe_3in_qty": None,
                "mat_pvc_pipe_4in_qty": None,
                "mat_pvc_pipe_5in_qty": None,
                "mat_flex_conduit_0p75in_qty": None,
                "mat_flex_conduit_2in_qty": None,
                "mat_flex_conduit_3in_qty": None,
                "mat_flex_conduit_20mm_qty": None,
                "mat_rack_open_42u_qty": None,
                "mat_rack_closed_8u_450_qty": None,
                "mat_rack_closed_20u_900_qty": None,
                "mat_rack_closed_42u_900_qty": None,
                "mat_rack_closed_42u_1150_qty": None,
                "mat_odc_144_qty": None,
                "mat_odc_288_qty": None,
                "mat_odc_576_qty": None,
                "mat_warning_tape_fo_6in_m": None,
                "mat_pole_fo_7m_2step_qty": calc_mat_pole_fo_7m_2step_qty,
                "mat_pole_fo_9m_3step_qty": calc_mat_pole_fo_9m_3step_qty,
                "mat_slack_support_50x50x3_qty": None,
                "mat_pole_fo_12m_3step_qty": None,
                "mat_slack_support_70x70x3_qty": calc_mat_slack_support_70x70x3_qty,
                "mat_patch_indoor_sm_sc_upc_sc_upc_5m": None,
                "mat_patch_indoor_sm_sc_upc_sc_upc_20m": None,
                "mat_patch_indoor_sm_sc_upc_lc_upc_5m": None,
                "mat_patch_indoor_sm_sc_upc_lc_upc_20m": None,
                "mat_patch_indoor_sm_lc_upc_lc_upc_5m": None,
                "mat_patch_indoor_sm_lc_upc_lc_upc_20m": None,
                "mat_patch_indoor_sm_fc_upc_lc_upc_5m": None,
                "mat_patch_indoor_sm_fc_upc_lc_upc_20m": None,
                "mat_patch_outdoor_sm_sc_upc_sc_upc_5m": 2 if seg_name not in recorded_segment else 0,
                "mat_patch_outdoor_sm_sc_upc_sc_upc_20m": None,
                "mat_patch_outdoor_sm_sc_upc_lc_upc_5m": 2 if (seg_name not in recorded_segment and sclc_enabled) else 0,
                "mat_patch_outdoor_sm_sc_upc_lc_upc_20m": None,
                "mat_patch_outdoor_sm_lc_upc_lc_upc_5m": None,
                "mat_patch_outdoor_sm_lc_upc_lc_upc_20m": None,
                "mat_patch_indoor_mm_lc_upc_lc_upc_5m": None,
                "mat_patch_indoor_mm_lc_upc_lc_upc_20m": None,
                "mat_pigtail_indoor_sc_upc_qty": None,
                "mat_adapter_fc_fc_qty": None,
                "mat_adapter_sc_sc_simplex_qty": None,
                "mat_adapter_sc_sc_duplex_qty": None,
                "mat_adapter_lc_lc_simplex_qty": None,
                "mat_adapter_lc_lc_duplex_qty": None,
                "mat_adapter_sc_lc_hybrid_qty": None,
                "mat_adapter_fc_lc_hybrid_qty": None,
                "mat_adapter_fc_sc_hybrid_qty": None,
                "mat_clamp_omega_qty": None,
                "mat_supporting_material_qty": int(is_first),

                # SERVICES
                "svc_trenching_reinstate_hotmix_m": None,
                "svc_trenching_reinstate_asphalt_m": None,
                "svc_trenching_reinstate_makadam_m": 20 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 20 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0),
                "svc_trenching_reinstate_floor_cement_m": None,
                "svc_trenching_reinstate_paving_m": None,
                "svc_trenching_reinstate_agregat_sirtu_m": None,
                "svc_trenching_reinstate_taman_m": None,
                "svc_trenching_reinstate_tanah_biasa_m": None,
                "svc_install_hdpe_subduct_32_27_m": calc_mat_hdpe_subduct_32_27_qty,
                "svc_install_hdpe_subduct_40_33_m": None,
                "svc_install_hdpe_subduct_50_43_m": None,
                "svc_install_hdpe_microduct_4way_12_10_m": None,
                "svc_install_hdpe_microduct_7way_12_10_m": None,
                "svc_install_hdpe_microduct_14way_12_10_m": None,
                "svc_install_hdpe_microduct_4way_18_14_m": None,
                "svc_install_hdpe_microduct_7way_18_14_m": None,
                "svc_install_hdpe_microduct_14way_18_14_m": None,
                "svc_install_galvanized_1inch_m": None,
                "svc_install_galvanized_1_5inch_m": None,
                "svc_install_galvanized_2inch_m": None,
                "svc_install_galvanized_3inch_m": None,
                "svc_install_galvanized_4inch_m": None,
                "svc_install_galvanized_5inch_m": None,
                "svc_install_pvc_20mm_m": None,
                "svc_install_pvc_1_5inch_m": None,
                "svc_install_pvc_2inch_m": None,
                "svc_install_pvc_3inch_m": None,
                "svc_install_pvc_4inch_m": None,
                "svc_install_pvc_5inch_m": None,
                "svc_install_pole_7m_2step_unit": calc_mat_pole_fo_7m_2step_qty,
                "svc_install_pole_9m_3step_unit": calc_mat_pole_fo_9m_3step_qty,
                "svc_install_slack_support_50x50x3_qty": None,
                "svc_install_slack_support_additional_qty": None,
                "svc_install_slack_support_70x70x3_qty": calc_mat_slack_support_70x70x3_qty,
                "svc_install_riser_galvanized_1_5inch_3m_qty": calc_mat_gi_pipe_1p5in_qty / 3,
                "svc_install_riser_galvanized_2inch_3m_qty": None,
                "svc_supply_install_temberang_tarik_unit": None,
                "svc_supply_install_bridge_crossing_single_pole_unit": None,
                "svc_supply_install_bridge_crossing_double_pole_unit": None,
                "svc_install_galvanized_2inch_atb_m": 0, # Perlu cross bridge information
                "svc_install_galvanized_4inch_atb_m": None,
                "svc_install_galvanized_4inch_self_support_m": None,
                "svc_boring_manual_subduct_m": None,
                "svc_boring_manual_crossing_road_m": None,
                "svc_boring_manual_crossing_toll_m": None,
                "svc_boring_manual_crossing_railway_m": qty_obs_rail * 70,
                "svc_boring_machine_crossing_railway_m": None,
                "svc_boring_manual_crossing_river_m": None,
                "svc_boring_machine_crossing_river_m": None,
                "svc_supply_install_marking_post_unit": None,
                "svc_install_manhole_100x100x120_unit": None,
                "svc_install_handhole_80x80x120_unit": qty_obs_rail * 2,
                "svc_install_handhole_60x60x120_unit": None,
                "svc_install_odc_foundation_50x70x50_unit": None,
                "svc_pulling_fo_aerial_incl_pole_m": round(calc_svc_pulling_fo_aerial_incl_pole_m, 0),
                "svc_pulling_fo_aerial_excl_pole_m": round(len_pole_m, 0),
                "svc_pulling_fo_burial_m": calc_mat_hdpe_subduct_32_27_qty,
                "svc_pulling_fo_direct_buried_m": None,
                "svc_air_blown_fo_m": None,
                "svc_pulling_fo_drop_wire_m": None,
                "svc_pulling_fo_inbuilding_m": None,
                "svc_splicing_fusion_qty": calc_svc_splicing_fusion_qty, # Recheck Long Formula
                "svc_termination_fusion_qty": calc_svc_termination_fusion_qty, # Recheck Long Formula
                "svc_bobok_tembok_bor_qty": None,
                "svc_install_patch_indoor_sc_sc_5m_qty": None,
                "svc_install_patch_indoor_sc_sc_20m_qty": None,
                "svc_install_patch_indoor_sc_lc_5m_qty": None,
                "svc_install_patch_indoor_sc_lc_20m_qty": None,
                "svc_install_patch_indoor_lc_lc_5m_qty": None,
                "svc_install_patch_indoor_lc_lc_20m_qty": None,
                "svc_install_patch_outdoor_sc_sc_5m_qty": 2 if seg_name not in recorded_segment else 0,
                "svc_install_patch_outdoor_sc_sc_20m_qty": None,
                "svc_install_patch_outdoor_sc_lc_5m_qty": 2 if (seg_name not in recorded_segment and sclc_enabled) else 0,
                "svc_install_patch_outdoor_sc_lc_20m_qty": None,
                "svc_install_patch_outdoor_lc_lc_5m_qty": None,
                "svc_install_patch_outdoor_lc_lc_20m_qty": None,
                "svc_install_patch_indoor_mm_lc_lc_5m_qty": None,
                "svc_install_patch_indoor_mm_lc_lc_20m_qty": None,
                "svc_install_pigtail_indoor_qty": None,
                "svc_install_adapter_fc_fc_qty": None,
                "svc_install_adapter_sc_sc_simplex_qty": None,
                "svc_install_adapter_sc_sc_duplex_qty": None,
                "svc_install_adapter_lc_lc_simplex_qty": None,
                "svc_install_adapter_lc_lc_duplex_qty": None,
                "svc_install_adapter_sc_lc_hybrid_qty": None,
                "svc_install_adapter_fc_lc_hybrid_qty": None,
                "svc_install_adapter_fc_sc_hybrid_qty": None,
                "svc_install_otb_12c_sc_upc_qty": None,
                "svc_install_otb_12c_fc_upc_qty": None,
                "svc_install_otb_24c_sc_upc_qty": qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else None,
                "svc_install_otb_24c_fc_upc_qty": qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else None,
                "svc_install_otb_48c_sc_upc_qty": None,
                "svc_install_otb_96c_sc_upc_qty": None,
                "svc_install_otb_144c_sc_upc_qty": None,
                "svc_install_otb_288c_sc_upc_qty": None,
                "svc_install_otb_288c_lc_upc_hd_qty": None,
                "svc_install_odp_4c_sc_upc_qty": qty_odp_by_core.get(4, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_8c_sc_upc_qty": qty_odp_by_core.get(8, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_12c_sc_upc_qty": qty_odp_by_core.get(12, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_24c_sc_upc_qty": qty_odp_by_core.get(24, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_48c_sc_upc_qty": qty_odp_by_core.get(48, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_96c_sc_upc_qty": qty_odp_by_core.get(96, 0) if (is_sc and is_odp) else None,
                "svc_support_integration_qty": 1,
                "test_otdr_2lambda_2way_ls": calc_test_otdr_2lambda_2way_ls,
                "test_opm_2lambda_2way_ls": calc_test_otdr_2lambda_2way_ls,
                "doc_hardcopy_ls": int(is_first),
                "doc_softcopy_ls": int(is_first),
                "overlap_fo": calc_total_overlap_m,
            }

            # BOQ PR Record
            pr_record = {
                "num": num,
                "site_type": program if program != "NA" else "Intersite FO",
                "stip_category": None,
                "operator": operator.upper(),
                "spk_wo": None,
                "segment_type": "Segment",
                "sonumb": None,
                "ring_id": ring_name,
                "segment_id": seg_name,
                "permit": calc_permission_pu,
                "shop_incl_pole_aerial": round(calc_svc_pulling_fo_aerial_incl_pole_m, 0),
                "shop_excl_pole_aerial": round(len_pole_m, 0),
                "shop_burial": calc_mat_hdpe_subduct_32_27_qty,
                "comcase": calc_permission_pu,
                "mat_pole_7": calc_mat_pole_fo_7m_2step_qty,
                "mat_pole_9": calc_mat_pole_fo_9m_3step_qty,
                "mat_cable_fo_24": calc_fo_cable_m,
                "mat_cable_fo_96": len_cable_by_core_m.get(96, 0),
                "mat_hdpe": calc_mat_hdpe_subduct_32_27_qty,
                "total_e2e_dist": calc_fo_cable_m + calc_total_overlap_m,
                "pole_exist": len_pole_m,
            }
            
            boq_records.append(seg_record)
            boq_pr_records.append(pr_record)

            # Update Segment Info
            recorded_segment.add(seg_name)
            is_first = False
            num += 1

    piv_df = pd.DataFrame(piv_records)
    boq_df = pd.DataFrame(boq_records)
    boq_pr_df = pd.DataFrame(boq_pr_records)

    # ---------------------------
    # Write into Excel template
    # ---------------------------
    template_path = os.path.join(DATA_DIR, "template", "boq", "Template_BOQ_Report.xlsx")
    output_path = os.path.join(export_dir, "BOQ Report.xlsx")

    if not os.path.exists(template_path):
        raise ValueError("BOQ template file not found in template directory.")

    shutil.copy2(template_path, output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        if boq_df.empty:
            logger.info("❌ No BOQ records to write.")
        else:
            boq_df = boq_df.reset_index(drop=True)
            # excel_styler(boq_df).to_excel(writer, sheet_name="Pivot Data", index=False)
            logger.info(
                f"📊 Excel sheet 'Pivot Data' written with {len(boq_df):,} records."
            )

    # ---------------------------
    # PROCESS BOQ EXCEL
    # ---------------------------
    wb = load_workbook(output_path)

    # Style
    named_style = NamedStyle(name="BOQ Row")
    side_style = Side(style="thin", border_style="thin")
    border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
    named_style.font = Font(name="Arial", size=11)
    named_style.border = border

    if "BOQ Row" not in [s for s in wb.named_styles]:
        wb.add_named_style(named_style)

    # BOQ Sheet
    boq_sheet = wb["BOQ"]
    start_data_row = 8
    col_index = {col: num for num, col in enumerate(boq_df.columns, start=1)}
    for idx, record in enumerate(boq_df.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = boq_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "BOQ Row"

            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # BOQ PR Sheet
    boq_pr_sheet = wb["BOQ PR"]
    start_data_row = 3
    col_index = {col: num for num, col in enumerate(boq_pr_df.columns, start=1)}
    for idx, record in enumerate(boq_pr_df.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = boq_pr_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "BOQ Row"

            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    logger.info("✅ BOQ Excel saved.")


if __name__ == "__main__":
    kmz_path = r"D:\JACOBS\PROJECT\TASK\2026\FEB\W1\DRM FORMAT\20250716-H2B2NewSiteCoverage-TBG-v9 (BoQ).kmz"
    export_dir = r"D:\JACOBS\PROJECT\TASK\2026\FEB\W1\DRM FORMAT\20250716-H2B2NewSiteCoverage-TBG-v9"
    sep = "-"

    os.makedirs(export_dir, exist_ok=True)
    drm_format(kmz_path=kmz_path, export_dir=export_dir, sep=sep)