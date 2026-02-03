# ===============
# ROUTE EXTRACTOR
# ===============
import os
import sys
import time
import geopandas as gpd
import pandas as pd
import numpy as np
import shapely
import simplekml
import zipfile
import shutil
import math

from tqdm import tqdm
from datetime import datetime
from shapely.strtree import STRtree
from shapely.geometry import Point, LineString, MultiLineString
from shapely.ops import nearest_points
from shapely.ops import linemerge
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import NamedStyle, Border, Side, Font, Alignment
from enum import Enum

sys.path.append(r"D:\JACOBS\SERVICE\API")

from modules.kml import export_kml, sanitize_kml, validate_kmz_design, validate_kmz_ipl
from modules.table import excel_styler
from modules.utils import auto_group, admin_information
from modules.geometry import relative_intersection, geodesic_length
from core.config import settings

MAINDATA_DIR = settings.MAINDATA_DIR
DATA_DIR = settings.DATA_DIR

# ------------------------------------------------------
# LOGGER
# ------------------------------------------------------
from core.logger import create_logger

logger = create_logger(__file__)


# ENUMS
# ------------------------------------------------------
class Operator(str, Enum):
    IOH = "ioh"
    XL = "xl"
    SURGE = "surge"
    TSEL = "tsel"


class DeviceType(str, Enum):
    OTB = "OTB"
    ODP = "ODP"


class ConnectorType(str, Enum):
    SC = "SC"
    FC = "FC"

class BoQType(str, Enum):
    INTERSITE = "intersite"
    MMP = "mmp"


def detect_turn(
    nodes_gdf: gpd.GeoDataFrame,
    edges_gdf: gpd.GeoDataFrame,
    angle_thresh_deg: float = 150.0,
):
    """
    Faster turn detection using only topology and angles.

    - No buffers
    - No gpd.overlay
    - Classification:
        * turn_isec >= 3  -> 'branch'
        * turn_isec == 2 & angle < angle_thresh -> 'turn'
        * else -> 'straight'
    """
    nodes = nodes_gdf.copy().reset_index(drop=True)
    nodes["turn_note"] = "straight"
    nodes["turn_isec"] = 0
    nodes["turn_ratio"] = 1.0  # not really used anymore
    nodes["area_count"] = 1

    # build map: node_id -> list of direction vectors (dx, dy)
    node_dirs = {nid: [] for nid in nodes["node_id"]}

    def edge_dirs_at_node(edge_geom, node_point: Point):
        coords = list(edge_geom.coords)
        d0 = node_point.distance(Point(coords[0]))
        d1 = node_point.distance(Point(coords[-1]))
        if d0 <= d1:
            # direction from node to next coord
            if len(coords) >= 2:
                dx = coords[1][0] - coords[0][0]
                dy = coords[1][1] - coords[0][1]
            else:
                dx = dy = 0.0
        else:
            # direction from node to previous coord
            if len(coords) >= 2:
                dx = coords[-2][0] - coords[-1][0]
                dy = coords[-2][1] - coords[-1][1]
            else:
                dx = dy = 0.0
        return (dx, dy)

    # build a quick lookup for node geometries
    node_geom_map = dict(zip(nodes["node_id"], nodes.geometry))

    # iterate edges, add direction vectors to node_dirs
    for _, e in edges_gdf.iterrows():
        for side in ("node_start", "node_end"):
            nid = e[side]
            if nid not in node_dirs:
                continue
            node_pt = node_geom_map.get(nid)
            if node_pt is None:
                continue
            dx, dy = edge_dirs_at_node(e.geometry, node_pt)
            norm = np.hypot(dx, dy)
            if norm == 0:
                continue
            node_dirs[nid].append((dx / norm, dy / norm))

    # classify each node
    for idx, row in nodes.iterrows():
        nid = row["node_id"]
        dirs = node_dirs.get(nid, [])
        k = len(dirs)

        if k == 0 or k == 1:
            # isolated or dead-end
            nodes.at[idx, "turn_isec"] = k
            nodes.at[idx, "turn_note"] = "straight"
            continue

        if k >= 3:
            nodes.at[idx, "turn_isec"] = k
            nodes.at[idx, "turn_note"] = "branch"
            continue

        # k == 2 -> check angle between two vectors
        (dx1, dy1), (dx2, dy2) = dirs[:2]
        dot = dx1 * dx2 + dy1 * dy2
        dot = max(-1.0, min(1.0, dot))
        angle_rad = np.arccos(dot)
        angle_deg = np.degrees(angle_rad)

        # if angle close to 180° => straight, else turn
        # angle here is the "inside" angle; 180 ~ straight, 90 ~ corner
        if angle_deg >= angle_thresh_deg:
            nodes.at[idx, "turn_isec"] = 0
            nodes.at[idx, "turn_note"] = "straight"
        else:
            nodes.at[idx, "turn_isec"] = 2
            nodes.at[idx, "turn_note"] = "turn"

    return nodes


def route_preprocess(gdf: gpd.GeoDataFrame, decimals: int = 12):
    """
    Generate nodes and edges (u, v) from LineString/MultiLineString geometries.
    Each vertex is treated as a node.
    Automatically snaps close endpoints (within `tol`) across different lines.
    """
    # --- VALIDATE GEOMETRY ---
    geom_types = gdf.geom_type.unique().tolist()
    invalid = [gt for gt in geom_types if gt not in ["LineString", "MultiLineString"]]
    if invalid:
        raise ValueError(f"Unsupported geometry types: {invalid}")

    # --- PREPARE DATA ---
    crs_input = gdf.crs
    gdf["id_line"] = gdf.index + 1
    gdf = gdf.explode(ignore_index=True)
    gdf = gdf.drop_duplicates(subset="geometry").reset_index(drop=True)

    # --- EXTRACT EDGES ---
    edges = []
    for _, row in gdf.iterrows():
        geom = row.geometry
        if geom.is_empty:
            continue
        lines = [geom] if geom.geom_type == "LineString" else geom.geoms
        for line in lines:
            coords = list(line.coords)
            for i in range(len(coords) - 1):
                u = Point(round(coords[i][0], decimals), round(coords[i][1], decimals))
                v = Point(
                    round(coords[i + 1][0], decimals), round(coords[i + 1][1], decimals)
                )
                edges.append(
                    {
                        "id_line": row["id_line"],
                        "geometry": LineString([u, v]),
                        "u": u,
                        "v": v,
                        **{k: v for k, v in row.items()},
                    }
                )
    edges_gdf = gpd.GeoDataFrame(edges, geometry="geometry", crs=gdf.crs)

    # --- BUILD NODES ---
    nodes = []
    for _, e in edges_gdf.iterrows():
        nodes.append({"id_line": e["id_line"], "geometry": e["u"]})
        nodes.append({"id_line": e["id_line"], "geometry": e["v"]})
    nodes_gdf = gpd.GeoDataFrame(nodes, geometry="geometry", crs=gdf.crs)

    nodes_gdf["x"] = nodes_gdf.geometry.x.round(decimals)
    nodes_gdf["y"] = nodes_gdf.geometry.y.round(decimals)
    nodes_gdf["coord_key"] = list(zip(nodes_gdf["x"], nodes_gdf["y"]))
    node_counts = nodes_gdf.groupby("coord_key").size().rename("count")
    nodes_gdf = (
        nodes_gdf.drop_duplicates("coord_key")
        .merge(node_counts, left_on="coord_key", right_index=True, how="left")
        .reset_index(drop=True)
    )

    # --- MAP NODE IDs ---
    nodes_gdf["node_id"] = [f"N{i+1:07d}" for i in range(len(nodes_gdf))]

    def find_node(pt):
        key = (round(pt.x, decimals), round(pt.y, decimals))
        match = nodes_gdf[nodes_gdf["coord_key"] == key]
        return match["node_id"].values[0] if not match.empty else None

    edges_gdf["node_start"] = edges_gdf["u"].apply(find_node)
    edges_gdf["node_end"] = edges_gdf["v"].apply(find_node)
    edges_gdf["length"] = edges_gdf.geometry.length

    # --- TURN ---
    nodes_gdf = detect_turn(nodes_gdf, edges_gdf, angle_thresh_deg=150)

    # --- CLEAN OUTPUT ---
    edges_gdf = edges_gdf.drop(columns=["u", "v"])
    nodes_gdf = nodes_gdf[
        [
            "node_id",
            "x",
            "y",
            "count",
            "turn_isec",
            "turn_ratio",
            "turn_note",
            "geometry",
        ]
    ]

    # CRS
    nodes_gdf = nodes_gdf.to_crs(crs_input)
    edges_gdf = edges_gdf.to_crs(crs_input)

    return nodes_gdf, edges_gdf


def snap_geom(g1: LineString | MultiLineString, g2: shapely.Geometry, threshold: float):
    from shapely.ops import nearest_points

    coordinates = []
    geom_type = g1.geom_type
    if geom_type == "LineString":
        for x, y in g1.coords:
            point = Point(x, y)
            p1, p2 = nearest_points(point, g2)
            if p1.distance(p2) <= threshold:
                coordinates.append(p2.coords[0])
            else:
                coordinates.append((x, y))
    elif geom_type == "MultiLineString":
        geoms = list(g1.geoms)
        for geom in geoms:
            for x, y in geom.coords:
                point = Point(x, y)
                p1, p2 = nearest_points(point, g2)
                if p1.distance(p2) <= threshold:
                    coordinates.append(p2.coords[0])
                else:
                    coordinates.append((x, y))
    return LineString(coordinates)


def substring_overlay(
    source_gdf: gpd.GeoDataFrame, ref_gdf: gpd.GeoDataFrame
) -> gpd.GeoDataFrame:
    import geopandas as gpd
    from shapely.ops import substring
    from shapely.geometry import Point, LineString, MultiLineString, MultiPoint

    """
    For each source line, extract the full segment that lies between its first and last
    intersection with any reference line. Works with LineString/MultiLineString and
    both point and overlap intersections.

    Returns a GeoDataFrame of substring segments with source attributes + index_right.
    """
    if source_gdf.crs != ref_gdf.crs:
        ref_gdf = ref_gdf.to_crs(epsg=3857)
    if source_gdf.crs.to_epsg() != 3857:
        source_gdf = source_gdf.to_crs(3857)

    # CANDIDATE
    candidates = gpd.sjoin(source_gdf, ref_gdf, how="inner", predicate="intersects")
    candidates["length"] = candidates.geometry.length
    candidates = candidates.sort_values("length", ascending=False)
    candidates = candidates.drop_duplicates("geometry")
    if candidates.empty:
        return gpd.GeoDataFrame(columns=list(source_gdf.columns), crs=source_gdf.crs)

    out_rows = []
    ref_dict = ref_gdf.geometry.to_dict()

    for src_idx, row in candidates.iterrows():
        src_geom = row.geometry
        src_length = src_geom.length
        ref_idx = row["index_right"]
        ref_geom = ref_dict[ref_idx]

        # inter = shapely.intersection(src_geom, ref_geom)
        # inter_length = inter.length
        # length_ratio = inter_length / src_length

        # if inter.is_empty:
        #     continue

        # anchors = []
        # # POINT
        # if isinstance(inter, Point):
        #     anchors.append(inter)
        # elif isinstance(inter, MultiPoint):
        #     anchors.extend(list(inter.geoms))

        # # LINESTRING
        # if isinstance(inter, LineString):
        #     coords = list(inter.coords)
        #     if len(coords) >= 2:
        #         anchors.append(Point(coords[0]))
        #         anchors.append(Point(coords[-1]))
        # elif isinstance(inter, MultiLineString):
        #     for seg in inter.geoms:
        #         coords = list(seg.coords)
        #         if len(coords) >= 2:
        #             anchors.append(Point(coords[0]))
        #             anchors.append(Point(coords[-1]))

        # if len(anchors) < 2:
        #     continue

        # # ANCHOR MAPPING DISTANCE
        # dists = [src_geom.project(pt) for pt in anchors]
        # start_d, end_d = min(dists), max(dists)
        # if end_d - start_d <= 0:
        #     continue

        # # SEGMENT
        # seg = substring(src_geom, start_d, end_d)
        seg, new_line = relative_intersection(src_geom, ref_geom, tolerance=20)

        # DIFF
        diff = shapely.difference(src_geom, ref_geom)
        if isinstance(diff, MultiLineString):
            for geom in diff.geoms:
                length = geom.length
                if length > 1000:
                    seg = shapely.difference(seg, geom)
        elif isinstance(diff, LineString):
            length = diff.length
            if length > 1000:
                seg = shapely.difference(seg, diff)

        if seg.is_empty or seg.length <= 0:
            continue

        attrs = {k: v for k, v in row.items() if k != "geometry"}
        attrs["geometry"] = seg
        out_rows.append(attrs)

    if not out_rows:
        return gpd.GeoDataFrame(columns=list(source_gdf.columns), crs=source_gdf.crs)

    result = gpd.GeoDataFrame(out_rows, crs=source_gdf.crs)
    result = result.explode(ignore_index=True)
    logger.info(f"🟢 Substring overlay success")
    return result


def obstacle_detection(lines_gdf: gpd.GeoDataFrame, sep="-"):
    lines_gdf = lines_gdf.copy()
    lines_gdf["line_id"] = lines_gdf.index

    ring_name = lines_gdf["ring_name"].mode()[0]
    osm_railway = gpd.read_parquet(
        f"{MAINDATA_DIR}/03. Road Network/railway_osm.parquet"
    )
    osm_toll = gpd.read_parquet(f"{MAINDATA_DIR}/03. Road Network/toll.parquet")
    osm_railway = osm_railway.rename(columns={"name": "rail_name"})
    osm_toll = osm_toll.rename(columns={"name": "toll_name"})

    lines_gdf = lines_gdf.to_crs(epsg=3857)
    osm_railway = osm_railway.to_crs(epsg=3857)
    osm_toll = osm_toll.to_crs(epsg=3857)

    union_lines = lines_gdf.geometry.union_all().buffer(20)
    osm_railway = osm_railway[osm_railway.geometry.intersects(union_lines)].copy()
    osm_toll = osm_toll[osm_toll.geometry.intersects(union_lines)].copy()

    # RAILWAY
    if not osm_railway.empty:
        isec_rail = gpd.overlay(
            lines_gdf,
            osm_railway[["rail_name", "geometry"]],
            how="intersection",
            keep_geom_type=False,
        )
        isec_rail["geometry"] = isec_rail.geometry.representative_point()
        isec_rail["remark"] = isec_rail["near_end"] + sep + isec_rail["far_end"]
        isec_rail["obstacle_railway"] = isec_rail.geometry.to_wkt()

        group = auto_group(isec_rail, distance=100)
        group = group.rename(columns={"region": "group"})
        isec_rail = gpd.sjoin(isec_rail, group[["group", "geometry"]]).drop(
            columns="index_right"
        )
        isec_rail = isec_rail.drop_duplicates(subset="group")
        logger.info(f"🟠 Found {len(isec_rail)} intersect with railway")
        lines_gdf = lines_gdf.merge(
            isec_rail[["line_id", "rail_name", "obstacle_railway"]],
            how="left",
            on="line_id",
        )
    else:
        lines_gdf["obstacle_railway"] = None
        logger.info(f"🟢 No railway obstacle in {ring_name}.")

    # TOLL
    if not osm_toll.empty:
        isec_toll = gpd.overlay(
            lines_gdf,
            osm_toll[["toll_name", "geometry"]],
            how="intersection",
            keep_geom_type=False,
        )
        isec_toll["geometry"] = isec_toll.geometry.representative_point()
        isec_toll["remark"] = isec_toll["near_end"] + sep + isec_toll["far_end"]
        isec_toll["obstacle_toll"] = isec_toll.geometry.to_wkt()

        group = auto_group(isec_toll, distance=100)
        group = group.rename(columns={"region": "group"})
        isec_toll = gpd.sjoin(isec_toll, group[["group", "geometry"]]).drop(
            columns="index_right"
        )
        isec_toll = isec_toll.drop_duplicates(subset="group")
        logger.info(f"🟠 Found {len(isec_toll)} intersect with highway")
        lines_gdf = lines_gdf.merge(
            isec_toll[["line_id", "toll_name", "obstacle_toll"]],
            how="left",
            left_index=True,
            right_index=True,
        )
    else:
        lines_gdf["obstacle_toll"] = None
        logger.info(f"🟢 No highway obstacle in {ring_name}.")

    # # JOIN ADMIN
    # admin_2024 = admin_2024.to_crs(epsg=3857)
    # clean_col = ['kabkot', 'provinsi', 'kecamatan', 'desa']
    # intersected.columns = intersected.columns.str.lower()

    # for col in clean_col:
    #     if col in intersected.columns:
    #         intersected = intersected.drop(columns=col)
    # intersected = gpd.sjoin(intersected, admin_2024).drop(columns='index_right')

    # # EXPORT
    # intersected = intersected.to_crs(epsg=4326)
    # intersected['long'] = intersected.geometry.x
    # intersected['lat'] = intersected.geometry.y
    # intersected.to_parquet(fr"{export_dir}\Intersect Railway.parquet")

    return lines_gdf


def bill_of_quantity(
    points: gpd.GeoDataFrame, lines: gpd.GeoDataFrame, sep="-", operator: str = None
):
    import shapely
    from shapely.ops import split, snap, linemerge
    from shapely.geometry import LineString
    import geopandas as gpd
    import numpy as np
    import pandas as pd

    # =============================
    # LOAD FO REFERENCE GEOMETRY
    # =============================
    fo_route_path = rf"{DATA_DIR}/FO TBG Only_01062025.parquet"
    fo_route = gpd.read_parquet(fo_route_path)
    fo_route = fo_route.to_crs(epsg=3857)
    fo_route.columns = fo_route.columns.str.lower()
    fo_route = fo_route.rename(columns={"name": "fiber"})

    # NORMALIZE OPERATOR
    operator = operator.lower().strip()
    valid_operator = ["ioh", "tsel", "xl", "surge"]
    if operator not in valid_operator:
        raise ValueError(
            f"Invalid operator value. Should be {(",").join(valid_operator)} instead of '{operator}'."
        )

    # =============================
    # PREPARE INPUT DATA
    # =============================
    points = points.copy().to_crs(epsg=3857)
    lines = lines.copy().to_crs(epsg=3857)

    if points.empty:
        raise ValueError("Points data is empty.")
    if lines.empty:
        raise ValueError("Lines data is empty.")

    # safe ring name
    if "ring_name" in points.columns and not points["ring_name"].dropna().empty:
        ring_name = points["ring_name"].mode().iloc[0]
    else:
        ring_name = "Unknown Ring"

    logger.info(f"🌏 {ring_name} BOQ running ...")
    lines["id_line"] = lines.index + 1

    # =============================
    # ROUTE PREPROCESS
    # =============================
    nodes, edges = route_preprocess(lines)
    nodes = nodes.to_crs(epsg=3857)
    edges = edges.to_crs(epsg=3857)

    # =============================
    # IDENTIFY TURN / BRANCH POINTS
    # =============================
    turn_data = nodes[nodes["turn_note"].str.contains("turn|branch", case=False, na=False)].copy()
    turn_data = turn_data.rename(columns={"node_id": "turn_id"})
    branch = turn_data[turn_data["turn_isec"] > 2].copy()
    branch = branch.rename(columns={"turn_id": "branch_id"})

    # nearest joins
    points = gpd.sjoin_nearest(
        points,
        nodes[["geometry", "node_id", "turn_ratio"]],
        how="left",
        exclusive=True,
    ).drop(columns="index_right")

    points = gpd.sjoin_nearest(
        points,
        turn_data[["geometry", "turn_id"]],
        how="left",
        distance_col="dist_turn",
        exclusive=True,
    ).drop(columns="index_right")

    if not branch.empty:
        points = gpd.sjoin_nearest(
            points,
            branch[["geometry", "branch_id"]],
            how="left",
            distance_col="dist_branch",
            exclusive=True,
            max_distance=500,
        ).drop(columns="index_right")
    else:
        points["branch_id"] = np.nan
        points["dist_branch"] = -1

    points["dist_turn"] = points["dist_turn"].fillna(-1)
    points["dist_branch"] = points["dist_branch"].fillna(-1)

    node_geom_map = {
        nid: geom.wkt for nid, geom in zip(nodes["node_id"], nodes.geometry)
    }
    turn_geom_map = {
        tid: geom.wkt for tid, geom in zip(turn_data["turn_id"], turn_data.geometry)
    }
    branch_geom_map = {
        bid: geom for bid, geom in zip(branch["branch_id"], branch.geometry)
    }

    # Assign OTB and ODP
    points["otb"] = points.geometry.to_wkt()
    points["odp"] = points["turn_id"].map(turn_geom_map)
    points["otb_type"] = 24
    points["odp_type"] = 24

    mask_no_odp = points["odp"].isna()
    points.loc[mask_no_odp & points["node_id"].notna(), "odp"] = (
        points.loc[mask_no_odp & points["node_id"].notna(), "node_id"]
        .map(node_geom_map)
    )

    for idx, row in points.iterrows():
        node_id = row["node_id"]
        turn_id = row["turn_id"]
        site_type = row["site_type"]
        dist_turn = row["dist_turn"]
        branch_id = row["branch_id"]
        dist_branch = row["dist_branch"]
        branch_ratio = row.get("turn_ratio", 1.0)

        # -- No nearest turn
        if (pd.isna(turn_id) or dist_turn < 0 or dist_turn > 500) and pd.notna(node_id):
            points.at[idx, "odp"] = node_geom_map[node_id]

        # -- Nearest branch if exist and ratio < 0.5
        if dist_branch > 0 and branch_ratio < 0.5 and pd.notna(branch_id):
            geom_branch = branch_geom_map.get(branch_id)
            if geom_branch is not None:
                points.at[idx, "odp"] = geom_branch.wkt

        # IOH Hub Core 48
        if operator == "ioh" and ("hub" in str(site_type).lower()):
            points.at[idx, "odp_type"] = 48

    # =============================
    # IDENTIFY ROUTE (ACCESS + BACKBONE)
    # =============================
    points_idx = points.set_index(points["site_id"].astype(str), drop=False)
    lines["cable_type"] = 24

    for idx, row in lines.iterrows():
        line_geom = row.geometry
        line_geom = (
            linemerge(line_geom)
            if line_geom.geom_type == "MultiLineString"
            else line_geom
        )

        ne = str(row["near_end"]).strip()
        fe = str(row["far_end"]).strip()

        ne_row = points_idx.loc[ne] if ne in points_idx.index else None
        fe_row = points_idx.loc[fe] if fe in points_idx.index else None

        # Near End missing Far End found
        if ne_row is None and fe_row is not None:
            logger.warning(
                f"⚠️ NE '{ne}' not found but FE '{fe}' found; "
                f"treating FE as NE (possible reversed) for line idx={idx}."
            )
            ne, fe = fe, ne
            ne_row, fe_row = fe_row, None  # FE becomes "ring" / non-site

        # Both Missing
        if ne_row is None and fe_row is None:
            logger.error(
                f"❌ Neither NE '{ne}' nor FE '{fe}' found in points for line idx={idx}. "
                f"Using full line as backbone."
            )
            lines.at[idx, "backbone"] = line_geom.wkt
            lines.at[idx, "access_ne"] = None
            lines.at[idx, "access_fe"] = None
            continue

        # Duplicate > First row
        if isinstance(ne_row, gpd.GeoDataFrame):
            ne_row = ne_row.iloc[0]
        if isinstance(fe_row, gpd.GeoDataFrame):
            fe_row = fe_row.iloc[0]

        # types & odp
        ne_type = str(ne_row.get("site_type", "")).lower() if ne_row is not None else ""
        if fe_row is not None:
            fe_type = str(fe_row.get("site_type", "")).lower()
        else:
            fe_type = "hub"

        odp_ne_wkt = ne_row.get("odp") if ne_row is not None else None
        odp_fe_wkt = fe_row.get("odp") if fe_row is not None else None

        access_ne = None
        access_fe = None
        lines.at[idx, "access_ne"] = None
        lines.at[idx, "access_fe"] = None

        # --- NEAR END ---
        if isinstance(odp_ne_wkt, str) and "hub" not in ne_type:
            odp_ne_geom = shapely.from_wkt(odp_ne_wkt)
            odp_ne_geom = snap(odp_ne_geom, line_geom, tolerance=5)

            splitted_ne = split(line_geom, odp_ne_geom)
            geoms = list(splitted_ne.geoms)

            if len(geoms) > 1:
                ne_pt = ne_row.geometry
                access_ne = min(geoms, key=lambda g: g.distance(ne_pt))
                lines.at[idx, "access_ne"] = access_ne.wkt


        # --- FAR END ---
        if isinstance(odp_fe_wkt, str) and "hub" not in fe_type:
            odp_fe_geom = shapely.from_wkt(odp_fe_wkt)
            odp_fe_geom = snap(odp_fe_geom, line_geom, tolerance=5)

            splitted_fe = split(line_geom, odp_fe_geom)
            geoms = list(splitted_fe.geoms)

            if len(geoms) > 1:
                fe_pt = fe_row.geometry
                access_fe = min(geoms, key=lambda g: g.distance(fe_pt))
                lines.at[idx, "access_fe"] = access_fe.wkt


        # --- BACKBONE ---
        backbone = line_geom
        if access_ne:
            backbone = shapely.difference(backbone, access_ne)
        if access_fe:
            backbone = shapely.difference(backbone, access_fe)
        lines.at[idx, "backbone"] = backbone.wkt

    # =============================
    # IDENTIFY EXISTING FO ROUTES
    # =============================
    union_lines = lines.geometry.union_all().buffer(30)
    fo_route_clip = fo_route[fo_route.geometry.intersects(union_lines)].copy()
    fo_route_clip["geometry"] = fo_route_clip.geometry.buffer(30)

    _ = substring_overlay(lines, fo_route_clip)
    existing_route = gpd.overlay(
        lines, fo_route_clip, how="intersection", keep_geom_type=True
    )

    lines["fo_exist"] = [{} for _ in range(len(lines))]
    lines["pole_exist"] = [{} for _ in range(len(lines))]
    lines["closure"] = [{} for _ in range(len(lines))]

    if existing_route.empty:
        logger.info("⚠️ No existing FO intersections found.")
        lines = obstacle_detection(lines, sep=sep)
        return points, lines

    existing_route = existing_route[["id_line", "fiber", "geometry"]].reset_index(
        drop=True
    )
    existing_route = existing_route.dissolve(["id_line", "fiber"]).reset_index()
    existing_route["geometry"] = existing_route.geometry.apply(
        lambda g: linemerge(g) if g.geom_type == "MultiLineString" else g
    )
    existing_route["length"] = existing_route.geometry.length
    existing_route = existing_route.sort_values("length", ascending=False)

    dropped = []
    for i, row in existing_route.iterrows():
        if i in dropped:
            continue
        geom = row.geometry.buffer(5)
        within_idx = existing_route[
            (existing_route.index != i) & (existing_route.within(geom))
        ]
        if not within_idx.empty:
            dropped.extend(within_idx.index.to_list())

    if dropped:
        existing_route = existing_route.drop(index=dropped)
        logger.info(f"ℹ️ Dropped {len(dropped)} overlapped lines.")
    existing_route = existing_route.drop_duplicates("geometry").reset_index(drop=True)

    # =============================
    # CLASSIFY EXISTING & POLE EXISTING
    # =============================
    for idx, row in lines.iterrows():
        id_line = row["id_line"]
        backbone = shapely.from_wkt(row["backbone"])
        fo_lines = existing_route[existing_route["id_line"] == id_line].copy()
        fo_exist_dict = {}
        pole_exist_dict = {}
        closure_dict = {}

        for _, fo_row in fo_lines.iterrows():
            fiber_name = fo_row["fiber"]
            fo_geom = fo_row.geometry
            if fo_geom.is_empty:
                continue

            if fo_geom.length > 1000:
                logger.info(f"ℹ️ FO Existing: {fiber_name} | Length: {fo_geom.length}")
                backbone = shapely.difference(backbone, fo_geom)
                closure = shapely.intersection(fo_geom, backbone)
                fo_exist_dict[fiber_name] = fo_geom.wkt

                if not closure.is_empty:
                    logger.info(closure)
                    closure_dict[fiber_name] = closure.wkt
            elif 100 < fo_geom.length < 1000:
                logger.info(f"ℹ️ Pole Existing: {fiber_name} | Length: {fo_geom.length}")
                pole_exist_dict[fiber_name] = fo_geom.wkt
            else:
                continue

        lines.at[idx, "fo_exist"] = fo_exist_dict
        lines.at[idx, "closure"] = closure_dict
        lines.at[idx, "pole_exist"] = pole_exist_dict
        lines.at[idx, "backbone"] = backbone.wkt

    # =============================
    # CLASSIFY OBSTACLE
    # =============================
    lines = obstacle_detection(lines, sep=sep)

    # Surge Preference: Cross KAI use Core 96
    if operator == "surge":
        points_idx = points.set_index(points["site_id"].astype(str), drop=False)
        lines_intersected = lines[(lines["obstacle_railway"].notna())].copy()
        lines.loc[lines.index.isin(lines_intersected.index), "cable_type"] = 96
        for idx, row in lines_intersected.iterrows():
            ne = str(row["near_end"])
            fe = str(row["far_end"])
            mask_odp = points["site_id"].astype(str).isin([ne, fe])
            points.loc[mask_odp, "odp_type"] = 96

    logger.info(f"🟢 {ring_name} BOQ Processing complete.\n")
    return points, lines


def parallel_boq(
    points_gdf: gpd.GeoDataFrame, lines_gdf: gpd.GeoDataFrame, operator: str, **kwargs
):
    ringlist = set(points_gdf["ring_name"])
    task_celery = kwargs.get("task_celery", False)
    sep = kwargs.get("sep", "-")

    if "index_right" in points_gdf.columns:
        points_gdf = points_gdf.drop(columns="index_right")
    if "index_right" in lines_gdf.columns:
        lines_gdf = lines_gdf.drop(columns="index_right")

    with ProcessPoolExecutor(max_workers=4) as executor:
        futures = {}
        for ring in ringlist:
            points_ring = points_gdf[points_gdf["ring_name"] == ring].copy()
            lines_ring = lines_gdf[lines_gdf["ring_name"] == ring].copy()
            future = executor.submit(
                bill_of_quantity, points_ring, lines_ring, sep=sep, operator=operator
            )
            futures[future] = ring

        points_compiled = []
        lines_compiled = []

        for future in tqdm(
            as_completed(futures), total=len(futures), desc="Process BOQ..."
        ):
            ring = futures[future]
            try:
                result = future.result()
                if result:
                    points_result, lines_result = result
                    points_compiled.append(points_result)
                    lines_compiled.append(lines_result)

                    if task_celery:
                        task_celery.update_state(
                            state="PROGRESS",
                            meta={
                                "status": (
                                    f"Completed BOQ for {len(lines_compiled)}/{len(ringlist):,} rings"
                                )
                            },
                        )
            except Exception as e:
                logger.error(f"🔴 Error BOQ in {ring}: {e}")
                continue

        points_compiled = pd.concat(points_compiled)
        lines_compiled = pd.concat(lines_compiled)

    return points_compiled, lines_compiled


def compile_dict(data_gdf: gpd.GeoDataFrame, column: str):
    data_list = []
    for idx, row in data_gdf.iterrows():
        col_data = row[column]

        if not isinstance(col_data, dict):
            continue
        if len(col_data) < 1:
            continue

        for col_name, geom in col_data.items():
            segment = {
                **{k: v for k, v in row.items() if k != "geometry"},
                column: col_name,
                "geometry": shapely.from_wkt(geom),
            }
            data_list.append(segment)

    if data_list:
        data_gdf = gpd.GeoDataFrame(data_list, geometry="geometry", crs=data_gdf.crs)
    else:
        data_gdf = gpd.GeoDataFrame(
            columns=data_gdf.columns, geometry="geometry", crs=data_gdf.crs
        )
    return data_gdf


def identify_connection(
    ring: str,
    target_fiber: gpd.GeoDataFrame,
    target_point: gpd.GeoDataFrame,
    start_column: str = "near_end",
) -> tuple:

    import numpy as np
    import geopandas as gpd

    # --- CRS normalize ---
    if target_fiber.crs != "EPSG:3857":
        target_fiber = target_fiber.to_crs(epsg=3857)
    if target_point.crs != "EPSG:3857":
        target_point = target_point.to_crs(epsg=3857)

    # --- Flatten any list/array values ---
    for col in ["near_end", "far_end"]:
        if col in target_fiber.columns:
            target_fiber[col] = target_fiber[col].apply(
                lambda x: x[0] if isinstance(x, (list, tuple, np.ndarray)) else x
            )

    # --- Validate start column ---
    if start_column == "near_end":
        opposite_column = "far_end"
    elif start_column == "far_end":
        opposite_column = "near_end"
    else:
        raise ValueError("start_column must be either 'near_end' or 'far_end'.")

    # --- Separate hub and site list ---
    fo_hub = target_point[
        target_point["site_type"].str.lower().str.contains("hub")
    ].drop_duplicates("geometry")
    site_list = target_point[
        ~target_point["site_type"].str.lower().str.contains("hub")
    ].drop_duplicates("geometry")

    # --- Identify starting hub ---
    hub_ids = fo_hub["site_id"].astype(str).tolist()
    start_hub = target_fiber[target_fiber[start_column].astype(str).isin(hub_ids)][
        start_column
    ].values
    if len(start_hub) == 0:
        start_hub = target_fiber[
            target_fiber[opposite_column].astype(str).isin(hub_ids)
        ][opposite_column].values
    if len(start_hub) == 0:
        print(fo_hub["site_id"].tolist())
        print(site_list["site_id"].tolist())
        print(hub_ids)
        raise ValueError(f"❌ No FO Hub found in ring {ring}")

    start_hub = start_hub[0]

    # --- Sequential connection search ---
    connection = [start_hub]
    visited = set([start_hub])
    frontier = [start_hub]  # support branching

    while frontier:
        current = frontier.pop(0)

        # find all fiber segments connected to this site
        matches = target_fiber[
            (target_fiber[start_column] == current)
            | (target_fiber[opposite_column] == current)
        ]

        for _, seg in matches.iterrows():
            if seg[start_column] == current:
                next_sites = [seg[opposite_column]]
            else:
                next_sites = [seg[start_column]]

            for next_site in next_sites:
                if next_site not in visited:
                    visited.add(next_site)
                    connection.append(next_site)
                    frontier.append(next_site)

    # --- Build ordered GeoDataFrame of connection points ---
    points_sequential = []
    for site_id in connection:
        site_id = str(site_id)
        if site_id in fo_hub["site_id"].astype(str).values:
            row = fo_hub[fo_hub["site_id"].astype(str) == site_id].iloc[0].to_dict()
        elif site_id in site_list["site_id"].astype(str).values:
            row = (
                site_list[site_list["site_id"].astype(str) == site_id].iloc[0].to_dict()
            )
        else:
            logger.info(f"⚠️ Site {site_id} not found.")
            continue
        points_sequential.append(row)

    if not points_sequential:
        logger.info(f"⚠️ No valid points found for ring {ring}")
        return None, None

    points_sequential = gpd.GeoDataFrame(
        points_sequential, crs="EPSG:3857"
    ).reset_index(drop=True)
    return points_sequential, connection


def auto_sorter(df: pd.DataFrame | gpd.GeoDataFrame, column: str, sort_list: list):
    from itertools import groupby

    sort_list = [key for key, _ in groupby(sort_list)]
    order_map = {str(v): i for i, v in enumerate(sort_list)}
    if not df.empty:
        df["order"] = df[column].astype(str).map(order_map)
        df = df.sort_values("order", na_position="last").drop(columns="order")
    return df


def create_topology(
    points_gdf: gpd.GeoDataFrame, merge: bool = True
) -> gpd.GeoDataFrame:
    if points_gdf.crs != "EPSG:3857":
        points_gdf = points_gdf.to_crs(epsg=3857)

    points_gdf = points_gdf[
        points_gdf.geometry.notnull() & ~points_gdf.geometry.is_empty
    ].copy()
    points_gdf["geometry"] = points_gdf.geometry.force_2d()

    ring_list = points_gdf["ring_name"].unique().tolist()
    topology_records = []

    for ring in ring_list:
        ring_points = points_gdf[points_gdf["ring_name"] == ring].reset_index(drop=True)
        if ring_points.empty:
            continue

        region = next(
            (x for x in ring_points.get("region", []) if pd.notna(x)), "Unknown Region"
        )
        project = next(
            (x for x in ring_points.get("project", []) if pd.notna(x)),
            "Unknown Project",
        )
        fo_hub_count = len(
            ring_points[ring_points["site_type"].str.lower().str.contains("hub")]
        )

        for i in range(len(ring_points)):
            start_point = ring_points.iloc[i]
            end_point = ring_points.iloc[(i + 1) % len(ring_points)]

            # skip bad geometries
            if start_point.geometry is None or end_point.geometry is None:
                logger.info(f"⚠️ Skipping segment in ring {ring}: invalid geometry.")
                continue

            # handle FO hub cases
            match fo_hub_count:
                case 1:
                    pass
                case 2:
                    if (i + 1) % len(ring_points) == 0:
                        continue
                case _:
                    raise ValueError(
                        f"Ring {ring} has {fo_hub_count} FO Hubs, which is not supported."
                    )

            try:
                start_coords = list(start_point.geometry.coords)[0][:2]
                end_coords = list(end_point.geometry.coords)[0][:2]
                line_geom = LineString([start_coords, end_coords])
            except Exception as e:
                logger.info(f"⚠️ Failed to create line in ring {ring}: {e}")
                continue

            record = {
                "name": f"{start_point['site_id']}-{end_point['site_id']}",
                "near_end": start_point["site_id"],
                "far_end": end_point["site_id"],
                "ring_name": ring,
                "region": region,
                "project": project,
                "length": line_geom.length,
                "route_type": "Topology",
                "fo_note": "topology",
                "geometry": line_geom,
            }
            topology_records.append(record)

    if not topology_records:
        logger.info("⚠️ No topology records created.")
        return gpd.GeoDataFrame(
            columns=["geometry"], geometry="geometry", crs="EPSG:3857"
        )

    topology_gdf = gpd.GeoDataFrame(
        topology_records, geometry="geometry", crs="EPSG:3857"
    )

    if merge and not topology_gdf.empty:
        topology_gdf = topology_gdf.dissolve(by="ring_name")
        topology_gdf = topology_gdf[["geometry", "region", "project"]].reset_index()
        topology_gdf["name"] = "Connection"
        topology_gdf["geometry"] = topology_gdf["geometry"].apply(
            lambda geom: (
                linemerge(geom) if geom.geom_type == "MultiLineString" else geom
            )
        )
    return topology_gdf


def compile_boq(
    points_boq: gpd.GeoDataFrame,
    lines_boq: gpd.GeoDataFrame,
    sep: str = "-",
    device_in_branch: str = "ODP",
    device_in_site: str = "OTB",
):

    # BILL OF QUANTITY
    # Lines Based
    lines_boq = lines_boq.copy()
    lines_boq["near_end"] = lines_boq["near_end"].astype(str).str.strip()
    lines_boq["far_end"] = lines_boq["far_end"].astype(str).str.strip()
    lines_boq["segment_name"] = lines_boq["near_end"] + sep + lines_boq["far_end"]
    backbone = lines_boq[
        [
            "segment_name",
            "near_end",
            "far_end",
            "ring_name",
            "backbone",
            "cable_type",
            "geometry",
        ]
    ].copy()
    backbone = backbone.rename(columns={"cable_type": "core"})
    backbone = backbone.dropna(subset=["backbone"])
    if not backbone.empty:
        backbone["geometry"] = backbone["backbone"].apply(
            lambda geom: shapely.from_wkt(geom)
        )
        backbone["segment_name"] = backbone["near_end"] + sep + backbone["far_end"]
        backbone["name"] = "BB " + backbone["segment_name"]
        backbone["name"] = np.where(
            backbone["core"] == 96,
            backbone["name"] + "_FO" + backbone["core"].astype(str) + "C",
            backbone["name"],
        )
        backbone["geometry"] = backbone["geometry"].apply(
            lambda geom: (
                linemerge(geom) if geom.geom_type == "MultiLineString" else geom
            )
        )
        backbone = backbone.drop(columns="backbone")
        backbone = backbone.to_crs(epsg=4326)

    access_ne = lines_boq[
        ["segment_name", "near_end", "far_end", "ring_name", "access_ne", "geometry"]
    ].copy()
    access_ne = access_ne.dropna(subset=["access_ne"])
    if not access_ne.empty:
        access_ne["geometry"] = access_ne["access_ne"].apply(
            lambda geom: shapely.from_wkt(geom)
        )
        access_ne["segment_name"] = access_ne["near_end"] + sep + access_ne["far_end"]
        access_ne["name"] = "Akses " + access_ne["segment_name"]
        access_ne["geometry"] = access_ne["geometry"].apply(
            lambda geom: (
                linemerge(geom) if geom.geom_type == "MultiLineString" else geom
            )
        )
        access_ne = access_ne.drop(columns="access_ne")
        access_ne = access_ne.to_crs(epsg=4326)

    access_fe = lines_boq[
        ["segment_name", "near_end", "far_end", "ring_name", "access_fe", "geometry"]
    ].copy()
    access_fe = access_fe.dropna(subset=["access_fe"])
    if not access_fe.empty:
        access_fe["geometry"] = access_fe["access_fe"].apply(
            lambda geom: shapely.from_wkt(geom)
        )
        access_fe["segment_name"] = access_fe["near_end"] + sep + access_fe["far_end"]
        access_fe["name"] = "Akses " + access_fe["segment_name"]
        access_fe["geometry"] = access_fe["geometry"].apply(
            lambda geom: (
                linemerge(geom) if geom.geom_type == "MultiLineString" else geom
            )
        )
        access_fe = access_fe.drop(columns="access_fe")
        access_fe = access_fe.to_crs(epsg=4326)

    # Points Based
    odp = points_boq[
        ["site_id", "site_type", "ring_name", "odp", "odp_type", "geometry"]
    ].copy()
    odp = odp.rename(columns={"odp_type": "core"})
    odp = odp.dropna(subset=["odp"])

    if not odp.empty:
        odp["geometry"] = odp["odp"].map(shapely.from_wkt)
        odp["name"] = f"{device_in_branch} " + odp["site_id"].astype(str)

        mask_core = odp["core"].isin([48, 96])
        odp.loc[mask_core, "name"] = (
            f"{device_in_branch}_"
            + odp.loc[mask_core, "core"].astype(str)
            + " "
            + odp.loc[mask_core, "site_id"].astype(str)
        )
        odp["ext_note"] = odp["name"].duplicated().astype(int)
        mask_ext = odp["ext_note"].eq(1)
        odp.loc[mask_ext, "name"] = odp.loc[mask_ext, "name"].str.replace(
            r"^(?P<device>ODP|OTB|Closure)(?P<core>_\d{2})? (?P<site_id>\w+)$",
            r"\g<device>\g<core>_EXT \g<site_id>",
            regex=True,
        )

        odp = odp.drop(columns="odp")
        odp = odp.to_crs(epsg=4326)
        odp["long"] = odp.geometry.x
        odp["lat"] = odp.geometry.y

    otb = points_boq[
        ["site_id", "site_type", "ring_name", "otb", "otb_type", "geometry"]
    ].copy()
    otb = otb.rename(columns={"otb_type": "core"})
    otb = otb.dropna(subset=["otb"])
    if not otb.empty:
        otb["geometry"] = otb["otb"].apply(lambda geom: shapely.from_wkt(geom))
        otb["name"] = f"{device_in_site} " + otb["site_id"]
        mask_core = otb["core"].isin([48, 96])
        otb.loc[mask_core, "name"] = (
            f"{device_in_site}_"
            + otb.loc[mask_core, "core"].astype(str)
            + " "
            + otb.loc[mask_core, "site_id"].astype(str)
        )
        otb["ext_note"] = otb["name"].duplicated().astype(int)
        mask_ext = otb["ext_note"].eq(1)
        otb.loc[mask_ext, "name"] = otb.loc[mask_ext, "name"].str.replace(
            r"^(?P<device>ODP|OTB|Closure)(?P<core>_\d{2})? (?P<site_id>\w+)$",
            r"\g<device>\g<core>_EXT \g<site_id>",
            regex=True,
        )
        otb = otb.drop(columns="otb")
        otb = otb.to_crs(epsg=4326)
        otb["long"] = otb.geometry.x
        otb["lat"] = otb.geometry.y

    # DIVIDED BY INTERSECTION FIBER
    fo_exist = lines_boq[
        ["segment_name", "near_end", "far_end", "ring_name", "fo_exist", "geometry"]
    ].copy()
    fo_exist = compile_dict(fo_exist, "fo_exist")
    if not fo_exist.empty:
        fo_exist["name"] = (
            fo_exist["segment_name"] + "/" + fo_exist["fo_exist"].astype(str)
        )
        fo_exist["geometry"] = fo_exist["geometry"].apply(
            lambda geom: (
                linemerge(geom) if geom.geom_type == "MultiLineString" else geom
            )
        )
        fo_exist = fo_exist.to_crs(epsg=4326)

    pole_exist = lines_boq[
        ["segment_name", "near_end", "far_end", "ring_name", "pole_exist", "geometry"]
    ].copy()
    pole_exist = compile_dict(pole_exist, "pole_exist")
    if not pole_exist.empty:
        pole_exist["name"] = pole_exist["segment_name"] + "/POLE EXT"
        pole_exist["geometry"] = pole_exist["geometry"].apply(
            lambda geom: (
                linemerge(geom) if geom.geom_type == "MultiLineString" else geom
            )
        )
        pole_exist = pole_exist.to_crs(epsg=4326)

    closure = lines_boq[
        [
            "segment_name",
            "near_end",
            "far_end",
            "ring_name",
            "closure",
            "cable_type",
            "geometry",
        ]
    ].copy()
    closure = closure.rename(columns={"cable_type": "core"})
    closure = compile_dict(closure, "closure")
    closure = closure.to_crs(epsg=4326)
    if not closure.empty:
        closure = closure.explode(ignore_index=True)
        closure["name"] = "Closure " + closure["segment_name"]
        closure["ext_note"] = np.where(closure[["name", "geometry"]].duplicated(), 1, 0)
        mask_ext = closure["ext_note"].eq(1)
        closure.loc[mask_ext, "name"] = (
            "Closure_EXT " + closure.loc[mask_ext, "segment_name"]
        )
        closure["long"] = closure.geometry.x
        closure["lat"] = closure.geometry.y

    obstacle_railway = lines_boq[
        [
            "segment_name",
            "near_end",
            "far_end",
            "ring_name",
            "obstacle_railway",
            "geometry",
        ]
    ].copy()
    obstacle_railway = obstacle_railway.dropna(subset=["obstacle_railway"])
    if not obstacle_railway.empty:
        obstacle_railway["geometry"] = obstacle_railway["obstacle_railway"].apply(
            lambda geom: shapely.from_wkt(geom)
        )
        obstacle_railway = obstacle_railway.drop(columns="obstacle_railway")
        obstacle_railway = obstacle_railway.to_crs(epsg=4326)
        obstacle_railway["long"] = obstacle_railway.geometry.x
        obstacle_railway["lat"] = obstacle_railway.geometry.y
        obstacle_railway["name"] = "Obstacle Rail " + obstacle_railway["segment_name"]

    obstacle_toll = lines_boq[
        [
            "segment_name",
            "near_end",
            "far_end",
            "ring_name",
            "obstacle_toll",
            "geometry",
        ]
    ].copy()
    obstacle_toll = obstacle_toll.dropna(subset=["obstacle_toll"])
    if not obstacle_toll.empty:
        obstacle_toll["geometry"] = obstacle_toll["obstacle_toll"].apply(
            lambda geom: shapely.from_wkt(geom)
        )
        obstacle_toll = obstacle_toll.drop(columns="obstacle_toll")
        obstacle_toll = obstacle_toll.to_crs(epsg=4326)
        obstacle_toll["name"] = "Obstacle Toll " + obstacle_toll["segment_name"]
        obstacle_toll["long"] = obstacle_toll.geometry.x
        obstacle_toll["lat"] = obstacle_toll.geometry.y
    return (
        odp,
        otb,
        closure,
        backbone,
        access_ne,
        access_fe,
        fo_exist,
        pole_exist,
        obstacle_railway,
        obstacle_toll,
    )


def excel_boq(
    points_boq: gpd.GeoDataFrame,
    lines_boq: gpd.GeoDataFrame,
    export_dir: str,
    device_in_site: str = "OTB",
    device_in_branch: str = "ODP",
    **kwargs,
):
    program = kwargs.get("program", "N/A")
    vendor = kwargs.get("vendor", "TBG")
    sep = kwargs.get("sep", ";")

    lines_boq = lines_boq.copy()
    points_boq = points_boq.copy()

    if "long" not in points_boq.columns or "lat" not in points_boq.columns:
        points_boq["long"] = points_boq.geometry.to_crs(epsg=4326).x
        points_boq["lat"] = points_boq.geometry.to_crs(epsg=4326).y
    if "vendor" not in points_boq.columns:
        points_boq["vendor"] = vendor
    if "program" not in points_boq.columns:
        points_boq["program"] = program

    # used_columns = {
    #     "ring_name": "Ring ID",
    #     "site_id": "Site ID",
    #     "site_name": "Site Name" if "site_name" in points_boq.columns else "N/A",
    #     "long": "Long",
    #     "lat": "Lat",
    #     "region": "Region",
    #     "vendor": "Vendor" if "vendor" in points_boq.columns else "N/A",
    #     "program": "Program" if "program" in points_boq.columns else "N/A",
    #     "geometry": "geometry",
    #     }

    # available_col = [col for col in used_columns.keys() if col in points_boq.columns]

    # -- Sitelist & Hub --
    sitelist = points_boq[
        points_boq["site_type"].str.lower().str.contains("site")
    ].copy()
    hubs = points_boq[points_boq["site_type"].str.lower().str.contains("hub")].copy()

    # sitelist = sitelist[available_col].rename(columns=used_columns)
    # hubs = hubs[available_col].rename(columns=used_columns)

    sitelist = sitelist.drop_duplicates("geometry")
    hubs = hubs.drop_duplicates("geometry")
    sitelist = sitelist.to_crs(epsg=3246)
    hubs = hubs.to_crs(epsg=3246)

    # -- Route --
    route = lines_boq.copy()
    route["length"] = route.geometry.to_crs(epsg=4326).apply(geodesic_length)
    route_columns = ["near_end", "far_end", "geometry", "ring_name", "length"]
    route = route[route_columns].copy()
    route["name"] = route["near_end"] + sep + route["far_end"]

    # BOQ
    result_boq = compile_boq(
        points_boq,
        lines_boq,
        sep=sep,
        device_in_site=device_in_site,
        device_in_branch=device_in_branch,
    )
    (
        odp,
        otb,
        closure,
        backbone,
        access_ne,
        access_fe,
        fo_exist,
        pole_exist,
        obstacle_railway,
        obstacle_toll,
    ) = result_boq

    # SUMMARY
    # Columns for Points: site_id, site_name, ring_name, long, lat, region, vendor, program, type, geometry
    # Columns for Lines: near_end, far_end, ring_name, length, region, vendor, program, type, geometry

    # -- Sitelist Summary --
    sitelist["type"] = "Sitelist"
    sitelist["long"] = sitelist.geometry.to_crs(epsg=4326).x
    sitelist["lat"] = sitelist.geometry.to_crs(epsg=4326).y

    hubs["type"] = "FO Hub"
    hubs["long"] = hubs.geometry.to_crs(epsg=4326).x
    hubs["lat"] = hubs.geometry.to_crs(epsg=4326).y
    cols_sitelist = [
        "site_id",
        "site_name",
        "site_type",
        "type",
        "algo",
        "region",
        "ring_name",
        "vendor",
        "program",
        "geometry",
    ]

    valid_col = []
    for col in cols_sitelist:
        if col in sitelist.columns and col in hubs.columns:
            valid_col.append(col)
    hubs = hubs[valid_col]
    sitelist = sitelist[valid_col]

    sheet_sitelist = pd.concat([hubs, sitelist], join="inner")
    sheet_sitelist = sheet_sitelist.sort_values("ring_name")
    sheet_sitelist = sheet_sitelist.drop_duplicates(["ring_name", "geometry"])
    sheet_sitelist = sheet_sitelist.drop(columns="geometry")
    sheet_sitelist.columns = sheet_sitelist.columns.str.lower().str.replace(" ", "_")

    # -- Device --
    odp["type"] = "ODP"
    odp = odp.to_crs(epsg=4326)
    if not odp.empty and "geometry" in odp:
        odp["long"] = odp.geometry.to_crs(epsg=4326).x
        odp["lat"] = odp.geometry.to_crs(epsg=4326).y

    otb["type"] = device_in_site
    otb = otb.to_crs(epsg=4326)
    if not otb.empty and "geometry" in otb:
        otb["long"] = otb.geometry.to_crs(epsg=4326).x
        otb["lat"] = otb.geometry.to_crs(epsg=4326).y

    closure["type"] = "CL"
    closure = closure.to_crs(epsg=4326)
    if not closure.empty and "geometry" in closure:
        closure["long"] = closure.geometry.to_crs(epsg=4326).x
        closure["lat"] = closure.geometry.to_crs(epsg=4326).y

    sheet_devices = pd.concat([odp, otb, closure], join="inner")
    sheet_devices = sheet_devices.sort_values("ring_name")
    sheet_devices = sheet_devices.drop_duplicates(["ring_name", "geometry"])
    sheet_devices = sheet_devices.drop(columns="geometry")
    sheet_devices.columns = sheet_devices.columns.str.lower().str.replace(" ", "_")

    # -- Lines --
    route["type"] = "Route"
    route = route.to_crs(epsg=4326)
    if not route.empty and "geometry" in route:
        route["length"] = route.geometry.to_crs(epsg=4326).apply(geodesic_length)

    backbone["type"] = "Backbone"
    backbone = backbone.to_crs(epsg=4326)
    if not backbone.empty and "geometry" in backbone:
        backbone["length"] = backbone.geometry.to_crs(epsg=4326).apply(geodesic_length)

    access_fe["type"] = "Access"
    access_fe = access_fe.to_crs(epsg=4326)
    if not access_fe.empty and "geometry" in access_fe:
        access_fe["length"] = access_fe.geometry.to_crs(epsg=4326).apply(geodesic_length)

    fo_exist["type"] = "FO Existing"
    fo_exist = fo_exist.to_crs(epsg=4326)
    if not fo_exist.empty and "geometry" in fo_exist:
        fo_exist["length"] = fo_exist.geometry.to_crs(epsg=4326).apply(geodesic_length)

    pole_exist["type"] = "Pole Existing"
    pole_exist = pole_exist.to_crs(epsg=4326)
    if not pole_exist.empty and "geometry" in pole_exist:
        pole_exist["length"] = pole_exist.geometry.to_crs(epsg=4326).apply(geodesic_length)

    sheet_routes = pd.concat([route, backbone, access_fe, fo_exist, pole_exist])
    sheet_routes = sheet_routes.sort_values("ring_name")
    sheet_routes = sheet_routes.drop_duplicates(["ring_name", "geometry"])
    sheet_routes = sheet_routes.drop(columns="geometry")
    sheet_routes.columns = sheet_routes.columns.str.lower().str.replace(" ", "_")

    # -- Obstacle --
    obstacle_railway["type"] = "Obstacle Railway"
    obstacle_railway = obstacle_railway.to_crs(epsg=4326)
    if not obstacle_railway.empty and "geometry" in obstacle_railway:
        obstacle_railway["long"] = obstacle_railway.geometry.x
        obstacle_railway["lat"] = obstacle_railway.geometry.y

    obstacle_toll["type"] = "Obstacle Toll"
    obstacle_toll = obstacle_toll.to_crs(epsg=4326)
    if not obstacle_toll.empty and "geometry" in obstacle_toll:
        obstacle_toll["long"] = obstacle_toll.geometry.x
        obstacle_toll["lat"] = obstacle_toll.geometry.y

    sheet_obstacle = pd.concat([obstacle_railway, obstacle_toll])
    sheet_obstacle = sheet_obstacle.sort_values("ring_name")
    sheet_obstacle = sheet_obstacle.drop_duplicates(["ring_name", "geometry"])
    sheet_obstacle = sheet_obstacle.drop(columns="geometry")
    sheet_obstacle.columns = sheet_obstacle.columns.str.lower().str.replace(" ", "_")

    # -- Summary --
    summ_sitelist = (
        sheet_sitelist.groupby(["ring_name", "type"]).size().unstack(fill_value=0)
    )
    summ_devices = (
        sheet_devices.groupby(["ring_name", "type"]).size().unstack(fill_value=0)
    )
    summ_routes = (
        sheet_routes.groupby(["ring_name", "type"])["length"]
        .sum()
        .unstack(fill_value=0)
    )
    summ_obstacle = (
        sheet_obstacle.groupby(["ring_name", "type"]).size().unstack(fill_value=0)
    )
    summary_compiled = summ_sitelist.copy()
    summary_compiled = (
        summ_sitelist.join(summ_devices).join(summ_routes).join(summ_obstacle).fillna(0)
    )

    # EXPORT EXCEL
    excel_path = os.path.join(export_dir, f"BOQ Report.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        if not summary_compiled.empty:
            sheet_name = "Summary"
            summary_compiled = summary_compiled.reset_index()
            excel_styler(summary_compiled).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
            logger.info(
                f"📊 Excel sheet '{sheet_name}' with {len(summary_compiled):,} records written."
            )
        if not sheet_sitelist.empty:
            sheet_name = "Sitelist Information"
            sheet_sitelist = sheet_sitelist.reset_index(drop=True)
            excel_styler(sheet_sitelist).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
            logger.info(
                f"📊 Excel sheet '{sheet_name}' with {len(sheet_sitelist):,} records written."
            )
        if not sheet_devices.empty:
            sheet_name = "Devices Information"
            sheet_devices = sheet_devices.reset_index(drop=True)
            excel_styler(sheet_devices).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
            logger.info(
                f"📊 Excel sheet '{sheet_name}' with {len(sheet_devices):,} records written."
            )
        if not sheet_routes.empty:
            sheet_name = "Routes Information"
            sheet_routes = sheet_routes.reset_index(drop=True)
            excel_styler(sheet_routes).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
            logger.info(
                f"📊 Excel sheet '{sheet_name}' with {len(sheet_routes):,} records written."
            )
        if not sheet_obstacle.empty:
            sheet_name = "Obstacle"
            sheet_obstacle = sheet_obstacle.reset_index(drop=True)
            excel_styler(sheet_obstacle).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
            logger.info(
                f"📊 Excel sheet '{sheet_name}' with {len(sheet_obstacle):,} records written."
            )
    logger.info("✅ Save Excel file BOQ Done.")


def kmz_boq(
    main_kml,
    lines_boq: gpd.GeoDataFrame,
    points_boq: gpd.GeoDataFrame,
    boq_data: tuple,
    folder: str,
    device_in_site="OTB",
    **kwargs,
):
    program = kwargs.get("program", "N/A")
    vendor = kwargs.get("vendor", "TBG")
    sep = kwargs.get("sep", ";")

    lines_boq = lines_boq.copy()
    points_boq = points_boq.copy()

    def safe_get_geometry(site_id):
        match = points_boq.loc[
            points_boq["site_id"].astype(str).str.strip() == str(site_id), "geometry"
        ]
        if not match.empty:
            return match.iloc[0]
        else:
            logger.info(
                f"⚠️ Missing geometry for site_id: {site_id} in folder {folder}."
            )
            return None

    lines_boq["start"] = (
        lines_boq["near_end"].astype(str).str.strip().apply(safe_get_geometry)
    )
    lines_boq["end"] = (
        lines_boq["far_end"].astype(str).str.strip().apply(safe_get_geometry)
    )

    lines_boq = lines_boq.reset_index(drop=True)
    filename = folder.replace("/", "-")
    if "long" not in points_boq.columns or "lat" not in points_boq.columns:
        points_boq["long"] = points_boq.geometry.to_crs(epsg=4326).x
        points_boq["lat"] = points_boq.geometry.to_crs(epsg=4326).y
    if "vendor" not in points_boq.columns:
        points_boq["vendor"] = vendor
    if "program" not in points_boq.columns:
        points_boq["program"] = program

    used_columns = {
        "ring_name": "Ring ID",
        "site_id": "Site ID",
        "site_name": "Site Name" if "site_name" in points_boq.columns else "N/A",
        "long": "Long",
        "lat": "Lat",
        "region": "Region",
        "vendor": "Vendor" if "vendor" in points_boq.columns else "N/A",
        "program": "Program" if "program" in points_boq.columns else "N/A",
        "geometry": "geometry",
    }

    available_col = [col for col in used_columns.keys() if col in points_boq.columns]

    # DESIGN
    # -- Topology --
    try:
        logger.info(f"ℹ️ Total Point {len(points_boq)}")
        ring = folder.split("/")[-1]
        point_conn, connection = identify_connection(
            ring=ring, target_fiber=lines_boq, target_point=points_boq
        )
    except Exception as e:
        logger.error(f"Failed identify connection: {e}")
        return main_kml

    points_boq = point_conn.copy()
    ring_topology = create_topology(points_boq)
    ring_topology = ring_topology.to_crs(epsg=4326)
    ring_topology["connection"] = "Connection"

    # -- Route --
    ring_route = lines_boq.copy()
    ring_route["length"] = ring_route.geometry.to_crs(epsg=4326).apply(geodesic_length)
    route_columns = ["near_end", "far_end", "geometry", "ring_name", "length"]
    ring_route = ring_route[route_columns].copy()
    ring_route["name"] = ring_route["near_end"] + sep + ring_route["far_end"]

    sorted_route = []
    for num, ne in enumerate(connection, start=1):
        near_end = ring_route[
            ring_route["near_end"].astype(str).str.strip() == str(ne).strip()
        ].copy()
        if not near_end.empty:
            sorted_route.append(near_end)
        else:
            far_end = ring_route[
                ring_route["far_end"].astype(str).str.strip() == str(ne).strip()
            ].copy()
            if not far_end.empty:
                logger.info(f"🟢 {ne} not found as NE, but found as FE")
                sorted_route.append(far_end)
            else:
                logger.info(f"🔴 {ne} not found in ring route.")
                logger.info(ring_route[["near_end", "far_end"]])

    sorted_route = pd.concat(sorted_route)
    sorted_route = sorted_route.drop_duplicates("geometry").reset_index(drop=True)
    ring_route = sorted_route.copy()

    # -- Sitelist & Hub --
    ring_sites = points_boq[
        ~points_boq["site_type"].str.lower().str.contains("hub")
    ].copy()
    ring_hub = points_boq[
        points_boq["site_type"].str.lower().str.contains("hub")
    ].copy()

    ring_sites = ring_sites[available_col].rename(columns=used_columns)
    ring_hub = ring_hub[available_col].rename(columns=used_columns)

    ring_sites = ring_sites.drop_duplicates("geometry")
    ring_hub = ring_hub.drop_duplicates("geometry")

    # -- DESIGN --
    ring_topology = ring_topology.to_crs(epsg=4326)
    ring_route = ring_route.to_crs(epsg=4326)
    ring_sites = ring_sites.to_crs(epsg=4326)
    ring_hub = ring_hub.to_crs(epsg=4326)

    kml_updated = export_kml(
        ring_topology,
        main_kml,
        filename,
        subfolder=folder,
        name_col="connection",
        color="#FF00FF",
        size=2,
        popup=False,
    )
    kml_updated = export_kml(
        ring_route,
        kml_updated,
        filename,
        subfolder=f"{folder}/Route",
        name_col="name",
        color="#0000FF",
        size=3,
        popup=False,
    )
    kml_updated = export_kml(
        ring_sites,
        kml_updated,
        filename,
        subfolder=f"{folder}/Site List",
        name_col="Site ID",
        color="#FFFF00",
        size=0.8,
        popup=True,
    )
    kml_updated = export_kml(
        ring_hub,
        kml_updated,
        filename,
        subfolder=f"{folder}/FO Hub",
        name_col="Site ID",
        icon="http://maps.google.com/mapfiles/kml/paddle/A.png",
        size=0.8,
        popup=True,
    )

    # -- BOQ --
    # result_boq = compile_boq(points_boq, lines_boq, sep=sep, device_in_site=device_in_site)
    (
        odp,
        otb,
        closure,
        backbone,
        access_ne,
        access_fe,
        fo_exist,
        pole_exist,
        obstacle_railway,
        obstacle_toll,
    ) = boq_data

    backbone = backbone.to_crs(epsg=4326)
    access_fe = access_fe.to_crs(epsg=4326)
    fo_exist = fo_exist.to_crs(epsg=4326)
    pole_exist = pole_exist.to_crs(epsg=4326)
    odp = odp.to_crs(epsg=4326)
    otb = otb.to_crs(epsg=4326)
    closure = closure.to_crs(epsg=4326)
    obstacle_railway = obstacle_railway.to_crs(epsg=4326)
    obstacle_toll = obstacle_toll.to_crs(epsg=4326)

    kml_updated = export_kml(
        backbone,
        kml_updated,
        filename,
        subfolder=f"{folder}/Route Backbone",
        name_col="name",
        color="#0000FF",
        size=3,
        popup=False,
    )
    kml_updated = export_kml(
        access_fe,
        kml_updated,
        filename,
        subfolder=f"{folder}/Route Akses",
        name_col="name",
        color="#FF0000",
        size=3,
        popup=False,
    )
    kml_updated = export_kml(
        odp,
        kml_updated,
        filename,
        subfolder=f"{folder}/ODP",
        name_col="name",
        icon="http://maps.google.com/mapfiles/kml/shapes/triangle.png",
        color="#00FF00",
        size=0.8,
        popup=False,
    )
    kml_updated = export_kml(
        otb,
        kml_updated,
        filename,
        subfolder=f"{folder}/OTB",
        name_col="name",
        icon="http://maps.google.com/mapfiles/kml/shapes/triangle.png",
        color="#00FF00",
        size=0.8,
        popup=False,
    )
    kml_updated = export_kml(
        closure,
        kml_updated,
        filename,
        subfolder=f"{folder}/Closure",
        name_col="name",
        icon="http://maps.google.com/mapfiles/kml/shapes/triangle.png",
        color="#00FF00",
        size=0.8,
        popup=False,
    )
    kml_updated = export_kml(
        fo_exist,
        kml_updated,
        filename,
        subfolder=f"{folder}/FO Existing",
        name_col="name",
        color="#00FF00",
        size=6,
        popup=False,
    )
    kml_updated = export_kml(
        pole_exist,
        kml_updated,
        filename,
        subfolder=f"{folder}/Pole Existing",
        name_col="name",
        color="#FFFFFF",
        size=6,
        popup=False,
    )
    kml_updated = export_kml(
        obstacle_railway,
        kml_updated,
        filename,
        subfolder=f"{folder}/Obstacle",
        name_col="name",
        icon="http://maps.google.com/mapfiles/kml/shapes/rail.png",
        color="#FFFFFF",
        size=0.8,
        popup=False,
    )
    kml_updated = export_kml(
        obstacle_toll,
        kml_updated,
        filename,
        subfolder=f"{folder}/Obstacle",
        name_col="name",
        icon="http://maps.google.com/mapfiles/kml/shapes/cabs.png",
        color="#FFFFFF",
        size=0.8,
        popup=False,
    )

    return kml_updated


def save_boq(points_boq: gpd.GeoDataFrame, lines_boq: gpd.GeoDataFrame, export_dir: str, sep="-"):
    result_boq = compile_boq(points_boq, lines_boq, sep=sep)
    (
        odp,
        otb,
        closure,
        backbone,
        access_ne,
        access_fe,
        fo_exist,
        pole_exist,
        obstacle_railway,
        obstacle_toll,
    ) = result_boq

    # CLEAN GEOMETRY
    clean_col = [
        "otb",
        "odp",
        "backbone",
        "fo_exist",
        "pole_exist",
        "closure",
        "obstacle_railway",
        "obstacle_toll",
    ]
    for col in clean_col:
        if col in points_boq.columns:
            points_boq = points_boq.drop(columns=col)
        if col in lines_boq.columns:
            lines_boq = lines_boq.drop(columns=col)

    # EXPORT
    points_boq.to_parquet(os.path.join(export_dir, "Points_BOQ.parquet"))
    lines_boq.to_parquet(os.path.join(export_dir, "Routes_BOQ.parquet"))
    if not odp.empty:
        odp.to_parquet(os.path.join(export_dir, "ODP_BOQ.parquet"))
    if not otb.empty:
        otb.to_parquet(os.path.join(export_dir, "OTB_BOQ.parquet"))
    if not backbone.empty:
        backbone.to_parquet(os.path.join(export_dir, "Backbone_BOQ.parquet"))
    if not access_ne.empty:
        access_ne.to_parquet(os.path.join(export_dir, "Access_NE_BOQ.parquet"))
    if not access_fe.empty:
        access_fe.to_parquet(os.path.join(export_dir, "Access_FE_BOQ.parquet"))
    if not closure.empty:
        closure.to_parquet(os.path.join(export_dir, "Closure_BOQ.parquet"))
    if not fo_exist.empty:
        fo_exist.to_parquet(os.path.join(export_dir, "FO_Exist_BOQ.parquet"))
    if not pole_exist.empty:
        pole_exist.to_parquet(os.path.join(export_dir, "Pole_Exist_BOQ.parquet"))
    if not obstacle_railway.empty:
        obstacle_railway.to_parquet(
            os.path.join(export_dir, "Obstacle_Railway_BOQ.parquet")
        )
    if not obstacle_toll.empty:
        obstacle_toll.to_parquet(os.path.join(export_dir, "Obstacle_Toll_BOQ.parquet"))
    logger.info(f"✅ Save BOQ Parquet Done.")

def boq_generation(
    kmz_path: str,
    export_dir: str,
    operator: Operator | str = Operator.XL,
    sep: str = ";",
    interval_pole_m: int = 80,
    cable_percentage: int = 10,
    sclc_enabled: bool = False,
    device_in_site: DeviceType = DeviceType.OTB,
    device_in_branch: DeviceType = DeviceType.ODP,
    connector_in_site: ConnectorType = ConnectorType.SC,
    connector_in_branch: ConnectorType = ConnectorType.SC,
    program_name: str = "Intersite FO"
):
    def even_excel(x:float|int):
        return math.ceil(x/2) * 2
    
    validated = validate_kmz_ipl(kmz_path, sep=sep)
    if validated is None:
        logger.info(f"❌ BOQ generation failed (invalid KMZ): {kmz_path}")
        return

    # ---------------------------
    # Inputs (GeoDataFrames)
    # ---------------------------
    gdf_points = validated["points_data"]
    gdf_lines = validated["lines_data"]

    gdf_hub = validated["fo_hub"]
    gdf_sitelist = validated["sitelist"]
    gdf_odp = validated["odp"]
    gdf_otb = validated["otb"]
    gdf_closure = validated["closure"]
    gdf_topology = validated["topology"]

    gdf_route = validated["route"]
    gdf_backbone = validated["backbone"]
    gdf_access = validated["access"]
    gdf_fo_exist = validated["fo_exist"]
    gdf_pole_exist = validated["pole_exist"]
    gdf_obstacle = validated["obstacle"]

    # ---------------------------
    # Enrich Metadata and CRS normalize
    # ---------------------------
    gdf_hub = admin_information(gdf_hub, level="kabkot")

    target_crs = 3857
    gdf_points = gdf_points.to_crs(epsg=target_crs)
    gdf_lines = gdf_lines.to_crs(epsg=target_crs)

    gdf_hub = gdf_hub.to_crs(epsg=target_crs)
    gdf_sitelist = gdf_sitelist.to_crs(epsg=target_crs)
    gdf_odp = gdf_odp.to_crs(epsg=target_crs)
    gdf_otb = gdf_otb.to_crs(epsg=target_crs)
    gdf_closure = gdf_closure.to_crs(epsg=target_crs)
    gdf_topology = gdf_topology.to_crs(epsg=target_crs)

    gdf_route = gdf_route.to_crs(epsg=target_crs)
    gdf_backbone = gdf_backbone.to_crs(epsg=target_crs)
    gdf_access = gdf_access.to_crs(epsg=target_crs)
    gdf_fo_exist = gdf_fo_exist.to_crs(epsg=target_crs)
    gdf_pole_exist = gdf_pole_exist.to_crs(epsg=target_crs)
    gdf_obstacle = gdf_obstacle.to_crs(epsg=target_crs)

    # ---------------------------
    # Compile BOQ records
    # ---------------------------
    piv_records: list[dict] = []
    boq_records: list[dict] = []
    boq_pr_records: list[dict] = []
    recorded_segment = set()
    num = 1
    gdf_route = gdf_route.reset_index(drop=True)
    for ring_name, gdf_ring_route in gdf_route.groupby("ring_name"):
        gdf_ring_segments = gdf_ring_route.drop_duplicates().copy()
        gdf_ring_sites = gdf_sitelist[gdf_sitelist["ring_name"] == ring_name].copy()
        gdf_ring_hubs = gdf_hub[gdf_hub["ring_name"] == ring_name].copy()

        # Metadata
        program = (gdf_ring_sites["program"].mode().iloc[0] if "program" in gdf_ring_sites.columns and not gdf_ring_sites.empty else None)
        region = (gdf_ring_sites["region"].mode().iloc[0] if "region" in gdf_ring_sites.columns and not gdf_ring_sites.empty else None)
        city = (gdf_ring_hubs["Kabkot"].mode().iloc[0] if "Kabkot" in gdf_ring_hubs.columns and not gdf_ring_hubs.empty else None)
        
        is_otb = DeviceType.OTB in [device_in_site, device_in_branch]
        is_odp = DeviceType.ODP in [device_in_site, device_in_branch]
        is_sc = ConnectorType.SC in [connector_in_branch, connector_in_site]
        is_fc = ConnectorType.FC in [connector_in_branch, connector_in_site]

        # Ring-level slices
        gdf_ring_backbone = gdf_backbone[gdf_backbone["ring_name"] == ring_name].copy()
        gdf_ring_access = gdf_access[gdf_access["ring_name"] == ring_name].copy()
        gdf_ring_fo_exist = gdf_fo_exist[gdf_fo_exist["ring_name"] == ring_name].copy()
        gdf_ring_pole_exist = gdf_pole_exist[gdf_pole_exist["ring_name"] == ring_name].copy()
        gdf_ring_otb = gdf_otb[gdf_otb["ring_name"] == ring_name].copy()
        gdf_ring_odp = gdf_odp[gdf_odp["ring_name"] == ring_name].copy()
        gdf_ring_closure = gdf_closure[gdf_closure["ring_name"] == ring_name].copy()
        gdf_ring_obstacle = gdf_obstacle[gdf_obstacle["ring_name"] == ring_name].copy()

        is_first = True
        for idx, seg_row in gdf_ring_segments.iterrows():
            seg_name = seg_row["name"]
            seg_ne = seg_row["near_end"]
            seg_fe = seg_row["far_end"]
            seg_ctx = f"ring={ring_name} seg={seg_name} ne={seg_ne} fe={seg_fe}"

            len_route_m = (
                round(float(seg_row['length']), 3)
                if seg_row.length is not None
                else 0.0
            )
            seg_core = int(seg_row.get("core", 24) or 24)
            
            # Previous Route
            if idx == 0:
                prev_ring = None
                len_prev_access_m = 0
                len_prev_access_ext_m = 0
            else:
                prev_seg = gdf_route.loc[idx-1, :]
                prev_ring = prev_seg['ring_name']
                prev_ne = prev_seg['near_end']
                prev_fe = prev_seg['far_end']
                prev_df_access = gdf_ring_access[(gdf_ring_access["near_end"] == prev_ne) & (gdf_ring_access["far_end"] == prev_fe)].copy()
                len_prev_access_m = (float(sum(prev_df_access['length'])) if not prev_df_access.empty else 0.0)
                len_prev_access_ext_m = 0

            # ---------------------------
            # Segment slices
            # ---------------------------
            df_bb = gdf_ring_backbone[
                (gdf_ring_backbone["near_end"] == seg_ne)
                & (gdf_ring_backbone["far_end"] == seg_fe)
            ].copy()
            df_access = gdf_ring_access[
                (gdf_ring_access["near_end"] == seg_ne)
                & (gdf_ring_access["far_end"] == seg_fe)
            ].copy()
            df_overlap = gdf_ring_fo_exist[
                (gdf_ring_fo_exist["near_end"] == seg_ne)
                & (gdf_ring_fo_exist["far_end"] == seg_fe)
            ].copy()
            df_pole = gdf_ring_pole_exist[
                (gdf_ring_pole_exist["near_end"] == seg_ne)
                & (gdf_ring_pole_exist["far_end"] == seg_fe)
            ].copy()

            df_otb = gdf_ring_otb[gdf_ring_otb["segment"] == seg_name].copy()
            df_otb_new = df_otb[df_otb["ext_note"] == 0].copy()
            df_otb_ext = df_otb[df_otb["ext_note"] == 1].copy()

            df_odp = gdf_ring_odp[gdf_ring_odp["segment"] == seg_name].copy()
            df_odp_new = df_odp[df_odp["ext_note"] == 0].copy()
            df_odp_ext = df_odp[df_odp["ext_note"] == 1].copy()

            if df_odp.empty:
                raise ValueError(
                    f"[ODP_NOT_FOUND] {seg_ctx}\n"
                    f"Segment snapshot:\n"
                    f"{gdf_ring_segments[['segment', 'near_end', 'far_end']].head(10).to_string(index=False)}"
                )

            logger.info(f"🟢 Processing {seg_ctx}")

            df_closure = gdf_ring_closure[gdf_ring_closure["segment"] == seg_name].copy()
            df_closure_new = df_closure[df_closure["ext_note"] == 0].copy()
            df_closure_ext = df_closure[df_closure["ext_note"] == 1].copy()

            df_obs_seg = gdf_ring_obstacle[(gdf_ring_obstacle["near_end"] == seg_ne) & (gdf_ring_obstacle["far_end"] == seg_fe)].copy()
            df_obs_toll = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("toll", case=False, na=False)].copy()
            df_obs_rail = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("rail", case=False, na=False)].copy()
            df_obs_bridge = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("bridge", case=False, na=False)].copy()

            # ---------------------------
            # Core parsing
            # ---------------------------
            core_bb = 24
            if not df_bb.empty and "name" in df_bb.columns:
                raw_name = str(df_bb["name"].iloc[0])
                tail = raw_name.split("_FO")[-1].replace("C", "")
                core_bb = int(tail) if tail.isdigit() else 24

            # ---------------------------
            # Length metrics
            # ---------------------------
            len_bb_m = float(sum(df_bb['length'])) if not df_bb.empty else 0.0
            len_access_m = (float(sum(df_access['length'])) if not df_access.empty else 0.0)
            len_overlap_m = ( float(sum(df_overlap['length'])) if not df_overlap.empty else 0.0)
            len_pole_m = (float(sum(df_pole['length'])) if not df_pole.empty else 0.0)
            len_access_ext_m = 0.0

            # Cable length by backbone core
            len_cable_by_core_m = {c: (len_route_m if core_bb == c else 0.0) for c in (24, 48, 72, 96, 120, 144)}
            
            # ---------------------------
            # Quantity metrics
            # ---------------------------
            qty_otb = len(df_otb)
            qty_otb_new = len(df_otb_new)
            qty_otb_ext = len(df_otb_ext)

            qty_odp = len(df_odp)
            qty_odp_new = len(df_odp_new)
            qty_odp_ext = len(df_odp_ext)

            df_otb_by_core = {
                c: df_otb_new[df_otb_new["core"] == c] for c in (24, 48, 72, 96, 120, 144)
            }
            df_odp_by_core = {
                c: df_odp_new[df_odp_new["core"] == c] for c in (24, 48, 72, 96, 120, 144)
            }
            qty_otb_by_core = {c: len(df_) for c, df_ in df_otb_by_core.items()}
            qty_odp_by_core = {c: len(df_) for c, df_ in df_odp_by_core.items()}

            qty_closure = len(df_closure)
            qty_closure_new = len(df_closure_new)
            qty_closure_ext = len(df_closure_ext)

            qty_obs_toll = len(df_obs_toll)
            qty_obs_rail = len(df_obs_rail)
            qty_obs_bridge = len(df_obs_bridge)

            # ---------------------------
            # Calculations
            # ---------------------------
            fo_factor = 1 + (cable_percentage/100)
            calc_permission_pu = math.floor(len_bb_m + len_access_m - len_pole_m + sum(len_cable_by_core_m.get(core, 0) for core in len_cable_by_core_m.keys() if int(core) != 24))
            calc_fo_cable_m = math.ceil(math.ceil(len_bb_m + len_access_m) * fo_factor / 100) * 100
            calc_closure_24_qty = qty_closure_new + (math.floor(calc_fo_cable_m / 4000) if calc_fo_cable_m >= 4000 else 0)
            calc_total_overlap_m = round((len_overlap_m + len_access_ext_m + len_prev_access_m + len_prev_access_ext_m if ring_name == prev_ring else len_overlap_m + len_access_ext_m) * fo_factor, 0)

            # Material
            calc_mat_hdpe_subduct_32_27_qty = 20 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 20 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0) + 70 * qty_obs_rail
            calc_mat_gi_pipe_1p5in_qty = 3 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 3 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0) + 3 * (2 * qty_obs_rail)
            calc_mat_pole_fo_9m_3step_qty = 0 if ((calc_permission_pu / interval_pole_m) < 3) else even_excel((calc_permission_pu / interval_pole_m) * 0.05) #=IF((S10/80)<3;0;EVEN(((S10)/80)*0,05))
            calc_mat_pole_fo_7m_2step_qty = 0 if calc_permission_pu < 0 else even_excel(calc_permission_pu/interval_pole_m) - calc_mat_pole_fo_9m_3step_qty #=IF((S10)<0;0;EVEN(((S10)/70)-DV10))
            calc_mat_slack_support_70x70x3_qty = 1 + math.floor((calc_mat_pole_fo_7m_2step_qty + calc_mat_pole_fo_9m_3step_qty)/4) if calc_mat_pole_fo_7m_2step_qty + calc_mat_pole_fo_9m_3step_qty > 0 else 0 # =IF(SUM(DU10;DV10)>0;1+ROUNDDOWN(SUM(DU10;DV10)/4;0);0)
            
            # Services
            otb_factor = (is_otb and (is_sc or is_fc))
            is_sc_odp = (is_sc and is_odp)

            calc_svc_pulling_fo_aerial_incl_pole_m = (calc_fo_cable_m + (len_cable_by_core_m.get(core_bb, 0) if core_bb != 24 else 0) - len_pole_m - calc_mat_hdpe_subduct_32_27_qty + 0 ) if (calc_fo_cable_m + (len_cable_by_core_m.get(core_bb, 0) if core_bb != 24 else 0) >= 20) else 0
            calc_splicing_fusion = ((calc_closure_24_qty + (qty_odp_by_core.get(24, 0) if is_sc_odp else 0)) * 24 + (qty_odp_by_core.get(48, 0) if is_sc_odp else 0) * 48 + (qty_odp_by_core.get(4, 0)  if is_sc_odp else 0) * 4 + (qty_odp_by_core.get(8, 0)  if is_sc_odp else 0) * 8 + (qty_odp_by_core.get(16, 0) if is_sc_odp else 0) * 16)
            calc_termination_fusion = sum((qty_otb_by_core.get(core, 0) if otb_factor else 0) * core for core in (12, 24, 48, 96, 144, 288))

            calc_svc_splicing_fusion_qty = 24 if (calc_splicing_fusion == 0 and calc_fo_cable_m > 0) else calc_splicing_fusion
            calc_svc_termination_fusion_qty = calc_termination_fusion
            
            # Testing
            calc_test_otdr_2lambda_2way_ls = (calc_svc_termination_fusion_qty if calc_svc_termination_fusion_qty > 0 else 96 if len_cable_by_core_m.get(96, 0) > 0 else 48 if len_cable_by_core_m.get(48, 0) > 0 else 24)

            # ---------------------------
            # Pivot record
            # ---------------------------
            piv_record = {
                "program": program,
                "operator": operator.upper(),
                "region": region,
                "city": city,
                "ring_name": ring_name,
                "segment_name": seg_name,
                "near_end": seg_ne,
                "far_end": seg_fe,
                "route_qty": None,
                "route_length_m": len_route_m,
                "bb_qty": None,
                "bb_length_m": len_bb_m,
                "access_qty": None,
                "access_length_m": len_access_m,
                "access_ext_qty": None,
                "access_ext_length_m": len_access_ext_m,
                "overlap_qty": None,
                "overlap_length_m": len_overlap_m,
                "pole_ext_length_m": len_pole_m,
                "otb_qty": qty_otb,
                "otb_new_qty": qty_otb_new,
                "otb_ext_qty": qty_otb_ext,
                "odp_qty": qty_odp,
                "odp_new_qty": qty_odp_new,
                "odp_ext_qty": qty_odp_ext,
                "odp_24_qty": qty_odp_by_core.get(24, 0),
                "odp_48_qty": qty_odp_by_core.get(48, 0),
                "odp_72_qty": qty_odp_by_core.get(72, 0),
                "odp_96_qty": qty_odp_by_core.get(96, 0),
                "odp_120_qty": qty_odp_by_core.get(120, 0),
                "odp_144_qty": qty_odp_by_core.get(144, 0),
                "closure_qty": qty_closure,
                "closure_new_qty": qty_closure_new,
                "closure_ext_qty": qty_closure_ext,
                "obstacle_toll_qty": qty_obs_toll,
                "obstacle_railway_qty": qty_obs_rail,
                "obstacle_bridge_qty": qty_obs_bridge,
                "permission_pu": calc_permission_pu,
                "fo_cable_m": calc_fo_cable_m,
                "total_overlap_m": calc_total_overlap_m,
                "cable_24_m": len_cable_by_core_m[24],
                "cable_48_m": len_cable_by_core_m[48],
                "cable_72_m": len_cable_by_core_m[72],
                "cable_96_m": len_cable_by_core_m[96],
                "cable_120_m": len_cable_by_core_m[120],
                "cable_144_m": len_cable_by_core_m[144],
            }
            piv_records.append(piv_record)

            # BOQ Record
            seg_record = {
                # Segment Info
                "no": num,
                "site_type": program if program != "NA" else program_name,
                "stip_category": None,
                "operator": operator.upper(),
                "spk_wo": None,
                "segment_type": "Segment",
                "sonumb": None,
                "ring_id": ring_name,
                "segment_id": seg_name,
                "area_city": city,
                "program": region,
                "work_code": None,
                "region_procurement": None,
                "e2e_distance_m": calc_fo_cable_m + calc_total_overlap_m, # len_cable_by_core_m[48] + len_cable_by_core_m[72] + len_cable_by_core_m[96] + len_cable_by_core_m[120] + len_cable_by_core_m[144]
                "boq_id": None,
                "subset_separator": None,

                # PREPARATION
                "prep_a1_rambu_papan": int(is_first),
                "prep_a2_direksi_keet": int(is_first),
                "prep_a3_koordinasi_pu_m": calc_permission_pu,
                "prep_a4_koordinasi_pjka": qty_obs_rail,
                "prep_a5_koordinasi_toll": qty_obs_toll,
                "prep_a6_koordinasi_private_supervise": None,
                "prep_a7_survey_abd": 1,
                "prep_a8_mobilisasi_transpor": int(is_first),

                # MATERIAL SUPPLY
                # Fo Cable
                "mat_fo_aerial_12_m": None,
                "mat_fo_aerial_24_m": None,
                "mat_fo_aerial_48_m": None,
                "mat_fo_aerial_72_m": None,
                "mat_fo_aerial_96_m": None,
                "mat_fo_aerial_144_m": None,
                "mat_fo_adss_12_m": len_cable_by_core_m.get(12, 0),
                "mat_fo_adss_24_m": calc_fo_cable_m, #  len_cable_by_core_m.get(24, 100) if calc_fo_cable_m is None else 
                "mat_fo_adss_48_m": len_cable_by_core_m.get(48, 0),
                "mat_fo_adss_72_m": len_cable_by_core_m.get(72, 0),
                "mat_fo_adss_96_m": len_cable_by_core_m.get(96, 0),
                "mat_fo_adss_144_m": len_cable_by_core_m.get(144, 0),
                "mat_fo_db_sj_12_m": None,
                "mat_fo_db_sj_24_m": None,
                "mat_fo_db_sj_48_m": None,
                "mat_fo_db_sj_72_m": None,
                "mat_fo_db_sj_96_m": None,
                "mat_fo_db_sj_144_m": None,
                "mat_fo_db_sj_264_m": None,
                "mat_fo_db_sj_288_m": None,
                "mat_fo_duct_12_m": None,
                "mat_fo_duct_24_m": None,
                "mat_fo_duct_48_m": None,
                "mat_fo_duct_72_m": None,
                "mat_fo_duct_96_m": None,
                "mat_fo_duct_144_m": None,
                "mat_fo_duct_264_m": None,
                "mat_fo_duct_288_m": None,
                "mat_fo_micro_24_m": None,
                "mat_fo_micro_48_m": None,
                "mat_fo_micro_96_m": None,
                "mat_fo_micro_144_m": None,
                "mat_fo_micro_288_m": None,
                "mat_fo_dropwire_1_m": None,
                "mat_fo_dropwire_2_m": None,
                "mat_closure_dome_24_qty": calc_closure_24_qty,
                "mat_closure_dome_48_qty": None,
                "mat_closure_dome_144_qty": None,
                "mat_closure_dome_96_qty": None,
                "mat_closure_dome_864_hd_qty": None,
                "mat_closure_inline_24_qty": None,
                "mat_closure_inline_48_qty": None,
                "mat_closure_inline_96_qty": None,
                "mat_closure_inline_144_qty": None,
                "mat_closure_inline_288_qty": None,
                "mat_otb_12_sc_upc_qty": qty_otb_by_core.get(12, 0) if (is_sc and is_otb) else None,
                "mat_otb_12_fc_upc_qty": qty_otb_by_core.get(12, 0) if (is_fc and is_otb) else None,
                "mat_otb_24_sc_upc_qty": qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else None,
                "mat_otb_24_fc_upc_qty": qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else None,
                "mat_otb_48_sc_upc_qty": qty_otb_by_core.get(48, 0) if (is_sc and is_otb) else None,
                "mat_otb_96_sc_upc_qty": qty_otb_by_core.get(96, 0) if (is_sc and is_otb) else None,
                "mat_otb_144_sc_upc_qty": qty_otb_by_core.get(144, 0) if (is_sc and is_otb) else None,
                "mat_otb_288_sc_upc_qty": qty_otb_by_core.get(288, 0) if (is_sc and is_otb) else None,
                "mat_otb_288_lc_upc_hd_qty": None,
                "mat_odp_4_sc_upc_qty": qty_odp_by_core.get(4, 0) if (is_sc and is_odp) else None,
                "mat_odp_8_sc_upc_qty": qty_odp_by_core.get(8, 0) if (is_sc and is_odp) else None,
                "mat_odp_16_sc_upc_qty": qty_odp_by_core.get(16, 0) if (is_sc and is_odp) else None,
                "mat_odp_24_sc_upc_qty": qty_odp_by_core.get(24, 0) if (is_sc and is_odp) else None,
                "mat_odp_96_sc_upc_qty": qty_odp_by_core.get(96, 0) if (is_sc and is_odp) else None,
                "mat_odp_pedestal_16_ug_qty": None,
                "mat_splitter_1to2_sc_upc_qty": None,
                "mat_splitter_1to4_sc_upc_qty": None,
                "mat_splitter_1to8_sc_upc_qty": None,
                "mat_splitter_1to16_sc_upc_qty": None,
                "mat_splitter_1to32_sc_upc_qty": None,
                "mat_rosset_2port_sc_upc_qty": None,
                "mat_hdpe_subduct_32_27_qty": calc_mat_hdpe_subduct_32_27_qty,
                "mat_hdpe_subduct_40_33_qty": None,
                "mat_hdpe_subduct_50_43_qty": None,
                "mat_hdpe_microduct_4way_12_10_qty": None,
                "mat_hdpe_microduct_7way_12_10_qty": None,
                "mat_hdpe_microduct_14way_12_10_qty": None,
                "mat_hdpe_microduct_4way_18_14_qty": None,
                "mat_hdpe_microduct_7way_18_14_qty": None,
                "mat_hdpe_microduct_14way_18_14_qty": None,
                "mat_gi_pipe_1in_qty": None,
                "mat_gi_pipe_1p5in_qty": calc_mat_gi_pipe_1p5in_qty,
                "mat_gi_pipe_2in_qty": None,
                "mat_gi_pipe_3in_qty": None,
                "mat_gi_pipe_4in_qty": None,
                "mat_gi_pipe_5in_qty": None,
                "mat_pvc_pipe_20mm_qty": None,
                "mat_pvc_pipe_1p5in_qty": None,
                "mat_pvc_pipe_2in_qty": None,
                "mat_pvc_pipe_3in_qty": None,
                "mat_pvc_pipe_4in_qty": None,
                "mat_pvc_pipe_5in_qty": None,
                "mat_flex_conduit_0p75in_qty": None,
                "mat_flex_conduit_2in_qty": None,
                "mat_flex_conduit_3in_qty": None,
                "mat_flex_conduit_20mm_qty": None,
                "mat_rack_open_42u_qty": None,
                "mat_rack_closed_8u_450_qty": None,
                "mat_rack_closed_20u_900_qty": None,
                "mat_rack_closed_42u_900_qty": None,
                "mat_rack_closed_42u_1150_qty": None,
                "mat_odc_144_qty": None,
                "mat_odc_288_qty": None,
                "mat_odc_576_qty": None,
                "mat_warning_tape_fo_6in_m": None,
                "mat_pole_fo_7m_2step_qty": calc_mat_pole_fo_7m_2step_qty,
                "mat_pole_fo_9m_3step_qty": calc_mat_pole_fo_9m_3step_qty,
                "mat_slack_support_50x50x3_qty": None,
                "mat_pole_fo_12m_3step_qty": None,
                "mat_slack_support_70x70x3_qty": calc_mat_slack_support_70x70x3_qty,
                "mat_patch_indoor_sm_sc_upc_sc_upc_5m": None,
                "mat_patch_indoor_sm_sc_upc_sc_upc_20m": None,
                "mat_patch_indoor_sm_sc_upc_lc_upc_5m": None,
                "mat_patch_indoor_sm_sc_upc_lc_upc_20m": None,
                "mat_patch_indoor_sm_lc_upc_lc_upc_5m": None,
                "mat_patch_indoor_sm_lc_upc_lc_upc_20m": None,
                "mat_patch_indoor_sm_fc_upc_lc_upc_5m": None,
                "mat_patch_indoor_sm_fc_upc_lc_upc_20m": None,
                "mat_patch_outdoor_sm_sc_upc_sc_upc_5m": 2 if seg_name not in recorded_segment else 0,
                "mat_patch_outdoor_sm_sc_upc_sc_upc_20m": None,
                "mat_patch_outdoor_sm_sc_upc_lc_upc_5m": 2 if (seg_name not in recorded_segment and sclc_enabled) else 0,
                "mat_patch_outdoor_sm_sc_upc_lc_upc_20m": None,
                "mat_patch_outdoor_sm_lc_upc_lc_upc_5m": None,
                "mat_patch_outdoor_sm_lc_upc_lc_upc_20m": None,
                "mat_patch_indoor_mm_lc_upc_lc_upc_5m": None,
                "mat_patch_indoor_mm_lc_upc_lc_upc_20m": None,
                "mat_pigtail_indoor_sc_upc_qty": None,
                "mat_adapter_fc_fc_qty": None,
                "mat_adapter_sc_sc_simplex_qty": None,
                "mat_adapter_sc_sc_duplex_qty": None,
                "mat_adapter_lc_lc_simplex_qty": None,
                "mat_adapter_lc_lc_duplex_qty": None,
                "mat_adapter_sc_lc_hybrid_qty": None,
                "mat_adapter_fc_lc_hybrid_qty": None,
                "mat_adapter_fc_sc_hybrid_qty": None,
                "mat_clamp_omega_qty": None,
                "mat_supporting_material_qty": int(is_first),

                # SERVICES
                "svc_trenching_reinstate_hotmix_m": None,
                "svc_trenching_reinstate_asphalt_m": None,
                "svc_trenching_reinstate_makadam_m": 20 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 20 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0),
                "svc_trenching_reinstate_floor_cement_m": None,
                "svc_trenching_reinstate_paving_m": None,
                "svc_trenching_reinstate_agregat_sirtu_m": None,
                "svc_trenching_reinstate_taman_m": None,
                "svc_trenching_reinstate_tanah_biasa_m": None,
                "svc_install_hdpe_subduct_32_27_m": calc_mat_hdpe_subduct_32_27_qty,
                "svc_install_hdpe_subduct_40_33_m": None,
                "svc_install_hdpe_subduct_50_43_m": None,
                "svc_install_hdpe_microduct_4way_12_10_m": None,
                "svc_install_hdpe_microduct_7way_12_10_m": None,
                "svc_install_hdpe_microduct_14way_12_10_m": None,
                "svc_install_hdpe_microduct_4way_18_14_m": None,
                "svc_install_hdpe_microduct_7way_18_14_m": None,
                "svc_install_hdpe_microduct_14way_18_14_m": None,
                "svc_install_galvanized_1inch_m": None,
                "svc_install_galvanized_1_5inch_m": None,
                "svc_install_galvanized_2inch_m": None,
                "svc_install_galvanized_3inch_m": None,
                "svc_install_galvanized_4inch_m": None,
                "svc_install_galvanized_5inch_m": None,
                "svc_install_pvc_20mm_m": None,
                "svc_install_pvc_1_5inch_m": None,
                "svc_install_pvc_2inch_m": None,
                "svc_install_pvc_3inch_m": None,
                "svc_install_pvc_4inch_m": None,
                "svc_install_pvc_5inch_m": None,
                "svc_install_pole_7m_2step_unit": calc_mat_pole_fo_7m_2step_qty,
                "svc_install_pole_9m_3step_unit": calc_mat_pole_fo_9m_3step_qty,
                "svc_install_slack_support_50x50x3_qty": None,
                "svc_install_slack_support_additional_qty": None,
                "svc_install_slack_support_70x70x3_qty": calc_mat_slack_support_70x70x3_qty,
                "svc_install_riser_galvanized_1_5inch_3m_qty": calc_mat_gi_pipe_1p5in_qty / 3,
                "svc_install_riser_galvanized_2inch_3m_qty": None,
                "svc_supply_install_temberang_tarik_unit": None,
                "svc_supply_install_bridge_crossing_single_pole_unit": None,
                "svc_supply_install_bridge_crossing_double_pole_unit": None,
                "svc_install_galvanized_2inch_atb_m": 0, # Perlu cross bridge information
                "svc_install_galvanized_4inch_atb_m": None,
                "svc_install_galvanized_4inch_self_support_m": None,
                "svc_boring_manual_subduct_m": None,
                "svc_boring_manual_crossing_road_m": None,
                "svc_boring_manual_crossing_toll_m": None,
                "svc_boring_manual_crossing_railway_m": qty_obs_rail * 70,
                "svc_boring_machine_crossing_railway_m": None,
                "svc_boring_manual_crossing_river_m": None,
                "svc_boring_machine_crossing_river_m": None,
                "svc_supply_install_marking_post_unit": None,
                "svc_install_manhole_100x100x120_unit": None,
                "svc_install_handhole_80x80x120_unit": qty_obs_rail * 2,
                "svc_install_handhole_60x60x120_unit": None,
                "svc_install_odc_foundation_50x70x50_unit": None,
                "svc_pulling_fo_aerial_incl_pole_m": round(calc_svc_pulling_fo_aerial_incl_pole_m, 0),
                "svc_pulling_fo_aerial_excl_pole_m": round(len_pole_m, 0),
                "svc_pulling_fo_burial_m": calc_mat_hdpe_subduct_32_27_qty,
                "svc_pulling_fo_direct_buried_m": None,
                "svc_air_blown_fo_m": None,
                "svc_pulling_fo_drop_wire_m": None,
                "svc_pulling_fo_inbuilding_m": None,
                "svc_splicing_fusion_qty": calc_svc_splicing_fusion_qty, # Recheck Long Formula
                "svc_termination_fusion_qty": calc_svc_termination_fusion_qty, # Recheck Long Formula
                "svc_bobok_tembok_bor_qty": None,
                "svc_install_patch_indoor_sc_sc_5m_qty": None,
                "svc_install_patch_indoor_sc_sc_20m_qty": None,
                "svc_install_patch_indoor_sc_lc_5m_qty": None,
                "svc_install_patch_indoor_sc_lc_20m_qty": None,
                "svc_install_patch_indoor_lc_lc_5m_qty": None,
                "svc_install_patch_indoor_lc_lc_20m_qty": None,
                "svc_install_patch_outdoor_sc_sc_5m_qty": 2 if seg_name not in recorded_segment else 0,
                "svc_install_patch_outdoor_sc_sc_20m_qty": None,
                "svc_install_patch_outdoor_sc_lc_5m_qty": 2 if (seg_name not in recorded_segment and sclc_enabled) else 0,
                "svc_install_patch_outdoor_sc_lc_20m_qty": None,
                "svc_install_patch_outdoor_lc_lc_5m_qty": None,
                "svc_install_patch_outdoor_lc_lc_20m_qty": None,
                "svc_install_patch_indoor_mm_lc_lc_5m_qty": None,
                "svc_install_patch_indoor_mm_lc_lc_20m_qty": None,
                "svc_install_pigtail_indoor_qty": None,
                "svc_install_adapter_fc_fc_qty": None,
                "svc_install_adapter_sc_sc_simplex_qty": None,
                "svc_install_adapter_sc_sc_duplex_qty": None,
                "svc_install_adapter_lc_lc_simplex_qty": None,
                "svc_install_adapter_lc_lc_duplex_qty": None,
                "svc_install_adapter_sc_lc_hybrid_qty": None,
                "svc_install_adapter_fc_lc_hybrid_qty": None,
                "svc_install_adapter_fc_sc_hybrid_qty": None,
                "svc_install_otb_12c_sc_upc_qty": None,
                "svc_install_otb_12c_fc_upc_qty": None,
                "svc_install_otb_24c_sc_upc_qty": qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else None,
                "svc_install_otb_24c_fc_upc_qty": qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else None,
                "svc_install_otb_48c_sc_upc_qty": None,
                "svc_install_otb_96c_sc_upc_qty": None,
                "svc_install_otb_144c_sc_upc_qty": None,
                "svc_install_otb_288c_sc_upc_qty": None,
                "svc_install_otb_288c_lc_upc_hd_qty": None,
                "svc_install_odp_4c_sc_upc_qty": qty_odp_by_core.get(4, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_8c_sc_upc_qty": qty_odp_by_core.get(8, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_12c_sc_upc_qty": qty_odp_by_core.get(12, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_24c_sc_upc_qty": qty_odp_by_core.get(24, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_48c_sc_upc_qty": qty_odp_by_core.get(48, 0) if (is_sc and is_odp) else None,
                "svc_install_odp_96c_sc_upc_qty": qty_odp_by_core.get(96, 0) if (is_sc and is_odp) else None,
                "svc_support_integration_qty": 1,
                "test_otdr_2lambda_2way_ls": calc_test_otdr_2lambda_2way_ls,
                "test_opm_2lambda_2way_ls": calc_test_otdr_2lambda_2way_ls,
                "doc_hardcopy_ls": int(is_first),
                "doc_softcopy_ls": int(is_first),
                "overlap_fo": calc_total_overlap_m,
            }

            # BOQ PR Record
            pr_record = {
                "num": num,
                "site_type": program if program != "NA" else "Intersite FO",
                "stip_category": None,
                "operator": operator.upper(),
                "spk_wo": None,
                "segment_type": "Segment",
                "sonumb": None,
                "ring_id": ring_name,
                "segment_id": seg_name,
                "permit": calc_permission_pu,
                "shop_incl_pole_aerial": round(calc_svc_pulling_fo_aerial_incl_pole_m, 0),
                "shop_excl_pole_aerial": round(len_pole_m, 0),
                "shop_burial": calc_mat_hdpe_subduct_32_27_qty,
                "comcase": calc_permission_pu,
                "mat_pole_7": calc_mat_pole_fo_7m_2step_qty,
                "mat_pole_9": calc_mat_pole_fo_9m_3step_qty,
                "mat_cable_fo_24": calc_fo_cable_m,
                "mat_cable_fo_96": len_cable_by_core_m.get(96, 0),
                "mat_hdpe": calc_mat_hdpe_subduct_32_27_qty,
                "total_e2e_dist": calc_fo_cable_m + calc_total_overlap_m,
                "pole_exist": len_pole_m,
            }
            
            boq_records.append(seg_record)
            boq_pr_records.append(pr_record)

            # Update Segment Info
            recorded_segment.add(seg_name)
            is_first = False
            num += 1

    piv_df = pd.DataFrame(piv_records)
    boq_df = pd.DataFrame(boq_records)
    boq_pr_df = pd.DataFrame(boq_pr_records)

    # ---------------------------
    # Write into Excel template
    # ---------------------------
    template_path = os.path.join(DATA_DIR, "template", "boq", "Template_BOQ_Report.xlsx")
    output_path = os.path.join(export_dir, "BOQ Report.xlsx")

    if not os.path.exists(template_path):
        raise ValueError("BOQ template file not found in template directory.")

    shutil.copy2(template_path, output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        if boq_df.empty:
            logger.info("❌ No BOQ records to write.")
        else:
            boq_df = boq_df.reset_index(drop=True)
            # excel_styler(boq_df).to_excel(writer, sheet_name="Pivot Data", index=False)
            logger.info(
                f"📊 Excel sheet 'Pivot Data' written with {len(boq_df):,} records."
            )

    # ---------------------------
    # PROCESS BOQ EXCEL
    # ---------------------------
    wb = load_workbook(output_path)

    # Style
    named_style = NamedStyle(name="BOQ Row")
    side_style = Side(style="thin", border_style="thin")
    border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
    named_style.font = Font(name="Arial", size=11)
    named_style.border = border

    if "BOQ Row" not in [s for s in wb.named_styles]:
        wb.add_named_style(named_style)

    # BOQ Sheet
    boq_sheet = wb["BOQ"]
    start_data_row = 8
    col_index = {col: num for num, col in enumerate(boq_df.columns, start=1)}
    for idx, record in enumerate(boq_df.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = boq_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "BOQ Row"

            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # BOQ PR Sheet
    boq_pr_sheet = wb["BOQ PR"]
    start_data_row = 3
    col_index = {col: num for num, col in enumerate(boq_pr_df.columns, start=1)}
    for idx, record in enumerate(boq_pr_df.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = boq_pr_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "BOQ Row"

            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    logger.info("✅ BOQ Excel saved.")

def boq_mmp(
    kmz_path: str,
    export_dir: str,
    operator: Operator | str = Operator.XL,
    sep: str = ";",
    interval_pole_m: int = 60,
    cable_percentage: int = 15,
    cable_multiplier: int = 2,
    sclc_enabled: bool = False,
    device_in_site: DeviceType = DeviceType.ODP,
    device_in_branch: DeviceType = DeviceType.ODP,
    connector_in_site: ConnectorType = ConnectorType.SC,
    connector_in_branch: ConnectorType = ConnectorType.SC,
    program_name: str = "TBG MMP"
):
    def even_excel(x:float|int):
        return math.ceil(x/2) * 2
    
    validated = validate_kmz_ipl(kmz_path, sep=sep)
    if validated is None:
        logger.info(f"❌ BOQ generation failed (invalid KMZ): {kmz_path}")
        return

    # ---------------------------
    # Inputs (GeoDataFrames)
    # ---------------------------
    gdf_points = validated["points_data"]
    gdf_lines = validated["lines_data"]

    gdf_hub = validated["fo_hub"]
    gdf_sitelist = validated["sitelist"]
    gdf_odp = validated["odp"]
    gdf_otb = validated["otb"]
    gdf_closure = validated["closure"]
    gdf_topology = validated["topology"]

    gdf_route = validated["route"]
    gdf_backbone = validated["backbone"]
    gdf_access = validated["access"]
    gdf_fo_exist = validated["fo_exist"]
    gdf_pole_exist = validated["pole_exist"]
    gdf_obstacle = validated["obstacle"]

    # ---------------------------
    # Enrich Metadata and CRS normalize
    # ---------------------------
    colopriming_data = pd.read_excel(f"{DATA_DIR}/Sitelist Dec 2025.xlsx")
    gdf_hub = admin_information(gdf_hub, level="kabkot")

    target_crs = 3857
    gdf_points = gdf_points.to_crs(epsg=target_crs)
    gdf_lines = gdf_lines.to_crs(epsg=target_crs)

    gdf_hub = gdf_hub.to_crs(epsg=target_crs)
    gdf_sitelist = gdf_sitelist.to_crs(epsg=target_crs)
    gdf_odp = gdf_odp.to_crs(epsg=target_crs)
    gdf_otb = gdf_otb.to_crs(epsg=target_crs)
    gdf_closure = gdf_closure.to_crs(epsg=target_crs)
    gdf_topology = gdf_topology.to_crs(epsg=target_crs)

    gdf_route = gdf_route.to_crs(epsg=target_crs)
    gdf_backbone = gdf_backbone.to_crs(epsg=target_crs)
    gdf_access = gdf_access.to_crs(epsg=target_crs)
    gdf_fo_exist = gdf_fo_exist.to_crs(epsg=target_crs)
    gdf_pole_exist = gdf_pole_exist.to_crs(epsg=target_crs)
    gdf_obstacle = gdf_obstacle.to_crs(epsg=target_crs)

    # ---------------------------
    # Compile BOQ records
    # ---------------------------
    boq_records: list[dict] = []
    recorded_segment = set()
    num = 1
    gdf_route = gdf_route.reset_index(drop=True)
    for ring_name, gdf_ring_route in gdf_route.groupby("ring_name"):
        gdf_ring_segments = gdf_ring_route.drop_duplicates().copy()
        gdf_ring_sites = gdf_sitelist[gdf_sitelist["ring_name"] == ring_name].copy()
        gdf_ring_hubs = gdf_hub[gdf_hub["ring_name"] == ring_name].copy()

        # Metadata
        program = (gdf_ring_sites["program"].mode().iloc[0] if "program" in gdf_ring_sites.columns and not gdf_ring_sites.empty else None)
        region = (gdf_ring_sites["region"].mode().iloc[0] if "region" in gdf_ring_sites.columns and not gdf_ring_sites.empty else None)
        city = (gdf_ring_hubs["Kabkot"].mode().iloc[0] if "Kabkot" in gdf_ring_hubs.columns and not gdf_ring_hubs.empty else None)
        
        is_otb = DeviceType.OTB in [device_in_site, device_in_branch]
        is_odp = DeviceType.ODP in [device_in_site, device_in_branch]
        is_sc = ConnectorType.SC in [connector_in_branch, connector_in_site]
        is_fc = ConnectorType.FC in [connector_in_branch, connector_in_site]

        # Ring-level slices
        gdf_ring_backbone = gdf_backbone[gdf_backbone["ring_name"] == ring_name].copy()
        gdf_ring_access = gdf_access[gdf_access["ring_name"] == ring_name].copy()
        gdf_ring_fo_exist = gdf_fo_exist[gdf_fo_exist["ring_name"] == ring_name].copy()
        gdf_ring_pole_exist = gdf_pole_exist[gdf_pole_exist["ring_name"] == ring_name].copy()
        gdf_ring_otb = gdf_otb[gdf_otb["ring_name"] == ring_name].copy()
        gdf_ring_odp = gdf_odp[gdf_odp["ring_name"] == ring_name].copy()
        gdf_ring_closure = gdf_closure[gdf_closure["ring_name"] == ring_name].copy()
        gdf_ring_obstacle = gdf_obstacle[gdf_obstacle["ring_name"] == ring_name].copy()

        is_first = True
        for idx, seg_row in gdf_ring_segments.iterrows():
            
            # Segment Metadata
            seg_name = seg_row["name"]
            seg_ne = seg_row["near_end"]
            seg_fe = seg_row["far_end"]
            seg_ctx = f"ring={ring_name} seg={seg_name} ne={seg_ne} fe={seg_fe}"

            site_row = colopriming_data.loc[colopriming_data["site_id"].astype(str) == str(seg_ne)].copy()
            site_row = site_row.iloc[0] if not site_row.empty else pd.Series()
            site_name = site_row.get("site_name", None)
            lat = site_row.get("lat", None)
            long = site_row.get("long", None)
            site_type = site_row.get("site_type", None)
            tower_type = site_row.get("tower_type", None)
            region = site_row.get("region", None)
            kabupaten = site_row.get("kabupaten", None)
            provinsi = site_row.get("provinsi", None)

            mmp_row = gdf_ring_sites[gdf_ring_sites["site_id"].astype(str) == str(seg_fe)].copy()
            mmp_row = mmp_row.to_crs(epsg=4326)
            mmp_row = mmp_row.iloc[0] if not mmp_row.empty else pd.Series()
            mmp_long = mmp_row.get("geometry", None).x
            mmp_lat = mmp_row.get("geometry", None).y

            len_route_m = (
                round(float(seg_row['length']), 3)
                if seg_row.length is not None
                else 0.0
            )
            seg_core = int(seg_row.get("core", 24) or 24)
            
            # Previous Route
            if idx == 0:
                prev_ring = None
                len_prev_access_m = 0
                len_prev_access_ext_m = 0
            else:
                prev_seg = gdf_route.loc[idx-1, :]
                prev_ring = prev_seg['ring_name']
                prev_ne = prev_seg['near_end']
                prev_fe = prev_seg['far_end']
                prev_df_access = gdf_ring_access[(gdf_ring_access["near_end"] == prev_ne) & (gdf_ring_access["far_end"] == prev_fe)].copy()
                len_prev_access_m = (float(sum(prev_df_access['length'])) if not prev_df_access.empty else 0.0)
                len_prev_access_ext_m = 0

            # ---------------------------
            # Segment slices
            # ---------------------------
            df_bb = gdf_ring_backbone[
                (gdf_ring_backbone["near_end"] == seg_ne)
                & (gdf_ring_backbone["far_end"] == seg_fe)
            ].copy()
            df_access = gdf_ring_access[
                (gdf_ring_access["near_end"] == seg_ne)
                & (gdf_ring_access["far_end"] == seg_fe)
            ].copy()
            df_overlap = gdf_ring_fo_exist[
                (gdf_ring_fo_exist["near_end"] == seg_ne)
                & (gdf_ring_fo_exist["far_end"] == seg_fe)
            ].copy()
            df_pole = gdf_ring_pole_exist[
                (gdf_ring_pole_exist["near_end"] == seg_ne)
                & (gdf_ring_pole_exist["far_end"] == seg_fe)
            ].copy()

            df_otb = gdf_ring_otb[gdf_ring_otb["segment"] == seg_name].copy()
            df_otb_new = df_otb[df_otb["ext_note"] == 0].copy()
            df_otb_ext = df_otb[df_otb["ext_note"] == 1].copy()

            df_odp = gdf_ring_odp[gdf_ring_odp["segment"] == seg_name].copy()
            df_odp_new = df_odp[df_odp["ext_note"] == 0].copy()
            df_odp_ext = df_odp[df_odp["ext_note"] == 1].copy()

            if df_odp.empty:
                raise ValueError(
                    f"[ODP_NOT_FOUND] {seg_ctx}\n"
                    f"Segment snapshot:\n"
                    f"{gdf_ring_segments[['segment', 'near_end', 'far_end']].head(10).to_string(index=False)}"
                )

            logger.info(f"🟢 Processing {seg_ctx}")

            df_closure = gdf_ring_closure[gdf_ring_closure["segment"] == seg_name].copy()
            df_closure_new = df_closure[df_closure["ext_note"] == 0].copy()
            df_closure_ext = df_closure[df_closure["ext_note"] == 1].copy()

            df_obs_seg = gdf_ring_obstacle[(gdf_ring_obstacle["near_end"] == seg_ne) & (gdf_ring_obstacle["far_end"] == seg_fe)].copy()
            df_obs_toll = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("toll", case=False, na=False)].copy()
            df_obs_rail = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("rail", case=False, na=False)].copy()
            df_obs_bridge = df_obs_seg[df_obs_seg["obstacle_type"].str.contains("bridge", case=False, na=False)].copy()

            # ---------------------------
            # Core parsing
            # ---------------------------
            core_bb = 24
            if not df_bb.empty and "name" in df_bb.columns:
                raw_name = str(df_bb["name"].iloc[0])
                tail = raw_name.split("_FO")[-1].replace("C", "")
                core_bb = int(tail) if tail.isdigit() else 24

            # ---------------------------
            # Length metrics
            # ---------------------------
            len_bb_m = float(sum(df_bb['length'])) if not df_bb.empty else 0.0
            len_access_m = (float(sum(df_access['length'])) if not df_access.empty else 0.0)
            len_overlap_m = ( float(sum(df_overlap['length'])) if not df_overlap.empty else 0.0)
            len_pole_m = (float(sum(df_pole['length'])) if not df_pole.empty else 0.0)
            len_access_ext_m = 0.0

            # Cable length by backbone core
            len_cable_by_core_m = {c: (len_route_m if core_bb == c else 0.0) for c in (24, 48, 72, 96, 120, 144)}
            
            # ---------------------------
            # Quantity metrics
            # ---------------------------
            qty_otb = len(df_otb)
            qty_otb_new = len(df_otb_new)
            qty_otb_ext = len(df_otb_ext)

            qty_odp = len(df_odp)
            qty_odp_new = len(df_odp_new)
            qty_odp_ext = len(df_odp_ext)

            df_otb_by_core = {
                c: df_otb_new[df_otb_new["core"] == c] for c in (24, 48, 72, 96, 120, 144)
            }
            df_odp_by_core = {
                c: df_odp_new[df_odp_new["core"] == c] for c in (24, 48, 72, 96, 120, 144)
            }
            qty_otb_by_core = {c: len(df_) for c, df_ in df_otb_by_core.items()}
            qty_odp_by_core = {c: len(df_) for c, df_ in df_odp_by_core.items()}

            qty_closure = len(df_closure)
            qty_closure_new = len(df_closure_new)
            qty_closure_ext = len(df_closure_ext)

            qty_obs_toll = len(df_obs_toll)
            qty_obs_rail = len(df_obs_rail)
            qty_obs_bridge = len(df_obs_bridge)

            # ---------------------------
            # Calculations
            # ---------------------------
            fo_factor = 1 + (cable_percentage/100)
            calc_permission_pu = math.floor(len_bb_m + len_access_m - len_pole_m + sum(len_cable_by_core_m.get(core, 0) for core in len_cable_by_core_m.keys() if int(core) != 24))
            calc_fo_cable_m = math.ceil(math.ceil(len_bb_m + len_access_m) * cable_multiplier * fo_factor / 100) * 100
            calc_closure_24_qty = qty_closure_new + (math.floor(calc_fo_cable_m / 4000) if calc_fo_cable_m >= 4000 else 0)
            calc_total_overlap_m = round((len_overlap_m + len_access_ext_m + len_prev_access_m + len_prev_access_ext_m if ring_name == prev_ring else len_overlap_m + len_access_ext_m) * fo_factor, 0)

            # Material
            calc_mat_hdpe_subduct_32_27_qty = 20 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 20 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0) + 70 * qty_obs_rail
            calc_mat_gi_pipe_1p5in_qty = 3 * (qty_otb_by_core.get(24, 0) if (is_sc and is_otb) else 0) + 3 * (qty_otb_by_core.get(24, 0) if (is_fc and is_otb) else 0) + 3 * (2 * qty_obs_rail)
            calc_mat_pole_fo_9m_3step_qty = 0 if ((calc_permission_pu / interval_pole_m) < 3) else even_excel((calc_permission_pu / interval_pole_m) * 0.05) #=IF((S10/80)<3;0;EVEN(((S10)/80)*0,05))
            calc_mat_pole_fo_7m_2step_qty = 0 if calc_permission_pu < 0 else even_excel(calc_permission_pu/interval_pole_m) - calc_mat_pole_fo_9m_3step_qty #=IF((S10)<0;0;EVEN(((S10)/70)-DV10))
            calc_mat_slack_support = math.ceil((calc_fo_cable_m + calc_total_overlap_m)/400) + 1
            calc_mat_slack_support_70x70x3_qty = 1 + math.floor((calc_mat_pole_fo_7m_2step_qty + calc_mat_pole_fo_9m_3step_qty)/4) if calc_mat_pole_fo_7m_2step_qty + calc_mat_pole_fo_9m_3step_qty > 0 else 0 # =IF(SUM(DU10;DV10)>0;1+ROUNDDOWN(SUM(DU10;DV10)/4;0);0)
            
            # Services
            otb_factor = (is_otb and (is_sc or is_fc))
            is_sc_odp = (is_sc and is_odp)

            calc_svc_pulling_fo_aerial_incl_pole_m = (calc_fo_cable_m + (len_cable_by_core_m.get(core_bb, 0) if core_bb != 24 else 0) - len_pole_m - calc_mat_hdpe_subduct_32_27_qty + 0 ) if (calc_fo_cable_m + (len_cable_by_core_m.get(core_bb, 0) if core_bb != 24 else 0) >= 20) else 0
            calc_splicing_fusion = ((calc_closure_24_qty + (qty_odp_by_core.get(24, 0) if is_sc_odp else 0)) * 24 + (qty_odp_by_core.get(48, 0) if is_sc_odp else 0) * 48 + (qty_odp_by_core.get(4, 0)  if is_sc_odp else 0) * 4 + (qty_odp_by_core.get(8, 0)  if is_sc_odp else 0) * 8 + (qty_odp_by_core.get(16, 0) if is_sc_odp else 0) * 16)
            calc_termination_fusion = sum((qty_otb_by_core.get(core, 0) if otb_factor else 0) * core for core in (12, 24, 48, 96, 144, 288))

            calc_svc_splicing_fusion_qty = 24 if (calc_splicing_fusion == 0 and calc_fo_cable_m > 0) else calc_splicing_fusion
            calc_svc_termination_fusion_qty = calc_termination_fusion
            
            # Testing
            calc_test_otdr_2lambda_2way_ls = (calc_svc_termination_fusion_qty if calc_svc_termination_fusion_qty > 0 else 96 if len_cable_by_core_m.get(96, 0) > 0 else 48 if len_cable_by_core_m.get(48, 0) > 0 else 24)

            seg_record = {
            "spk_site": program_name,
            "boq_sent_date": None,
            "no": num,
            "tp": "TBG",
            "site_id": None,
            "site_name": site_name,
            "lat_site": lat,
            "lon_site": long,
            "site_type": site_type,
            "tower_type": tower_type,
            "region": region,
            "area_city": kabupaten,
            "area_province": provinsi,
            "operator": operator.upper(),
            "operator_site_id": seg_ne,
            "spk_wo": program,
            "e2e_distance_m": len_route_m,
            "prep_general_cost": 1,
            "prep_transport_material_cost": 1,
            "mat_fo_adss_24_m": calc_fo_cable_m,
            "mat_odp_24_sc_upc_qty": qty_odp_by_core.get(24, 0) if is_odp else 0,
            "mat_otb_24_sc_upc_qty": qty_otb_by_core.get(24, 0) if is_otb else 0,
            "mat_closure_dome_24_qty": calc_closure_24_qty,
            "mat_hdpe_subduct_32_27_qty": calc_mat_hdpe_subduct_32_27_qty,
            "mat_pole_fo_7m_2step_qty": calc_mat_pole_fo_7m_2step_qty,
            "mat_pole_fo_9m_3step_qty": calc_mat_pole_fo_9m_3step_qty,
            "mat_pole_accesories": calc_mat_pole_fo_7m_2step_qty,
            "mat_slack_support_qty": calc_mat_slack_support,
            "mat_gi_pipe_3in_qty": 1,
            "mat_patch_outdoor_sm_sc_upc_lc_upc_20m": 2 if (seg_name not in recorded_segment) else 0,
            "mat_patch_outdoor_sm_sc_upc_sc_upc_5m": 2 if (seg_name not in recorded_segment) else 0,
            "mat_adapter_lc_lc_duplex_qty": 0,
            "svc_pulling_fo_aerial_m": calc_fo_cable_m,
            "svc_pulling_fo_duct_m": 15,
            "svc_pulling_fo_inbuilding_m": 0,
            "svc_trenching_reinstate_makadam_m": 15,
            "svc_trenching_reinstate_makadam_out_m": 0,
            "svc_install_hdpe_placement": 15,
            "svc_install_pole_7m_2step_unit": calc_mat_pole_fo_7m_2step_qty,
            "svc_install_pole_9m_3step_unit": calc_mat_pole_fo_9m_3step_qty,
            "svc_install_pole_accesories": calc_mat_pole_fo_7m_2step_qty,
            "svc_install_slack_support_qty": calc_mat_slack_support,
            "svc_install_riser_galvanized_3m_qty": 1,
            "svc_splicing_termination_qty": 24*2,
            "svc_install_odp_otb_qty": qty_odp_by_core.get(24, 0),
            "svc_install_patch_outdoor_sc_lc_20m_qty": 2 if (seg_name not in recorded_segment) else 0,
            "svc_install_patch_outdoor_sc_sc_5m_qty": 2 if (seg_name not in recorded_segment) else 0,
            "svc_support_integration_qty": 1,
            "prep_special_permit_cost": None,
            "sow": "New MMP",
            "remark_site": None,
            "lat_mmp": mmp_lat,
            "lon_mmp": mmp_long,
            "remark_mmp": None,
            "existing_route_flag": len_overlap_m,
            }

            boq_records.append(seg_record)

            # Update Segment Info
            recorded_segment.add(seg_name)
            is_first = False
            num += 1

    boq_df = pd.DataFrame(boq_records)

    # ---------------------------
    # Write into Excel template
    # ---------------------------
    template_path = os.path.join(DATA_DIR, "template", "boq", "Template_BOQ_MMP.xlsx")
    output_path = os.path.join(export_dir, "BOQ MMP Report.xlsx")

    if not os.path.exists(template_path):
        raise ValueError("BOQ template file not found in template directory.")

    shutil.copy2(template_path, output_path)

    # ---------------------------
    # PROCESS BOQ EXCEL
    # ---------------------------
    wb = load_workbook(output_path)

    # Style
    named_style = NamedStyle(name="BOQ Row")
    side_style = Side(style="thin", border_style="thin")
    border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
    named_style.font = Font(name="Arial", size=10)
    named_style.border = border

    if "BOQ Row" not in [s for s in wb.named_styles]:
        wb.add_named_style(named_style)

    # BOQ Sheet
    boq_sheet = wb["BOQ"]
    start_data_row = 2
    col_index = {col: num for num, col in enumerate(boq_df.columns, start=1)}
    for idx, record in enumerate(boq_df.to_dict("records")):
        excel_row = start_data_row + idx

        for key, value in record.items():
            cell = boq_sheet.cell(
                row = excel_row,
                column = col_index[key],
                value = value
            )
            cell.style = "BOQ Row"

            if isinstance(value, (int, float)) and value is not None:
                if ('long' in str(key).lower()) or ('lat' in str(key).lower()):
                    continue

                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    logger.info("✅ BOQ Excel saved.")



def main_boq(
    points: gpd.GeoDataFrame,
    lines: gpd.GeoDataFrame,
    export_dir: str,
    sep: str = ";",
    operator= Operator.XL,
    boq_type: BoQType = BoQType.INTERSITE,
    interval_pole_m: int = 80,
    cable_percentage: int = 10,
    cable_multiplier: int = 1,
    sclc_enabled: bool = False,
    device_in_branch = DeviceType.ODP,
    device_in_site = DeviceType.OTB,
    connector_in_site: ConnectorType = ConnectorType.SC,
    connector_in_branch: ConnectorType = ConnectorType.SC,
    program_name: str = "Intersite FO",
    **kwargs,
):
    vendor = kwargs.get("vendor", "TBG")
    program = kwargs.get("program", "Not Defined")
    task_celery = kwargs.get("task_celery", False)

    boq_dir = os.path.join(export_dir, "BOQ")
    os.makedirs(boq_dir, exist_ok=True)

    start_time = time.time()
    points_boq, lines_boq = parallel_boq(
        points, lines, sep=sep, operator=operator, task_celery=task_celery
    )

    # EXPORT
    save_boq(points_boq, lines_boq, boq_dir, sep=sep)
    end_time = time.time()
    boq_time = round((end_time - start_time) / 60, 2)

    # EXCEL FILE
    # excel_boq(
    #     points_boq,
    #     lines_boq,
    #     boq_dir,
    #     sep=sep,
    #     operator=operator,
    #     device_in_branch=device_in_branch,
    #     device_in_site=device_in_site,
    # )

    # KMZ
    start_time = time.time()
    ring_names = sorted(points_boq["ring_name"].dropna().unique().tolist())
    output_kmz = os.path.join(boq_dir, "BOQ KMZ Design.kmz")

    result_boq = compile_boq(
        points_boq,
        lines_boq,
        sep=sep,
        device_in_site=device_in_site,
        device_in_branch=device_in_branch,
    )
    (
        odp,
        otb,
        closure,
        backbone,
        access_ne,
        access_fe,
        fo_exist,
        pole_exist,
        obstacle_railway,
        obstacle_toll,
    ) = result_boq

    main_kmz = simplekml.Kml()
    for num, ring in tqdm(
        enumerate(ring_names, start=1), total=len(ring_names), desc="Process KMZ BOQ"
    ):
        ring_points = points_boq[points_boq["ring_name"] == ring].copy()
        ring_lines = lines_boq[lines_boq["ring_name"] == ring].copy()

        # DATA BOQ
        ring_odp = odp[odp["ring_name"] == ring].copy()
        ring_otb = otb[otb["ring_name"] == ring].copy()
        ring_closure = closure[closure["ring_name"] == ring].copy()
        ring_backbone = backbone[backbone["ring_name"] == ring].copy()
        ring_access_fe = access_fe[access_fe["ring_name"] == ring].copy()
        ring_access_ne = access_ne[access_ne["ring_name"] == ring].copy()
        ring_fo_exist = fo_exist[fo_exist["ring_name"] == ring].copy()
        ring_pole_exist = pole_exist[pole_exist["ring_name"] == ring].copy()
        ring_obstacle_railway = obstacle_railway[
            obstacle_railway["ring_name"] == ring
        ].copy()
        ring_obstacle_toll = obstacle_toll[obstacle_toll["ring_name"] == ring].copy()
        boq_data = (
            ring_odp,
            ring_otb,
            ring_closure,
            ring_backbone,
            ring_access_ne,
            ring_access_fe,
            ring_fo_exist,
            ring_pole_exist,
            ring_obstacle_railway,
            ring_obstacle_toll,
        )

        if "region" in ring_points.columns:
            region = ring_points["region"].mode()[0]
            folder = f"{region}/{ring}"
        else:
            folder = ring

        try:
            main_kmz = kmz_boq(
                main_kmz,
                lines_boq=ring_lines,
                points_boq=ring_points,
                boq_data=boq_data,
                folder=folder,
                vendor=vendor,
                program=program,
                sep=sep,
            )
            if task_celery:
                task_celery.update_state(
                    state="PROGRESS",
                    meta={
                        "status": (f"Compile KMz for {num}/{len(ring_names):,} rings")
                    },
                )
            logger.info(f"🟢 {ring} BOQ KMZ inserted.")
        except Exception as e:
            logger.error(f"Error in ring {ring}: {e}")

    sanitize_kml(main_kmz)
    main_kmz.savekmz(output_kmz)
    end_time = time.time()
    kmz_time = round((end_time - start_time) / 60, 2)

    # BOQ FORMAT RESULT
    start_time = time.time()
    match boq_type:
        case BoQType.INTERSITE:
            boq_generation(
                kmz_path=output_kmz, 
                export_dir=boq_dir, 
                sep=sep, 
                operator=operator,  
                interval_pole_m = interval_pole_m,
                cable_percentage = cable_percentage,
                sclc_enabled = sclc_enabled,
                device_in_site = device_in_site,
                device_in_branch = device_in_branch,
                connector_in_site = connector_in_site,
                connector_in_branch = connector_in_branch,
                program_name = program_name
            )
        case BoQType.MMP:
            boq_mmp(
                kmz_path=output_kmz, 
                export_dir=boq_dir, 
                sep=sep, 
                operator=operator,  
                interval_pole_m = interval_pole_m,
                cable_percentage = cable_percentage,
                cable_multiplier = cable_multiplier,
                sclc_enabled = sclc_enabled,
                device_in_site = device_in_site,
                device_in_branch = device_in_branch,
                connector_in_site = connector_in_site,
                connector_in_branch = connector_in_branch,
                program_name = program_name
            )
    end_time = time.time()
    excel_time = round((end_time - start_time) / 60, 2)

    logger.info(f"✅ All BOQ Process Done.")
    logger.info(f"ℹ️ Time Consumed:")
    logger.info(f"BOQ Parallel Time   : {boq_time:,} minutes")
    logger.info(f"Excel Result Time   : {excel_time:,} minutes")
    logger.info(f"KMZ Result Time     : {kmz_time:,} minutes")
    logger.info(f"BOQ {operator.upper()} Time : {round((end_time-start_time)/60,2):,} minutes")


if __name__ == "__main__":
    kmz_path = r"D:\JACOBS\PROJECT\TASK\2026\FEB\W1\MMP BOQ\MMP XLS Batch 6 - SOKKA.kmz"
    export_dir = (r"D:\JACOBS\PROJECT\TASK\2026\FEB\W1\MMP BOQ\MMP XLS Batch 6 - SOKKA")
    sep= ";"
    boq_type = "mmp"
    operator = "tsel"

    match boq_type:
        case BoQType.INTERSITE:
            # INTERSITE CONFIG
            interval_pole_m = 80
            cable_percentage = 10
            cable_multiplier = 1
            device_in_branch = "ODP"
            device_in_site = "OTB"
            sclc_enabled = False

        case BoQType.MMP:
            # MMP CONFIG
            interval_pole_m = 60
            cable_percentage = 15
            cable_multiplier = 2
            device_in_branch = "ODP"
            device_in_site = "ODP"
            sclc_enabled = False

    os.makedirs(export_dir, exist_ok=True)
    points_kmz, lines_kmz = validate_kmz_design(kmz_path, sep=sep)
    main_boq(
        points=points_kmz,
        lines=lines_kmz,
        export_dir=export_dir,
        sep=sep,
        operator=operator,
        boq_type=boq_type,
        interval_pole_m=interval_pole_m,
        cable_percentage=cable_percentage,
        cable_multiplier=cable_multiplier,
        sclc_enabled = sclc_enabled,
        device_in_site=device_in_site,
        device_in_branch=device_in_branch,
        program_name="MMP DMT"
    )

    # ZIPFILE
    zip_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_BOQ_Task.zip"
    zip_filepath = os.path.join(export_dir, zip_filename)
    with zipfile.ZipFile(zip_filepath, "w", zipfile.ZIP_DEFLATED) as zipf:
        for root, _, files in os.walk(export_dir):
            for export_file in files:
                if (
                    export_file != zip_filename
                    and not export_file.endswith(".zip")
                    and "Checkpoint" not in export_file
                ):
                    export_file_path = os.path.join(root, export_file)
                    arcname = os.path.relpath(export_file_path, export_dir)
                    zipf.write(export_file_path, arcname)
    print(f"📦 Result files zipped.")

    # kmz_ipl = r"D:\JACOBS\PROJECT\TASK\2026\JAN\W4\BOQ Dev\0000005199_APD IPL Fiberisasi Newsite TBG_2025.kmz"
    # export_dir = (r"D:\JACOBS\PROJECT\TASK\2026\JAN\W4\BOQ Dev\\Export\0000005199_APD IPL Fiberisasi Newsite TBG_2025")
    # os.makedirs(export_dir, exist_ok=True)

    # boq_generation(kmz_ipl, export_dir=export_dir, sep=";", operator="xl")